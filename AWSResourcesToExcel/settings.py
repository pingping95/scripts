from openpyxl.styles import Font, Side, Border, Alignment, Color, PatternFill


class Common:
    def __init__(self, name: str):
        """
        Set resource common initial settings
        """
        self.sheet = None
        self.cell_start = 2
        self.name = name
        self.thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),)
    
    def add_cell(self, ro: int, co: int, values: str):
        """
        Add each cell
        """
        self.sheet.cell(row=ro, column=co).value = values
        self.sheet.cell(row=ro, column=co).font = Font(name="맑은 고딕", size=10)
        self.sheet.cell(row=ro, column=co).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        self.sheet.cell(row=ro, column=co).border = self.thin_border
    
    def return_humanbytes(self, byte):
        """
        Return the given bytes as a human friendly KiB, MiB, GiB or TiB Strings
        """
        B = float(byte)
        KiB = float(1024)
        MiB = float(KiB ** 2)  # 1,048,576
        GiB = float(KiB ** 3)  # 1,073,741,824
        TiB = float(KiB ** 4)  # 1,099,511,627,776
        if B < KiB:
            return "{0} {1}".format(B, "Bytes" if 0 == B > 1 else "Byte")
        elif KiB <= B < MiB:
            return "{0:.1f} KiB".format(B / KiB)
        elif MiB <= B < GiB:
            return "{0:.1f} MiB".format(B / MiB)
        elif GiB <= B < TiB:
            return "{0:.1f} GiB".format(B / GiB)
        elif TiB <= B:
            return "{0:.1f} TiB".format(B / TiB)
    
    def fit_cell_width(self, widths: list):
        dimensions = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", 'Q', 'R', 'S', 'T', 'U', 'V']
        for k in range(0, 22):
            self.sheet.column_dimensions[dimensions[k]].width = widths[k]

    def check_type(self, ip_type: str, number: int):
        """
        Transport port number into protocol type
        """
        transported_type = ip_type  # tcp, upd, icmp, icmpv6 or numbers
        target_dict = {22: "SSH", 25: "SMTP", 53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP", 389: "LDAP", 443: "HTTPS",
                    445: "SMB", 465: "SMTPS", 993: "IMAPS", 995: "POP3S", 1433: "MSSQL", 2049: "NFS", 3306: "MySQL/Aurora",
                    3389: "RDP", 5439: "Redshift", 5432: "PostgreSQL", 1521: "Oracle-RDS", 5985: "WirnRM-HTTP",
                    5986: "WinRM-HTTPS", 2007: "Elastic-Graphics"}
        if type(number) == int:
            if number in target_dict:
                transported_type = target_dict.get(number)
            elif type(number) == "tcp" or type(number) == "udp":
                transported_type = "Custom {0} Rule".format(ip_type.upper())
            return transported_type
        else:
            if "-" in str(number):
                transported_type = "Custom {0} Rule".format(ip_type.upper())
        return transported_type
    
    def make_header(self, start_row: int, title: str):
        """
        Make title of sheet.
        """
        self.sheet.cell(row = start_row, column = 2).value = title
        self.sheet.cell(row=start_row, column=2).font = Font(name="맑은 고딕", size=12, bold=True)
        self.cell_start += 1
    
    def make_cell_header(self, cell_start, head):
        """
        Make headers of sheet
        """
        for colu_index, hea in enumerate(head):
            self.sheet.cell(row=cell_start, column=colu_index + 2).value = hea
            self.sheet.cell(row=cell_start, column=colu_index + 2).font = Font(name="맑은 고딕", size=10, bold=True)
            self.sheet.cell(row=cell_start, column=colu_index + 2).alignment = Alignment(horizontal="center", vertical="center")
            self.sheet.cell(row=cell_start, column=colu_index + 2).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
            self.sheet.cell(row=cell_start, column=colu_index + 2).border = self.thin_border
        self.cell_start += 1