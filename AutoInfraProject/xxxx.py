import boto3
import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment, Color, PatternFill
import datetime
from dateutil.relativedelta import relativedelta
from time import sleep

# Profile Name와 Region 기입 필수


my_name = "taehun.kim"
cred_file_path = f'C:\\Users\\{my_name}\\.aws\\config'


storing_path = f'G:\\공유 드라이브\\CSU 공유 드라이브\\팀별 공유 드라이브\\DevOps 상품 본부\\SRE2 센터\\SRE 5팀\\SMB3\\[기타] 내부공유자료\\고객사 자산 내역\\'

# Open the file, Only read access
f = open(cred_file_path, 'r')

lines = f.readlines()

pr

p_name_list = []
r_name_list = []



for line in lines:
    if "profile" in line:
        p_name = line.split(" ")[-1].split("]")[0]
        p_name_list.append(p_name)
    if "region" in line:
        r_name = line.split(" = ")[-1]
        if "\n" in r_name:
            r_name = r_name.split("\n")[0]
        r_name_list.append(r_name)

f.close()        

cust_name_list = dict(zip(p_name_list, r_name_list))



# Settings

right_now = datetime.datetime.now()
now = right_now.date()
target_date = str(now + relativedelta(days=-2))

today = datetime.date.today()

# dd/mm/YY
today_date = today.strftime("%Y%m%d")

# cell box
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),)


# custom 함수 생성
def add_cell(sheet, ro, col, values):
    sheet.cell(row=ro, column=col).value = values
    sheet.cell(row=ro, column=col).font = Font(name="맑은 고딕", size=10)
    sheet.cell(row=ro, column=col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    sheet.cell(row=ro, column=col).border = thin_border


def humanbytes(B):
    """
    Return the given bytes as a human friendly KiB, MiB, GiB, or TiB string
    """
    B = float(B)
    KiB = float(1024)
    MiB = float(KiB ** 2)  # 1,048,576
    GiB = float(KiB ** 3)  # 1,073,741,824
    TiB = float(KiB ** 4)  # 1,099,511,627,776
    if B < KiB:
        return "{0} {1}".format(B, "Bytes" if 0 == B > 1 else "Byte")
    elif KiB <= B < MiB:
        return "{0:.1f} KiB".format(B / KiB)
    elif MiB <= B < GiB:
        return "{0:.1f} MiB".format(B / MiB)
    elif GiB <= B < TiB:
        return "{0:.1f} GiB".format(B / GiB)
    elif TiB <= B:
        return "{0:.1f} TiB".format(B / TiB)


# cell width
def sheet_cell_width(sheet, sheet_cell_widths):
    dimensions = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", 
                "N", "O", "P", 'Q', 'R', 'S', 'T', 'U', 'V']
    for k in range(0, 22):
        sheet.column_dimensions[dimensions[k]].width = sheet_cell_widths[k]


# Transport Port Number into Type ( Dictionary )
def check_type(ip_type, number):
    """
    Transport Port Number into Type
    ipType(str), number(int)
    """
    transported_type = ip_type  # tcp, upd, icmp, icmpv6 or numbers
    target_dict = {22: "SSH", 25: "SMTP", 53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP", 389: "LDAP", 443: "HTTPS",
                445: "SMB", 465: "SMTPS", 993: "IMAPS", 995: "POP3S", 1433: "MSSQL", 2049: "NFS", 3306: "MySQL/Aurora",
                3389: "RDP", 5439: "Redshift", 5432: "PostgreSQL", 1521: "Oracle-RDS", 5985: "WirnRM-HTTP",
                5986: "WinRM-HTTPS", 2007: "Elastic-Graphics"}

    if type(number) == int:
        if number in target_dict:
            transported_type = target_dict.get(number)
        elif type(number) == "tcp" or type(number) == "udp":
            transported_type = "Custom {0} Rule".format(ip_type.upper())
        return transported_type
    else:
        if "-" in str(number):
            transported_type = "Custom {0} Rule".format(ip_type.upper())

    return transported_type


def make_header(sheet, start_row, title):
    sheet.cell(row=start_row, column=2).value = title
    sheet.cell(row=start_row, column=2).font = Font(name="맑은 고딕", size=12, bold=True)


def make_cell_header(sheet, cell_start, head):
    for colu_index, hea in enumerate(head):
        sheet.cell(row=cell_start, column=colu_index + 2).value = hea
        sheet.cell(row=cell_start, column=colu_index + 2).font = Font(name="맑은 고딕", size=10, bold=True)
        sheet.cell(row=cell_start, column=colu_index + 2).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=cell_start, column=colu_index + 2).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
        sheet.cell(row=cell_start, column=colu_index + 2).border = thin_border


for p_name, r_name in cust_name_list.items():
    
    try:

        # Profile Name의 Credentials 정보를 이용하여 Session 맺음
        session = boto3.session.Session(profile_name=p_name)

        # Session을 이용하여 Resource or Client 객체 생성
        ec2_res = session.resource(service_name="ec2", region_name=r_name)
        ec2_cli = session.client(service_name="ec2", region_name=r_name)
        elb_cli = session.client(service_name="elb", region_name=r_name)  # CLB
        elbv2_cli = session.client(service_name="elbv2", region_name=r_name)  # ALB, NLB, GLB
        rds_cli = session.client("rds", region_name=r_name)
        s3_cli = session.client(service_name="s3", region_name=r_name)
        s3_res = session.resource(service_name="s3", region_name=r_name)
        cloudfront_cli = session.client(service_name="cloudfront", region_name=r_name)
        cloudtrail_cli = session.client(service_name="cloudtrail", region_name=r_name)
        cloudwatch_cli = session.client(service_name="cloudwatch", region_name=r_name)
        lambda_cli = session.client(service_name="lambda", region_name=r_name)
        iam_cli = session.client(service_name="iam", region_name=r_name)
        iam_res = session.resource(service_name="iam", region_name=r_name)
        route53_cli = session.client(service_name="route53", region_name=r_name)
        eks_cli = session.client(service_name="eks", region_name=r_name)
        ecs_cli = session.client(service_name="ecs", region_name=r_name)
        els_cli = session.client(service_name="elasticache", region_name=r_name)
        efs_cli = session.client(service_name="efs", region_name=r_name)
        ecr_cli = session.client(service_name="ecr", region_name=r_name)
        dynamo_cli = session.client(service_name="dynamodb", region_name=r_name)
        es_cli = session.client(service_name="es", region_name=r_name)



        # Workbook 생성
        wb = openpyxl.Workbook()
        print("###############################################################################################")
        print(f"profile name : {p_name}\n region name : {r_name}\n(고객 환경마다 엑셀 생성 시간 상이)\n")

        """
        sheet 1
        """
        sheet1 = wb.active
        sheet1.title = "VPC"
        # 10
        cell_widths = [5, 5, 18, 25, 25, 25, 17, 15, 20, 20, 5, 5, 18, 24, 24, 13, 13, 13, 13, 13, 7, 7]
        sheet_cell_width(sheet1, cell_widths)

        # VPCs
        sheet1_cell_start = 2
        make_header(sheet1, sheet1_cell_start, "VPC")
        sheet1_cell_start += 1

        cell_headers = ["No.", "VPC Name", "VPC ID", "VPC CIDR Block"]
        make_cell_header(sheet1, sheet1_cell_start, cell_headers)
        sheet1_cell_start += 1


        for idx, vpc in enumerate(ec2_res.vpcs.all()):
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=2, values=idx + 1)
            try:
                for tag in vpc.tags:
                    if tag["Key"] == "Name":
                        add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values=tag["Value"])
            except Exception as e:
                add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values="-")
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=4, values=vpc.id)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=5, values=vpc.cidr_block)

            sheet1_cell_start = sheet1_cell_start + 1


        # Network ACLS

        sheet1_cell_start = sheet1_cell_start + 2
        make_header(sheet1, sheet1_cell_start, "NACL")
        sheet1_cell_start += 1

        # 1. Inbound
        acl_start_row = sheet1_cell_start

        cell_header_upper = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Inbound Rule", "", "", "",  ""]

        make_cell_header(sheet1, sheet1_cell_start, cell_header_upper)
        sheet1_cell_start += 1

        acl_end_row = sheet1_cell_start

        cell_header_lower = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Rule", "Protocol", "Port Range", "Source","Allow / Deny",]

        make_cell_header(sheet1, sheet1_cell_start, cell_header_lower)
        sheet1_cell_start += 1

        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=2, end_column=2)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=3, end_column=3)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=4, end_column=4)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=5, end_column=5)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_start_row, start_column=6, end_column=10)


        # 2. OutBound
        sheet1_cell_start = acl_start_row

        acl_start_row = sheet1_cell_start

        cell_header_upper = ["No.","Network ACL Name","Network ACL Id","VPC ID","Outbound Rule","","","",""]
        for col_index, header in enumerate(cell_header_upper):
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).value = header
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).font = Font(name="맑은 고딕", size=10, bold=True)
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).alignment = Alignment(horizontal="center", vertical="center")
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).border = thin_border
        sheet1_cell_start = sheet1_cell_start + 1

        acl_end_row = sheet1_cell_start

        cell_header_lower = ["No.","Network ACL Name","Network ACL Id","VPC ID","Rule","Protocol","Port Range","Source","Allow / Deny"]

        for col_index, header in enumerate(cell_header_lower):
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).value = header
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).font = Font(name="맑은 고딕", size=10, bold=True)
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).alignment = Alignment(horizontal="center", vertical="center")
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
            sheet1.cell(row=sheet1_cell_start, column=col_index + 12).border = thin_border
        sheet1_cell_start = sheet1_cell_start + 1

        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=12, end_column=12)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=13, end_column=13)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=14, end_column=14)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=15, end_column=15)
        sheet1.merge_cells(start_row=acl_start_row, end_row=acl_start_row, start_column=16, end_column=20)


        # 값 입력
        # Inbound Rule

        init_sheet1_cell_start = sheet1_cell_start

        for idx, acl in enumerate(ec2_cli.describe_network_acls()["NetworkAcls"]):

            acl_start_row = sheet1_cell_start

            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=2, values=idx + 1)  # Number
            name = "-"
            try:
                for tag in acl["Tags"]:
                    if tag["Key"] == "Name":
                        name = tag["Value"]
            except Exception:
                pass

            acl_id = acl["NetworkAclId"]
            vpc_id = acl["VpcId"]

            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values=name)  # Network ACL Name
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=4, values=acl_id)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=5, values=vpc_id)

            in_count = 0
            for entry in acl["Entries"]:
                if entry["Egress"] == False:  # Ingress
                    in_count += 1

                    if "CidrBlock" in entry:
                        source = entry["CidrBlock"]
                    else:
                        source = "-"
                    protocol = "All"
                    portRange = "All"

                    # ruleNumber
                    if entry["RuleNumber"] >= 32737:
                        ruleNumber = "*"
                    else:
                        ruleNumber = entry["RuleNumber"]

                    # RuleAction
                    ruleAction = entry["RuleAction"]

                    # protocol, PortRange
                    if entry["Protocol"] == "-1":
                        protocol = "All"
                    elif entry["Protocol"] == "6":
                        protocol = "TCP"
                        if entry["PortRange"]["To"] == entry["PortRange"]["From"]:
                            portRange = entry["PortRange"]["From"]
                        else:
                            portRange = (
                                str(entry["PortRange"]["From"]) + " - " + str(entry["PortRange"]["To"])
                            )
                    elif entry["Protocol"] == "17":
                        protocol = "UDP"
                        if entry["PortRange"]["To"] == entry["PortRange"]["From"]:
                            portRange = entry["PortRange"]["From"]
                        else:
                            portRange = (
                                str(entry["PortRange"]["From"]) + " - " + str(entry["PortRange"]["To"])
                            )

                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=6, values=ruleNumber)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=7, values=protocol)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=8, values=portRange)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=9, values=source)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=10, values=ruleAction.capitalize())

                    sheet1_cell_start += 1

            acl_end_row = sheet1_cell_start - 1

            # merge cells
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=2, end_column=2)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=3, end_column=3)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=4, end_column=4)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=5, end_column=5)

        subnet_sheet1_cell_start = sheet1_cell_start

        # Outbound
        sheet1_cell_start = init_sheet1_cell_start

        for idx, acl in enumerate(ec2_cli.describe_network_acls()["NetworkAcls"]):
            acl_start_row = sheet1_cell_start

            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=12, values=idx + 1)  # Number
            name = "-"
            try:
                for tag in acl["Tags"]:
                    if tag["Key"] == "Name":
                        name = tag["Value"]
            except Exception:
                pass

            acl_id = acl["NetworkAclId"]
            vpc_id = acl["VpcId"]

            # Network ACL Name
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=13, values=name)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=14, values=acl_id)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=15, values=vpc_id)

            in_count = 0
            for entry in acl["Entries"]:
                if entry["Egress"] == True:  # Egress
                    in_count += 1

                    if "CidrBlock" in entry:
                        source = entry["CidrBlock"]
                    else:
                        source = "-"
                    protocol = "All"
                    portRange = "All"

                    # ruleNumber
                    if entry["RuleNumber"] >= 32737:
                        ruleNumber = "*"
                    else:
                        ruleNumber = entry["RuleNumber"]

                    # RuleAction
                    ruleAction = entry["RuleAction"]

                    # protocol, PortRange
                    if entry["Protocol"] == "-1":
                        protocol = "All"
                    elif entry["Protocol"] == "6":
                        protocol = "TCP"
                        if entry["PortRange"]["To"] == entry["PortRange"]["From"]:
                            portRange = entry["PortRange"]["From"]
                        else:
                            portRange = (str(entry["PortRange"]["From"]) + " - " + str(entry["PortRange"]["To"]))
                    elif entry["Protocol"] == "17":
                        protocol = "UDP"
                        if entry["PortRange"]["To"] == entry["PortRange"]["From"]:
                            portRange = entry["PortRange"]["From"]
                        else:
                            portRange = (str(entry["PortRange"]["From"]) + " - " + str(entry["PortRange"]["To"]))

                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=16, values=ruleNumber)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=17, values=protocol)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=18, values=portRange)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=19, values=source)
                    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=20, values=ruleAction.capitalize())

                    sheet1_cell_start += 1

            acl_end_row = sheet1_cell_start - 1

            # merge cells
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=12, end_column=12)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=13, end_column=13)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=14, end_column=14)
            sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=15, end_column=15)

        # Subnets

        sheet1_cell_start = subnet_sheet1_cell_start + 2
        make_header(sheet1, sheet1_cell_start, "Subnet Group")
        sheet1_cell_start += 1

        cell_headers = ["No.", "Subnet Type", "VPC ID", "Subnet Name", "Subnet ID", "Subnet CIDR Block", "Availability Zone",
                        "Network ACLs", "Route Tables"]
        make_cell_header(sheet1, sheet1_cell_start, cell_headers)
        sheet1_cell_start += 1

        # Public Subnet을 pubSubnetDict에 넣음
        pubSubnetDict = {}

        for routeTable in ec2_cli.describe_route_tables()["RouteTables"]:
            associations = routeTable["Associations"]
            routes = routeTable["Routes"]
            isPublic = False

            for route in routes:
                gid = route.get("GatewayId", "")
                if gid.startswith("igw-"):
                    isPublic = True

            if not isPublic:
                continue

            for assoc in associations:
                subnetId = assoc.get("SubnetId", None)  # This checks for explicit associations, only
                if subnetId:
                    pubSubnetDict[subnetId] = isPublic


        for idx, subnet in enumerate(ec2_res.subnets.all()):
            add_cell(sheet1, sheet1_cell_start, 2, idx + 1)  # Number

            # Subnet Type (Check if this subnet is Public or Private)
            if subnet.id in pubSubnetDict:
                add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values="Public Network")
            else:
                add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values="Private Network")

            add_cell(sheet1, sheet1_cell_start, 4, subnet.vpc_id)  # VPC Id

            # Subnet Name
            try:
                for tags in subnet.tags:
                    if tags["Key"] == "Name":
                        add_cell(sheet1, sheet1_cell_start, 5, tags["Value"])
            except Exception as e:
                add_cell(sheet1, sheet1_cell_start, 5, "-")

            add_cell(sheet1, sheet1_cell_start, 6, subnet.subnet_id)  # Subnet Id
            add_cell(sheet1, sheet1_cell_start, 7, subnet.cidr_block)  # Subnet CIDR Block
            add_cell(sheet1, sheet1_cell_start, 8, subnet.availability_zone)  # Subnet Availability Zone

            # NetworkAcls
            for acls in ec2_cli.describe_network_acls()["NetworkAcls"]:
                for i in acls["Associations"]:
                    if i.get("SubnetId") == subnet.subnet_id:
                        add_cell(sheet1, sheet1_cell_start, 9, i.get("NetworkAclId"))

            # Route Tables
            try:
                route_table = ec2_cli.describe_route_tables(
                    Filters=[
                        {
                            "Name": "association.subnet-id",
                            "Values": [
                                subnet.subnet_id,
                            ],
                        },
                    ],
                )["RouteTables"]
                add_cell(sheet1, sheet1_cell_start, 10, route_table[0].get("Associations")[0].get("RouteTableId"))
            except:
                add_cell(sheet1, sheet1_cell_start, 10, "-")

            sheet1_cell_start = sheet1_cell_start + 1

        """
        Route Table
        """

        sheet3 = wb.create_sheet("sheet3")
        sheet3.title = "Route Table"

        cell_widths = [5, 5, 25, 23, 22, 25, 16, 20, 22, 15, 20, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7]
        sheet_cell_width(sheet3, cell_widths)

        sheet3_cell_start = 3
        make_header(sheet3, sheet3_cell_start, "Route Table")
        sheet3_cell_start += 1

        cell_headers = ["No.", "Name", "ID", "Destination", "Target"]
        make_cell_header(sheet3, sheet3_cell_start, cell_headers)
        sheet3_cell_start += 1

        for idx, route in enumerate(ec2_res.route_tables.all()):
            long_start = sheet3_cell_start
            short_start = sheet3_cell_start
            route_table = ec2_res.RouteTable(route.id)
            route_asso = route_table.associations_attribute

            add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

            try:
                add_cell(sheet3, sheet3_cell_start, 3, route_table.tags[0].get("Value"))
            except:
                add_cell(sheet3, sheet3_cell_start, 3, "-")

            add_cell(sheet3, sheet3_cell_start, 4, route_table.id)

            for idx, info in enumerate(route_table.routes_attribute):
                add_cell(sheet3, long_start, 5, info.get("DestinationCidrBlock"))
                if info.get("GatewayId"):
                    add_cell(sheet3, long_start, 6, info.get("GatewayId"))
                elif info.get("NatGatewayId"):
                    add_cell(sheet3, long_start, 6, info.get("NatGatewayId"))
                elif info.get("TransitGatewayId"):
                    add_cell(sheet3, long_start, 6, info.get("TransitGatewayId"))
                elif info.get("LocalGatewayId"):
                    add_cell(sheet3, long_start, 6, info.get("LocalGatewayId"))
                else:
                    add_cell(sheet3, long_start, 6, info.get("NetworkInterfaceId"))
                long_start = long_start + 1

            # cell merge
            sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=2, end_column=2)
            sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=3, end_column=3)
            sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=4, end_column=4)

            sheet3_cell_start = long_start


        """
        Peering Connection
        """

        response = ec2_cli.describe_vpc_peering_connections()["VpcPeeringConnections"]
        if len(response) != 0:

            sheet3_cell_start = sheet3_cell_start + 2
            sheet3.cell(row=sheet3_cell_start, column=2).value = "Peering Connections"
            sheet3.cell(row=sheet3_cell_start, column=2).font = Font(name="맑은 고딕", size=12, bold=True)
            sheet3_cell_start = sheet3_cell_start + 1

            cell_headers = ["No.","Name","Status","Peering Connection ID","Requester VPC ID","Requester Region",
            "Requester CIDR Block","Accepter VPC ID","Accepter Region","Accepter CIDR Block"]
            
            make_cell_header(sheet3, sheet3_cell_start, cell_headers)
            
            sheet3_cell_start = sheet3_cell_start + 1

            for idx, peering in enumerate(response):
                # No.
                add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

                # Name
                peer_name = "-"
                try:
                    for i in peering.get("Tags"):
                        if i.get("Key") == "Name":
                            peer_name = i.get("Value")
                except Exception:
                    pass
                add_cell(sheet3, sheet3_cell_start, 3, peer_name)

                # Status, Accepter CIDR Block
                # Statue가 Active가 아닐 경우 Accepter의 CIDR Block을 알아낼 수 없음
                peer_status = peering["Status"].get("Code")
                if peer_status == "active":
                    add_cell(sheet3, sheet3_cell_start, 4, peer_status)

                    # Accepter CIDR Block
                    accepter_cidr = peering["AccepterVpcInfo"].get("CidrBlock")
                    add_cell(sheet3, sheet3_cell_start, 11, accepter_cidr)

                else:
                    add_cell(sheet3, sheet3_cell_start, 4, peer_status)
                    add_cell(sheet3, sheet3_cell_start, 11, "-")

                # Peering Connection ID
                peer_id = peering.get("VpcPeeringConnectionId")
                add_cell(sheet3, sheet3_cell_start, 5, peer_id)

                # Requester
                requester = peering.get("RequesterVpcInfo")

                # Requester VPC ID
                req_id = requester.get("VpcId")
                add_cell(sheet3, sheet3_cell_start, 6, req_id)

                # Requester Region
                req_region = requester.get("Region")
                add_cell(sheet3, sheet3_cell_start, 7, req_region)

                # Requester CIDR Block
                req_cidr = requester.get("CidrBlock")
                add_cell(sheet3, sheet3_cell_start, 8, req_cidr)

                # Accepter
                accepter = peering.get("AccepterVpcInfo")

                # Accepter VPC ID
                accep_id = accepter.get("VpcId")
                add_cell(sheet3, sheet3_cell_start, 9, accep_id)

                # Accepter Region
                accep_region = accepter.get("Region")
                add_cell(sheet3, sheet3_cell_start, 10, accep_region)

                sheet3_cell_start = sheet3_cell_start + 1

        else:
            print("There is no Peering Connections")

        """
        Transit Gateway
        """

        response = ec2_cli.describe_transit_gateways()["TransitGateways"]
        if len(response) != 0:

            sheet3_cell_start = sheet3_cell_start + 2
            make_header(sheet3, sheet3_cell_start, "Transit Gateway")
            sheet3_cell_start = sheet3_cell_start + 1

            cell_headers = ["No.", "Name", "State", "Transit Gateway ID", "Creation Date", "CIDR Blocks"]
            make_cell_header(sheet3, sheet3_cell_start, cell_headers)
            sheet3_cell_start = sheet3_cell_start + 1

            for idx, tgw in enumerate(response):
                # No.
                add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

                # Name
                tgw_name = "-"
                try:
                    for i in tgw.get("Tags"):
                        if i.get("Key") == "Name":
                            tgw_name = i.get("Value")
                except Exception:
                    pass
                add_cell(sheet3, sheet3_cell_start, 3, tgw_name)

                # State
                tgw_state = tgw.get("State")
                add_cell(sheet3, sheet3_cell_start, 4, tgw_state)

                # Transit Gateway ID
                tgw_id = tgw.get("TransitGatewayId")
                add_cell(sheet3, sheet3_cell_start, 5, tgw_id)

                # Creation Date
                tgw_created_date = str(tgw.get("CreationTime"))
                tgw_date = tgw_created_date.split(" ")[0]
                add_cell(sheet3, sheet3_cell_start, 6, tgw_date)

                # CIDR Blocks
                try:
                    tgw_cidr = ""
                    for i, cidr in enumerate(tgw["Options"]["TransitGatewayCidrBlocks"]):
                        if i != 0:
                            tgw_cidr += ", \n"
                        tgw_cidr += cidr
                    add_cell(sheet3, sheet3_cell_start, 7, tgw_cidr)
                except Exception as e:
                    add_cell(sheet3, sheet3_cell_start, 7, "-")

                sheet3_cell_start = sheet3_cell_start + 1

        else:
            print("There is no Transit Gateway")


        # sheet 4
        # Nat GW, IGW

        """
        Nat Gateway
        """

        sheet4 = wb.create_sheet("sheet4")
        sheet4.title = "NAT GW & IGW"

        cell_widths = [5, 5, 27, 22, 22, 22, 24, 20, 20, 20, 20, 20, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7]
        sheet_cell_width(sheet4, cell_widths)

        title = sheet4["B2"]
        title.value = "Nat Gateways"
        title.font = Font(name="맑은 고딕", size=12, bold=True)
        # title.alignment = Alignment(horizontal='center', vertical='center')


        sheet4_cell_start = 3
        cell_header = ["No.", "Name", "ID", "Elastic IP", "VPC", "Subnet", "Status"]

        for col_index, header in enumerate(cell_header):
            sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
            sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(
                name="맑은 고딕", size=10, bold=True
            )
            sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(
                patternType="solid", fgColor=Color("E3E3E3")
            )
            sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border

        sheet4_cell_start = sheet4_cell_start + 1

        """
        Nat Gateway
        """

        if ec2_cli.describe_nat_gateways()["NatGateways"]:

            for idx, nat in enumerate(ec2_cli.describe_nat_gateways()["NatGateways"]):
                add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
                natName = "-"
                try:
                    for i in nat.get("Tags"):
                        if i.get("Key") == "Name":
                            natName = i.get("Value")
                except Exception:
                    pass

                id = nat.get("NatGatewayId")
                eip = nat.get("NatGatewayAddresses")[0].get("PublicIp")
                vpc = nat.get("VpcId")
                subnet = nat.get("SubnetId")
                state = nat.get("State")

                add_cell(sheet4, sheet4_cell_start, 3, natName)
                add_cell(sheet4, sheet4_cell_start, 4, id)
                add_cell(sheet4, sheet4_cell_start, 5, eip)
                add_cell(sheet4, sheet4_cell_start, 6, vpc)
                add_cell(sheet4, sheet4_cell_start, 7, subnet)
                add_cell(sheet4, sheet4_cell_start, 8, state.capitalize())

                sheet4_cell_start = sheet4_cell_start + 1
        else:
            print("There is no NAT Gateway.")


        """
        IGW
        """
        if ec2_cli.describe_internet_gateways():

            sheet4_cell_start = sheet4_cell_start + 2

            sheet4.cell(row=sheet4_cell_start, column=2).value = "Internet Gateways"
            sheet4.cell(row=sheet4_cell_start, column=2).font = Font(name="맑은 고딕", size=12, bold=True)

            sheet4_cell_start = sheet4_cell_start + 1

            cell_headers = ["No.", "Name", "ID", "VPC", "Status"]
            make_cell_header(sheet4, sheet4_cell_start, cell_headers)

            sheet4_cell_start = sheet4_cell_start + 1

            for idx, igw in enumerate(ec2_cli.describe_internet_gateways()["InternetGateways"]):
                # No.
                add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
                # Name 가져옴, 없을 시 '-'
                igwName = "-"
                try:
                    for i in igw.get("Tags"):
                        if i.get("Key") == "Name":
                            igwName = i.get("Value")
                except Exception:
                    pass
                # igw ID, vpc ID, Status
                igwId = igw.get("InternetGatewayId")
                if len(igw["Attachments"]) != 0:
                    vpcId = igw["Attachments"][0].get("VpcId")
                    status = igw["Attachments"][0].get("State")
                else:
                    vpcId = "-"
                    status = "Detached"

                # Name
                add_cell(sheet4, sheet4_cell_start, 3, igwName)
                # IGW ID
                add_cell(sheet4, sheet4_cell_start, 4, igwId)
                # VPC ID
                add_cell(sheet4, sheet4_cell_start, 5, vpcId)
                # Status
                add_cell(sheet4, sheet4_cell_start, 6, status.capitalize())

                sheet4_cell_start = sheet4_cell_start + 1

        else:
            print("There is no Internet Gateway.")


        """
        VPN Connections
        """
        response = ec2_cli.describe_vpn_connections()["VpnConnections"]


        if len(response) != 0:
            sheet4_cell_start += 2
            make_header(sheet4, sheet4_cell_start, "VPN")

            sheet4_cell_start += 1
            cell_headers = ["No.","Name","VPN ID","State","Virtual Private Gateway","Transit Gateway",
                            "Customer Gateway", "Routing", "Type", "Local IPv4 CIDR", "Remote IPv4 CIDR"]
            make_cell_header(sheet4, sheet4_cell_start, cell_headers)
            sheet4_cell_start += 1

            for idx, vpn in enumerate(response):
                # No.
                add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
                # Name
                vpn_name = "-"
                try:
                    for i in vpn.get("Tags"):
                        if i.get("Key") == "Name":
                            vpn_name = i.get("Value")
                except Exception:
                    pass
                add_cell(sheet4, sheet4_cell_start, 3, vpn_name)

                # VPN ID
                vpn_id = vpn.get("VpnConnectionId")
                add_cell(sheet4, sheet4_cell_start, 4, vpn_id)

                # State
                vpn_state = vpn.get("State")
                add_cell(sheet4, sheet4_cell_start, 5, vpn_state.capitalize())

                # Virtual Private Gateway ID
                try:
                    vgw_id = vpn.get("VpnGatewayId")
                    add_cell(sheet4, sheet4_cell_start, 6, vgw_id)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 6, "-")

                # Transit Gateway ID
                try:
                    tgw_id = vpn.get("TransitGatewayId")
                    add_cell(sheet4, sheet4_cell_start, 7, tgw_id)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 7, "-")

                # Customer Gateway ID
                try:
                    cgw_id = vpn.get("CustomerGatewayId")
                    add_cell(sheet4, sheet4_cell_start, 8, cgw_id)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 8, "-")

                # Routing
                routing = vpn["Options"]["StaticRoutesOnly"]
                if routing == False:
                    add_cell(sheet4, sheet4_cell_start, 9, "Dynamic")
                else:
                    add_cell(sheet4, sheet4_cell_start, 9, "Static")

                # Type
                vpn_type = vpn.get("Type")
                add_cell(sheet4, sheet4_cell_start, 10, vpn_type)

                # Local IPv4 CIDR
                local_cidr = vpn["Options"]["LocalIpv4NetworkCidr"]
                add_cell(sheet4, sheet4_cell_start, 11, local_cidr)

                # Remote IPV4 CIDR
                remote_cidr = vpn["Options"]["RemoteIpv4NetworkCidr"]
                add_cell(sheet4, sheet4_cell_start, 12, remote_cidr)

                sheet4_cell_start = sheet4_cell_start + 1
        else:
            print("There is no VPN Connections.")


        """
        Virtual Private Gateway
        """

        response = ec2_cli.describe_vpn_gateways()["VpnGateways"]


        if len(response) != 0:

            sheet4_cell_start = sheet4_cell_start + 2
            make_header(sheet4, sheet4_cell_start, "VGW")
            sheet4_cell_start += 1

            cell_headers = ["No.", "Name", "ID", "State", "Type", "VPC", "Amazon Side SAN"]
            make_cell_header(sheet4, sheet4_cell_start, cell_headers)
            sheet4_cell_start += 1

            for idx, vgw in enumerate(response):
                # No.
                add_cell(sheet4, sheet4_cell_start, 2, idx + 1)

                # Name
                vgw_name = "-"
                try:
                    for i in vgw.get("Tags"):
                        if i.get("Key") == "Name":
                            vgw_name = i.get("Value")
                    add_cell(sheet4, sheet4_cell_start, 3, vgw_name)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 3, vgw_name)

                # ID
                vgw_id = vgw.get("VpnGatewayId")
                add_cell(sheet4, sheet4_cell_start, 4, vgw_id)

                # State
                vgw_state = vgw.get("State")
                add_cell(sheet4, sheet4_cell_start, 5, str(vgw_state).capitalize())

                # Type
                vgw_type = vgw.get("Type")
                add_cell(sheet4, sheet4_cell_start, 6, vgw_type)

                # VPC
                vpc_id = ""
                try:
                    for i in vgw.get("VpcAttachments"):
                        if i.get("State") == "attached":
                            vpc_id += i.get("VpcId")
                    add_cell(sheet4, sheet4_cell_start, 7, vpc_id)
                except Exception:
                    print("fail")
                    add_cell(sheet4, sheet4_cell_start, 7, vpc_id)

                # Amazon Side SAN
                try:
                    vgw_san = vgw.get("AmazonSideAsn")
                    add_cell(sheet4, sheet4_cell_start, 8, vgw_san)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 8, "-")

                sheet4_cell_start = sheet4_cell_start + 1

        else:
            print("There is no Virtual Private Gateways.")


        """
        Customer Gateway
        """

        response = ec2_cli.describe_customer_gateways()["CustomerGateways"]
        if len(response) != 0:
            sheet4_cell_start = sheet4_cell_start + 2
            make_header(sheet4, sheet4_cell_start, "CGW")
            sheet4_cell_start = sheet4_cell_start + 1

            cell_headers = ["No.", "Name", "ID", "State", "Type", "IP Address", "BGP ASN", "Device Name"]
            make_cell_header(sheet4, sheet4_cell_start, cell_headers)
            sheet4_cell_start = sheet4_cell_start + 1

            for idx, cgw in enumerate(response):

                # No.
                add_cell(sheet4, sheet4_cell_start, 2, idx + 1)

                # Name
                cgw_name = "-"
                try:
                    for i in cgw.get("Tags"):
                        if i.get("Key") == "Name":
                            cgw_name = i.get("Value")
                    add_cell(sheet4, sheet4_cell_start, 3, cgw_name)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 3, cgw_name)

                # CGW ID
                cgw_id = cgw.get("CustomerGatewayId")
                add_cell(sheet4, sheet4_cell_start, 4, cgw_id)

                # State
                cgw_state = cgw.get("State")
                add_cell(sheet4, sheet4_cell_start, 5, str(cgw_state).capitalize())

                # Type
                cgw_type = cgw.get("Type")
                add_cell(sheet4, sheet4_cell_start, 6, cgw_type)

                # IP Address
                cgw_ip = cgw.get("IpAddress")
                add_cell(sheet4, sheet4_cell_start, 7, cgw_id)

                # BGP SAN
                bgw_asn = cgw.get("BgpAsn")
                add_cell(sheet4, sheet4_cell_start, 8, bgw_asn)

                # Device Name
                try:
                    device_name = cgw["DeviceName"]
                    add_cell(sheet4, sheet4_cell_start, 9, device_name)
                except Exception:
                    add_cell(sheet4, sheet4_cell_start, 9, "-")

                sheet4_cell_start = sheet4_cell_start + 1


        else:
            print("There is no Customet Gateways.")


        # sheet5

        sheet5 = wb.create_sheet("sheet5")
        sheet5.title = "Security Group"
        cell_widths = [5, 5, 55, 20, 22, 13, 24, 45, 7, 7.8, 55, 23, 22, 13, 20, 30, 10, 11, 7, 7, 7, 7]
        sheet_cell_width(sheet5, cell_widths)

        sheet5_cell_start = 2
        make_header(sheet5, sheet5_cell_start, "Inbound")
        sheet5_cell_start += 1

        # Security Group
        cell_header1 = ["No.","Security Groups Name","Group ID","Inbound Rule","Inbound Rule","Inbound Rule","비고(Description)"]
        cell_header2 = [ "No.", "Security Groups Name", "Group ID", "Type", "Port Range","source","비고(Description)"]

        make_cell_header(sheet5, sheet5_cell_start, cell_header1)
        sheet5_cell_start = sheet5_cell_start + 1

        make_cell_header(sheet5, sheet5_cell_start, cell_header2)
        sheet5_cell_start = sheet5_cell_start + 1

        sheet5.merge_cells(start_row=3, end_row=4, start_column=2, end_column=2)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=3, end_column=3)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=4, end_column=4)
        sheet5.merge_cells(start_row=3, end_row=3, start_column=5, end_column=7)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=8, end_column=8)

        previous = 4
        flag = 0

        # Inbound Rule
        for idx, security_group in enumerate(ec2_res.security_groups.all()):
            short_start = sheet5_cell_start
            long_start1 = sheet5_cell_start
            short_start1 = sheet5_cell_start
            long_start2 = sheet5_cell_start
            short_start2 = sheet5_cell_start
            sec_group = ec2_res.SecurityGroup(security_group.id)

            # No.
            add_cell(sheet5, sheet5_cell_start, 2, idx + 1)
            # Security Group Name
            add_cell(sheet5, sheet5_cell_start, 3, sec_group.group_name)
            # Security Group Id
            add_cell(sheet5, sheet5_cell_start, 4, sec_group.group_id)

            # inbound
            # Type, Port Range, Source, Description
            for inbound in sec_group.ip_permissions:
                ip_range = []
                desc = []

                # Type
                if inbound.get("IpProtocol") == "-1":
                    ipType = "ALL Traffic"
                else:
                    ipType = inbound.get("IpProtocol")

                # portRange
                if inbound.get("FromPort") == inbound.get("ToPort"):
                    portrange = inbound.get("FromPort")
                else:
                    portrange = str(inbound.get("FromPort")) + " - " + str(inbound.get("ToPort"))

                if portrange == "0--1" or portrange == -1:
                    portrange = "N/A"

                if ipType == "ALL Traffic":
                    portrange = "All"

                ipType = check_type(ipType, portrange)

                # source
                for ips in inbound.get("IpRanges"):
                    ip_range.append(ips.get("CidrIp"))
                    description = ips.get("Description")
                    if description == None:
                        desc.append("-")
                    else:
                        desc.append(description)

                # 비고
                for ips in inbound.get("Ipv6Ranges"):
                    ip_range.append(ips.get("CidrIpv6"))
                    description = ips.get("Description")
                    if description == None:
                        desc.append("-")
                    else:
                        desc.append(description)

                for group in inbound.get("UserIdGroupPairs"):
                    ip_range.append(group.get("GroupId"))
                    description = group.get("Description")
                    if description == None:
                        desc.append("-")
                    else:
                        desc.append(description)

                add_cell(sheet5, short_start1, 5, ipType)
                add_cell(sheet5, short_start1, 6, portrange)

                tmp1 = long_start1

                for ip, desc in zip(ip_range, desc):
                    add_cell(sheet5, long_start1, 7, ip)
                    add_cell(sheet5, long_start1, 8, desc)
                    long_start1 += 1

                tmp1_1 = long_start1

                sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=5, end_column=5)
                sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=6, end_column=6)

                short_start1 = long_start1

            try:
                if long_start1 >= long_start2:
                    sheet5_cell_start = long_start1
                    # cell merge
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=2, end_column=2)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=3, end_column=3)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=4, end_column=4)
                else:
                    sheet5_cell_start = long_start2
                    # cell merge

                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=2, end_column=2)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=3, end_column=3)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=4, end_column=4)
            except:
                pass

        ##

        title = sheet5["J2"]
        title.value = "Outbound"
        title.font = Font(name="맑은 고딕", size=12, bold=True)

        sheet5_cell_start = 3
        cell_header1 = ["No.","Security Groups Name","Group ID","Outbound Rule","Outbound Rule","Outbound Rule","비고(Description)"]
        cell_header2 = ["No.","Security Groups Name","Group ID","Type","Port Range","source","비고(Description)"]

        for col_index, header in enumerate(cell_header1):
            add_cell(sheet5, sheet5_cell_start, col_index + 10, header)
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).font = Font(name="맑은 고딕", size=10, bold=True)
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).alignment = Alignment(horizontal="center", vertical="center")
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).border = thin_border
        sheet5_cell_start = sheet5_cell_start + 1

        for col_index, header in enumerate(cell_header2):
            add_cell(sheet5, sheet5_cell_start, col_index + 10, header)
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).font = Font(name="맑은 고딕", size=10, bold=True)
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).alignment = Alignment(horizontal="center", vertical="center")
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
            sheet5.cell(row=sheet5_cell_start, column=col_index + 10).border = thin_border
        sheet5_cell_start = sheet5_cell_start + 1

        sheet5.merge_cells(start_row=3, end_row=4, start_column=10, end_column=10)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=11, end_column=11)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=12, end_column=12)
        sheet5.merge_cells(start_row=3, end_row=3, start_column=13, end_column=15)
        sheet5.merge_cells(start_row=3, end_row=4, start_column=16, end_column=16)

        # outbound

        for idx, security_group in enumerate(ec2_res.security_groups.all()):
            add_cell(sheet5, sheet5_cell_start, 10, idx + 1)
            short_start = sheet5_cell_start
            long_start1 = sheet5_cell_start
            short_start1 = sheet5_cell_start
            long_start2 = sheet5_cell_start
            short_start2 = sheet5_cell_start

            sec_group = ec2_res.SecurityGroup(security_group.id)
            add_cell(sheet5, sheet5_cell_start, 11, sec_group.group_name)
            add_cell(sheet5, sheet5_cell_start, 12, sec_group.group_id)

            # inbound
            for outbound in sec_group.ip_permissions_egress:
                if outbound:
                    ip_range = []
                    desc = []

                    # Type
                    if outbound.get("IpProtocol") == "-1":
                        ipType = "ALL Traffic"
                    else:
                        ipType = outbound.get("IpProtocol")

                    # portRange
                    if outbound.get("FromPort") == outbound.get("ToPort"):
                        portrange = outbound.get("FromPort")
                    else:
                        portrange = str(outbound.get("FromPort")) + " - " + str(outbound.get("ToPort"))

                    if portrange == "0--1" or portrange == -1:
                        portrange = "N/A"

                    if ipType == "ALL Traffic":
                        portrange = "All"

                    ipType = check_type(ipType, portrange)

                    # source
                    for ips in outbound.get("IpRanges"):
                        ip_range.append(ips.get("CidrIp"))
                        description = ips.get("Description")
                        if description == None:
                            desc.append("-")
                        else:
                            desc.append(description)

                    # 비고
                    for ips in outbound.get("Ipv6Ranges"):
                        ip_range.append(ips.get("CidrIpv6"))
                        description = ips.get("Description")
                        if description == None:
                            desc.append("-")
                        else:
                            desc.append(description)

                    for group in outbound.get("UserIdGroupPairs"):
                        ip_range.append(group.get("GroupId"))
                        description = group.get("Description")
                        if description == None:
                            desc.append("-")
                        else:
                            desc.append(description)

                    add_cell(sheet5, short_start1, 10, idx + 1)
                    add_cell(sheet5, short_start1, 11, sec_group.group_name)
                    add_cell(sheet5, short_start1, 12, sec_group.group_id)

                    add_cell(sheet5, short_start1, 13, ipType)
                    add_cell(sheet5, short_start1, 14, portrange)

                    tmp1 = long_start1

                    for ip, desc in zip(ip_range, desc):
                        add_cell(sheet5, long_start1, 15, ip)
                        add_cell(sheet5, long_start1, 16, desc)
                        add_cell(sheet5, long_start1, 10, idx + 1)
                        add_cell(sheet5, long_start1, 11, sec_group.group_name)
                        add_cell(sheet5, long_start1, 12, sec_group.group_id)
                        long_start1 += 1

                    tmp1_1 = long_start1
                    try:
                        sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=13, end_column=13)
                        sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=14, end_column=14)
                    except:
                        pass
                    short_start1 = long_start1

                sheet5_cell_start = long_start1

            try:
                if long_start1 >= long_start2:
                    sheet5_cell_start = long_start1
                    # cell merge
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=10, end_column=10)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=11, end_column=11)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=12, end_column=12)
                else:
                    sheet5_cell_start = long_start2
                    # cell mrge

                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=10, end_column=10)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=11, end_column=11)
                    sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=12, end_column=12)
            except:
                pass


        # sheet6
        # ELB
        response = elb_cli.describe_load_balancers().get("LoadBalancerDescriptions")
        response2 = elbv2_cli.describe_load_balancers().get('LoadBalancers')
        if len(response) != 0 or len(response2) != 0:

            sheet6 = wb.create_sheet("sheet6")
            sheet6.title = "ELB"
            cell_widths = [5, 5, 13, 28, 65, 10, 25, 32, 18, 20.5, 12, 14, 11, 7, 7, 7, 7, 7, 7, 7,7,7]
            sheet_cell_width(sheet6, cell_widths)

            sheet6_cell_start = 2
            make_header(sheet6, sheet6_cell_start, "ELB")
            sheet6_cell_start += 1

            # ELB
            cell_headers = ["No.", "Scheme", "ELB Name", "DNS Name", "Type", "Port Configuration", "Instance IDs or Target Groups",
                            "Availability Zones", "ELB Security Group", "Cross-Zone","Idle Timeout (s)", "Access Logs"]
            make_cell_header(sheet6, sheet6_cell_start, cell_headers)
            sheet6_cell_start = sheet6_cell_start + 1


            num = 0

            # CLB
            for idx, response in enumerate(response):
                # No.
                num = idx + 1
                add_cell(sheet6, sheet6_cell_start, 2, idx + 1)

                # Scheme
                schemeType = str(response["Scheme"]).capitalize()
                add_cell(sheet6, sheet6_cell_start, 3, schemeType)

                # ELB Name
                ElbName = response["LoadBalancerName"]
                add_cell(sheet6, sheet6_cell_start, 4, ElbName)

                # DNS Name
                DnsName = response["DNSName"]
                add_cell(sheet6, sheet6_cell_start, 5, DnsName)

                # Type
                add_cell(sheet6, sheet6_cell_start, 6, "Classic")

                # Port Configuration
                ListenerDesc = response["ListenerDescriptions"]
                Listenerstr = ""
                for idx, Listener in enumerate(ListenerDesc):
                    if idx != 0:
                        Listenerstr += ", \n"
                    Listenerstr += (
                            str(Listener["Listener"].get("LoadBalancerPort"))
                            + " forwarding to "
                            + str(Listener["Listener"].get("InstancePort"))
                    )
                add_cell(sheet6, sheet6_cell_start, 7, Listenerstr)

                # Instance IDs
                idstr = ""
                if len(response["Instances"]) != 0:
                    for idx, id in enumerate(response["Instances"]):
                        if idx != 0:
                            idstr += ", \n"
                        idstr += id.get("InstanceId")
                    add_cell(sheet6, sheet6_cell_start, 8, idstr)
                else:
                    add_cell(sheet6, sheet6_cell_start, 8, "-")

                # Availability Zones
                zonestr = ""
                for idx, zone in enumerate(response["AvailabilityZones"]):
                    if idx != 0:
                        zonestr += ", \n"
                    zonestr += zone
                add_cell(sheet6, sheet6_cell_start, 9, zonestr)

                # Security Groups
                secstr = ""
                for idx, sec in enumerate(response["SecurityGroups"]):
                    if idx != 0:
                        secstr += ", \n"
                    secstr += sec
                add_cell(sheet6, sheet6_cell_start, 10, secstr)

                response_attr = elb_cli.describe_load_balancer_attributes(LoadBalancerName=ElbName)[
                    "LoadBalancerAttributes"
                ]

                # Cross-Zone Load Balancing
                crossZone = str(response_attr.get('CrossZoneLoadBalancing').get('Enabled')).lower().capitalize()
                add_cell(sheet6, sheet6_cell_start, 11, crossZone)

                # Idle Timeout
                idleTimeout = response_attr["ConnectionSettings"]["IdleTimeout"]
                add_cell(sheet6, sheet6_cell_start, 12, int(idleTimeout))

                # Access Logs
                access_log = str(response_attr.get('AccessLog').get('Enabled')).lower().capitalize()
                add_cell(sheet6, sheet6_cell_start, 13, access_log)

                sheet6_cell_start += 1

            # ALB, NLB, GWLB
            for idx, response in enumerate(response2):
                # No.
                num += 1
                add_cell(sheet6, sheet6_cell_start, 2, num)

                # Scheme
                schemeType = ""
                if "Scheme" in response:
                    schemeType = str(response["Scheme"]).capitalize()
                if len(schemeType) != 0:
                    add_cell(sheet6, sheet6_cell_start, 3, schemeType)
                else:
                    add_cell(sheet6, sheet6_cell_start, 3, "-")
                
                # ELB Name
                ElbName = response["LoadBalancerName"]
                add_cell(sheet6, sheet6_cell_start, 4, ElbName)

                # DNS Name
                DnsName = ""
                if "DNSName" in response:
                    DnsName = response["DNSName"]
                if len(DnsName) != 0:
                    add_cell(sheet6, sheet6_cell_start, 5, DnsName)
                else:
                    add_cell(sheet6, sheet6_cell_start, 5, "-")

                # Type
                lbType = str(response["Type"])
                add_cell(sheet6, sheet6_cell_start, 6, lbType.capitalize())

                # Port Configuration
                # Instance IDs
                target_group = elbv2_cli.describe_target_groups(
                    LoadBalancerArn=response["LoadBalancerArn"]
                    )
                
                portstr = ""
                targetstr = ""
                try:
                    for idx, port in enumerate(target_group["TargetGroups"]):
                        if idx != 0:
                            portstr += ", \n"
                            targetstr += ", \n"
                        portstr += str(port.get("Port")) + " " + port.get("Protocol")
                        targetstr += port.get("TargetGroupName")
                    add_cell(sheet6, sheet6_cell_start, 7, portstr)
                    add_cell(sheet6, sheet6_cell_start, 8, targetstr)

                except:
                    add_cell(sheet6, sheet6_cell_start, 7, "-")
                    add_cell(sheet6, sheet6_cell_start, 8, "-")

                # Availability Zones
                zonestr = ""
                for idx, zone in enumerate(response["AvailabilityZones"]):
                    if idx != 0:
                        zonestr += ", \n"
                    zonestr += zone.get("ZoneName")
                add_cell(sheet6, sheet6_cell_start, 9, zonestr)

                # Security Groups
                secstr = ""
                try:
                    for idx, sec in enumerate(response["SecurityGroups"]):
                        if idx != 0:
                            secstr += ", \n"
                        secstr += sec
                        add_cell(sheet6, sheet6_cell_start, 10, secstr)
                except:
                    add_cell(sheet6, sheet6_cell_start, 10, "-")

                # Attributes
                Attributes = elbv2_cli.describe_load_balancer_attributes(LoadBalancerArn=response["LoadBalancerArn"])["Attributes"]

                # GWLB : CorssZone
                if response["Type"] == "gateway":
                    accessLogsEnabled = "-"
                    crossZoneLBEnabled = Attributes[1]["Value"]
                    idleTimeout = "-"

                # NLB : AccessLogs, CrossZone
                elif response["Type"] == "network":
                    accessLogsEnabled = Attributes[0]["Value"]
                    crossZoneLBEnabled = Attributes[4]["Value"]
                    idleTimeout = "-"

                # ALB : AccessLogs, CrossZone, Idle Timeout
                else:
                    accessLogsEnabled = Attributes[0]["Value"]
                    crossZoneLBEnabled = "true"  # 기본으로 활성화되어 있음
                    idleTimeout = Attributes[3]["Value"]
                
                # Cross Zone
                add_cell(sheet6, sheet6_cell_start, 11, str(crossZoneLBEnabled).capitalize())

                # IdleTimeout
                add_cell(sheet6, sheet6_cell_start, 12, idleTimeout)

                # Access Logs
                try:
                    add_cell(sheet6, sheet6_cell_start, 13, str(accessLogsEnabled).lower().capitalize())
                except:
                    add_cell(sheet6, sheet6_cell_start, 13, "-")

                sheet6_cell_start += 1

        else:
            print("There is no ELB.")

        # Elastic IP 리스트 뽑아냄
        addr_list = ec2_cli.describe_addresses().get("Addresses")
        eip_list = []
        try:
            for eip in addr_list:
                eip_list.append(eip.get("PublicIp"))
        except Exception:
            pass

        response = ec2_cli.describe_instances()
        if response["Reservations"]:

            # sheet7
            sheet7 = wb.create_sheet("sheet7")
            sheet7.title = "EC2"
            cell_widths = [5, 5, 14, 35, 8, 8, 12, 11, 20, 23, 19, 13, 13, 13, 21, 15, 20, 38, 15, 21, 15, 10]
            sheet_cell_width(sheet7, cell_widths)

            sheet7_cell_start = 2
            make_header(sheet7, sheet7_cell_start, "EC2")
            sheet7_cell_start += 1

            """ EC2
            """
            cell_headers = ["No.", "Availability Zone", "Instance Name", "Status", "AMI (OS)", "Instance Type", "Subnet Type",
                            "VPC ID", "Subnet ID", "Instance ID", "Private IP", "Public IP", "Elstic IP", "Root Volume ID",
                            "Root Volume (GB)", "Key Pair", "Security Groups", "IAM role", "Data Volume ID", "Data Volume (GB)"]
            make_cell_header(sheet7, sheet7_cell_start, cell_headers)

            sheet7_cell_start += 1

            indx = 0
            for reservation in response["Reservations"]:
                for Instances in reservation["Instances"]:
                    indx += 1

                    # No.
                    add_cell(sheet7, sheet7_cell_start, 2, indx)

                    # Availability Zone
                    Zone = Instances["Placement"].get("AvailabilityZone")
                    add_cell(sheet7, sheet7_cell_start, 3, Zone)

                    # Instance Name
                    InstanceName = "-"
                    try:
                        for i in Instances["Tags"]:
                            if i.get("Key") == "Name":
                                InstanceName = i.get("Value")
                                break
                    except:
                        pass
                    add_cell(sheet7, sheet7_cell_start, 4, InstanceName)

                    # Status
                    # 'pending'|'running'|'shutting-down'|'terminated'|'stopping'|'stopped'
                    Status = Instances["State"].get("Name")
                    add_cell(sheet7, sheet7_cell_start, 5, Status)
                    if Status == "running" or Status == "stopped":

                        # AMI (OS)
                        try:
                            OS = Instances["Platform"]
                            add_cell(sheet7, sheet7_cell_start, 6, OS)
                        except:
                            add_cell(sheet7, sheet7_cell_start, 6, "Linux")
                        # Instance Type
                        Type = Instances["InstanceType"]
                        add_cell(sheet7, sheet7_cell_start, 7, Type)

                        # Subnet Type
                        if Instances.get("PublicIpAddress"):
                            add_cell(sheet7, sheet7_cell_start, 8, "Public")
                        else:
                            add_cell(sheet7, sheet7_cell_start, 8, "Private")

                        # VPC ID
                        vpc_id = Instances["VpcId"]
                        add_cell(sheet7, sheet7_cell_start, 9, vpc_id)

                        # Subnet ID
                        Subnet_ID = Instances["SubnetId"]
                        add_cell(sheet7, sheet7_cell_start, 10, Subnet_ID)

                        # Instance ID
                        Instance_ID = Instances["InstanceId"]
                        add_cell(sheet7, sheet7_cell_start, 11, Instance_ID)

                        # Private IP
                        Private_IP = Instances["PrivateIpAddress"]
                        add_cell(sheet7, sheet7_cell_start, 12, Private_IP)

                        # Public IP
                        if Instances.get("PublicIpAddress"):
                            Public_IP = Instances["PublicIpAddress"]
                            add_cell(sheet7, sheet7_cell_start, 13, Public_IP)
                        else:
                            Public_IP = ""
                            add_cell(sheet7, sheet7_cell_start, 13, Public_IP)

                        # EIP
                        if Public_IP in eip_list:
                            add_cell(sheet7, sheet7_cell_start, 14, Public_IP)
                        else:
                            add_cell(sheet7, sheet7_cell_start, 14, "-")

                        # data volume 리스트
                        data_volume_list = []

                        # Root Volume ID
                        root_volume_id = ""
                        for ebs in Instances["BlockDeviceMappings"]:
                            data_volume_list.append(ebs["Ebs"].get("VolumeId"))  # EC2 인스턴스에 있는 모든 volume을 추가

                            if ebs.get("DeviceName") == Instances["RootDeviceName"]:
                                data_volume_list.remove(ebs["Ebs"].get("VolumeId"))  # Root Volume은 리스트에서 제외
                                root_volume_id = str(ebs["Ebs"].get("VolumeId"))
                                add_cell(sheet7, sheet7_cell_start, 15, root_volume_id)

                        # Root Volume (GB)
                        try:
                            response = ec2_cli.describe_volumes(VolumeIds=[root_volume_id])
                            for volume in response["Volumes"]:
                                Size = volume.get("Size")
                                add_cell(sheet7, sheet7_cell_start, 16, Size)

                        # Root Volume이 없을 경우 오류 발생 가능성 있음
                        except Exception:
                            add_cell(sheet7, sheet7_cell_start, 16, "-")

                        # Key Pair
                        try:
                            KeyPair = Instances["KeyName"]
                            add_cell(sheet7, sheet7_cell_start, 17, KeyPair)
                        except:
                            pass
                        # Security Group
                        sec_groups = ""
                        for idx, sec in enumerate(Instances["SecurityGroups"]):
                            if idx != 0:
                                sec_groups += ", \n"
                            sec_groups += sec.get("GroupName")
                            # SG = sec.get('GroupName')
                        add_cell(sheet7, sheet7_cell_start, 18, sec_groups)
                        # IAM role
                        try:
                            IAM = Instances["IamInstanceProfile"].get("Arn").split("/")[-1]
                            add_cell(sheet7, sheet7_cell_start, 19, IAM)
                        except:
                            add_cell(sheet7, sheet7_cell_start, 19, "-")

                        # Data Volume이 존재할 경우
                        if len(data_volume_list) != 0:
                            # Data Volume ID
                            data_volume_id = ""
                            data_volume_size = ""
                            for idx, data_volume in enumerate(data_volume_list):
                                if idx != 0:
                                    data_volume_id += ", \n"
                                    data_volume_size += ", \n"
                                data_volume_id += data_volume
                                response2 = ec2_cli.describe_volumes(VolumeIds=[data_volume])
                                for volume in response2["Volumes"]:
                                    data_volume_size += str(volume.get("Size"))

                            add_cell(sheet7, sheet7_cell_start, 20, data_volume_id)
                            add_cell(sheet7, sheet7_cell_start, 21, data_volume_size)
                        else:
                            add_cell(sheet7, sheet7_cell_start, 20, "-")
                            add_cell(sheet7, sheet7_cell_start, 21, "-")

                        end_sheet7_row = sheet7_cell_start
                    else:
                        pass

                    sheet7_cell_start = sheet7_cell_start + 1

        else:
            print("There is no EC2.")

        # sheet8
        rdsres = rds_cli.describe_db_instances()

        if rdsres["DBInstances"]:

            sheet8 = wb.create_sheet("sheet8")
            sheet8.title = "RDS"
            cell_widths = [5, 5, 14, 35, 11, 13, 22, 15, 15, 21, 45, 49, 23, 23, 12, 27,22, 20, 7, 7, 7, 7]
            sheet_cell_width(sheet8, cell_widths)
            sheet8_cell_start = 2
            make_header(sheet8, sheet8_cell_start, "RDS")
            sheet8_cell_start += 1

            # RDS
            cell_headers = ["No.", "Availability Zone", "RDS Name", "RDS Engine", "Engine Version", "DB Instance Class", "Storage Type",
                            "Master Username", "Master Password", "VPC ID", "Subnet Group", "Parameter Group", "Option Group",
                            "Database Port","Preferred Maintenance Window", "Preferred Backup Window", "Backup Retention Time"]
            make_cell_header(sheet8, sheet8_cell_start, cell_headers)

            sheet8_cell_start += 1

            for idx, rdsdata in enumerate(rdsres["DBInstances"]):
                # pprint.pprint(rdsdata)
                # No.
                add_cell(sheet8, sheet8_cell_start, 2, idx + 1)
                # Availability Zone
                add_cell(sheet8, sheet8_cell_start, 3, rdsdata["AvailabilityZone"])
                # RDS Name
                add_cell(sheet8, sheet8_cell_start, 4, rdsdata["DBInstanceIdentifier"])
                # DB Engine
                add_cell(sheet8, sheet8_cell_start, 5, rdsdata["Engine"])
                # Engine Version
                add_cell(sheet8, sheet8_cell_start, 6, rdsdata["EngineVersion"])
                # DB Instance Class
                add_cell(sheet8, sheet8_cell_start, 7, rdsdata["DBInstanceClass"])
                # Storage Type
                add_cell(sheet8, sheet8_cell_start, 8, rdsdata["StorageType"])
                # Master Username
                add_cell(sheet8, sheet8_cell_start, 9, rdsdata["MasterUsername"])
                # Master Password
                add_cell(sheet8, sheet8_cell_start, 10, "-")
                # VPC ID
                add_cell(sheet8, sheet8_cell_start, 11, rdsdata["DBSubnetGroup"].get("VpcId"))
                # Subnet Group

                subnetstr = rdsdata["DBSubnetGroup"].get("DBSubnetGroupName")
                subnetstr += "\n( "
                for idx, subnet in enumerate(rdsdata["DBSubnetGroup"].get("Subnets")):
                    if idx != 0:
                        subnetstr += ", \n"
                    subnetstr += subnet.get("SubnetIdentifier")
                subnetstr += " )"
                add_cell(sheet8, sheet8_cell_start, 12, subnetstr)

                # Parameter Group
                for dbparam in rdsdata["DBParameterGroups"]:
                    add_cell(sheet8, sheet8_cell_start, 13, dbparam.get("DBParameterGroupName"))
                # Option Group
                for dboption in rdsdata["OptionGroupMemberships"]:
                    add_cell(sheet8, sheet8_cell_start, 14, dboption.get("OptionGroupName"))
                # Database Port
                add_cell(sheet8, sheet8_cell_start, 15, rdsdata["Endpoint"].get("Port"))

                # Maintenance Time
                add_cell(sheet8, sheet8_cell_start, 16, rdsdata["PreferredMaintenanceWindow"])
                # PreferredMaintenanceWindowPreferredBackupWindow Backup Retention Time
                add_cell(sheet8, sheet8_cell_start, 17, rdsdata["PreferredBackupWindow"])
                add_cell(sheet8, sheet8_cell_start, 18, rdsdata["BackupRetentionPeriod"])
                # 용도
                sheet8_cell_start = sheet8_cell_start + 1
        else:
            print("There is no RDS.")

        # sheet9
        response = s3_cli.list_buckets()

        if response["Buckets"]:
            sheet9 = wb.create_sheet("sheet9")
            sheet9.title = "S3"

            cell_widths = [5, 5, 55, 13, 14, 22, 11, 11, 11, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet9, cell_widths)

            sheet9_cell_start = 2
            make_header(sheet9, sheet9_cell_start, "S3")
            sheet9_cell_start += 1

            # S3
            cell_headers = ["No.", "Bucket Name", "Creation Date", "Region", "Size\n (" + str(target_date) + " 측정값)",
                            "Logging", "Versioning",]
            make_cell_header(sheet9, sheet9_cell_start, cell_headers)
            sheet9_cell_start += 1

            for idx, bucket in enumerate(s3_res.buckets.all()):
                try:
                    # No.
                    add_cell(sheet9, sheet9_cell_start, 2, idx + 1)

                    # Name
                    bucket_name = bucket.name
                    add_cell(sheet9, sheet9_cell_start, 3, bucket_name)

                    # Creation Date
                    bucket_creation_date = str(bucket.creation_date).split(" ")[0]
                    add_cell(sheet9, sheet9_cell_start, 4, bucket_creation_date)

                    # Bucket Region
                    bucket_region = s3_cli.get_bucket_location(
                        Bucket=bucket_name
                    ).get('LocationConstraint')
                    if bucket_region == None:
                        bucket_region = "us-east-1"
                    add_cell(sheet9, sheet9_cell_start, 5, bucket_region)


                    # Bucket Size
                    cloudwatch_temp_cli = session.client(service_name="cloudwatch", region_name=bucket_region)

                    readable_bucket_size = 0
                    response = cloudwatch_temp_cli.get_metric_statistics(
                        Namespace="AWS/S3",
                        MetricName="BucketSizeBytes",
                        Dimensions=[
                            {"Name": "BucketName", "Value": bucket_name},
                            {"Name": "StorageType", "Value": "StandardStorage"},
                        ],
                        Statistics=["Average"],
                        Period=3600,
                        StartTime=(now - datetime.timedelta(days=2)).isoformat(),
                        EndTime=now.isoformat(),
                    )
                    if len(response["Datapoints"]) != 0:
                        bucket_size = int(response["Datapoints"][0].get("Average"))
                    else:
                        bucket_size = "-"

                    if type(bucket_size) == int:
                        readable_bucket_size = humanbytes(bucket_size)
                    else:
                        readable_bucket_size = "-"

                    add_cell(sheet9, sheet9_cell_start, 6, readable_bucket_size)

                    try:
                        # Logging
                        bucket_logging = s3_cli.get_bucket_logging(Bucket=bucket_name)
                        if "LoggingEnabled" in bucket_logging:
                            add_cell(sheet9, sheet9_cell_start, 7, "Enabled")
                        else:
                            add_cell(sheet9, sheet9_cell_start, 7, "Disabled")

                        # Versioning
                        bucket_versioning = s3_cli.get_bucket_versioning(Bucket=bucket_name)
                        if "Status" in bucket_versioning:
                            add_cell(sheet9, sheet9_cell_start, 8, "Enabled")
                        else:
                            add_cell(sheet9, sheet9_cell_start, 8, "Disabled")
                    except Exception as e:
                        add_cell(sheet9, sheet9_cell_start, 7, "-")
                        add_cell(sheet9, sheet9_cell_start, 8, "-")
                        print(f"{bucket_name} : 비 정상적인 Bucket으로 추정", e)
                except:
                    pass


                    sheet9_cell_start += 1
        else:
            print("There is no S3.")
        
        # sheet10
        # CloudFront
        cloudfront = cloudfront_cli.list_distributions().get('DistributionList')

        if "Items" in cloudfront:
            try:
                sheet10 = wb.create_sheet("sheet10")
                sheet10.title = "CloudFront"
                cell_widths = [6, 6, 16, 28, 47, 16, 30, 10, 10, 25, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
                sheet_cell_width(sheet10, cell_widths)

                sheet10_cell_start = 2
                make_header(sheet10, sheet10_cell_start, "CloudFront")
                sheet10_cell_start += 1

                cell_headers = ["No.", "ID", "Domain Name", "Origin Domain Names", "Origin Groups (EA)", "CNAMEs", "Status", "State", "Price Class"]
                make_cell_header(sheet10, sheet10_cell_start, cell_headers)
                sheet10_cell_start += 1

                for idx, items in enumerate(cloudfront.get('Items')):
                    # No.
                    add_cell(sheet10, sheet10_cell_start, 2, idx + 1)

                    # Id
                    item_id = items.get("Id")
                    add_cell(sheet10, sheet10_cell_start, 3, item_id)

                    # Domain Name
                    distribution_domain_name = items.get('DomainName')
                    add_cell(sheet10, sheet10_cell_start, 4, distribution_domain_name)        

                    # Origin Domain Names
                    origin_domain = ""
                    origin = items.get('Origins')
                    if origin.get('Quantity') != 0:
                        for i, item in enumerate(origin["Items"]):
                            if i != 0:
                                origin_domain += ", \n"
                            origin_domain += item.get('DomainName')
                    else:
                        origin_domain = "-"
                    add_cell(sheet10, sheet10_cell_start, 5, origin_domain)

                    # Origin Groups (EA)
                    origin_group_cnt = items.get('OriginGroups').get('Quantity')
                    add_cell(sheet10, sheet10_cell_start, 6, origin_group_cnt)

                    # cf_cnames
                    cf_cnames = ""
                    aliases = items.get('Aliases')
                    if aliases.get('Quantity') != 0:
                        for i, cname in enumerate(aliases.get('Items')):
                            if i != 0:
                                cf_cnames += ", \n"
                            cf_cnames += cname
                    else:
                        cf_cnames = "-"
                    add_cell(sheet10, sheet10_cell_start, 7, cf_cnames)

                    # cf_status (Deployed, ..)
                    try:
                        cf_status = items.get('Status')
                    except Exception:
                        cf_status = "-"
                    add_cell(sheet10, sheet10_cell_start, 8, cf_status)        

                    # cf_state (Enabled, Disabled)
                    cf_state = items.get('Enabled')
                    if cf_state == True:
                        cf_state = "Enabled"
                    elif cf_state == False:
                        cf_state = "Disabled"
                    add_cell(sheet10, sheet10_cell_start, 9, cf_state)

                    # PriceClass
                    price_class = items.get('PriceClass')
                    if "All" in price_class:
                        price_class = "All Edge Locations \n(Best Performance)"
                    elif "100" in price_class:
                        price_class = "Only U.S, Canada and Europe"
                    elif "200" in price_class:
                        price_class = "Use U.S, Canada, Europe, \nAsia, Middle East and Afreeca"
                    add_cell(sheet10, sheet10_cell_start, 10, price_class)
                except:
                    pass
                    
                    sheet10_cell_start += 1

        else:
            print("There is no CloudFront.")
        
        # sheet11
        if cloudtrail_cli.describe_trails()["trailList"]:
            # print(cloudtrail_cli.describe_trails()['trailList'])

            sheet11 = wb.create_sheet("sheet11")
            sheet11.title = "CloudTrail"
            cell_widths = [6, 6, 25, 14, 16, 13, 15, 45, 8,8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet11, cell_widths)

            sheet11_cell_start = 2
            make_header(sheet11, sheet11_cell_start, "CloudTrail")
            sheet11_cell_start += 1

            cell_headers = ["No.", "Trail name", "Home Region", "Multi-region Trail", "Trail Insights",
                            "Organization Trail", "S3 bucket"]
            make_cell_header(sheet11, sheet11_cell_start, cell_headers)
            sheet11_cell_start += 1

            # NO.
            # Home Region
            # Trail name
            # Multi-region trails
            # trail	Insights
            # Organization trail
            # S3 bucket
            for idx, trail in enumerate(cloudtrail_cli.describe_trails()["trailList"]):
                add_cell(sheet11, sheet11_cell_start, 2, idx + 1)
                add_cell(sheet11, sheet11_cell_start, 3, trail.get("Name"))
                add_cell(sheet11, sheet11_cell_start, 4, trail.get("HomeRegion"))
                add_cell(sheet11, sheet11_cell_start, 5, trail.get("IsMultiRegionTrail"))
                add_cell(sheet11, sheet11_cell_start, 6, trail.get("HasInsightSelectors"))
                add_cell(sheet11, sheet11_cell_start, 7, trail.get("IsOrganizationTrail"))
                add_cell(sheet11, sheet11_cell_start, 8, trail.get("S3BucketName"))
                sheet11_cell_start = sheet11_cell_start + 1

            """ CloudTrail
            """
        else:
            print("There is no CloudTrail.")

        # sheet12
        # CloudWatch
        
        sheet12 = wb.create_sheet("sheet12")
        sheet12.title = "CloudWatch"
        cell_widths = [6, 6, 56, 56, 15, 23, 28, 10.75, 19, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
        sheet_cell_width(sheet12, cell_widths)

        sheet12_cell_start = 2
        make_header(sheet12, sheet12_cell_start, "CloudWatch")
        sheet12_cell_start += 1

        cell_headers = ["No.", "Dashboard Name", "Last Updated(UTC)", "Size"]
        make_cell_header(sheet12, sheet12_cell_start, cell_headers)
        sheet12_cell_start += 1

        dashboards = cloudwatch_cli.list_dashboards()
        for idx, dashboard in enumerate(dashboards["DashboardEntries"]):
            add_cell(sheet12, sheet12_cell_start, 2, idx + 1)
            add_cell(sheet12, sheet12_cell_start, 3, dashboard.get("DashboardName"))
            date = str(dashboard.get("LastModified"))
            add_cell(sheet12, sheet12_cell_start, 4, date.split("+")[0])
            add_cell(sheet12, sheet12_cell_start, 5, dashboard.get("Size"))
            sheet12_cell_start = sheet12_cell_start + 1

        sheet12_cell_start = sheet12_cell_start + 1
        cell_headers = ["No.", "Alarm Name", "AlarmDescription", "Namespace", "MetricName", "ComparisonOperator",
                    "Threshold", "StateValue"]

        make_cell_header(sheet12, sheet12_cell_start, cell_headers)
        sheet12_cell_start += 1

        alarms = cloudwatch_cli.describe_alarms()
        for idx, alarm in enumerate(alarms["MetricAlarms"]):
            add_cell(sheet12, sheet12_cell_start, 2, idx + 1)
            add_cell(sheet12, sheet12_cell_start, 3, alarm.get("AlarmName"))
            add_cell(sheet12, sheet12_cell_start, 4, alarm.get("AlarmDescription"))
            add_cell(sheet12, sheet12_cell_start, 5, alarm.get("Namespace"))
            add_cell(sheet12, sheet12_cell_start, 6, alarm.get("MetricName"))
            add_cell(sheet12, sheet12_cell_start, 7, alarm.get("ComparisonOperator"))
            add_cell(sheet12, sheet12_cell_start, 8, alarm.get("Threshold"))
            add_cell(sheet12, sheet12_cell_start, 9, alarm.get("StateValue"))
            sheet12_cell_start = sheet12_cell_start + 1
        

        # sheet13
        sheet13 = wb.create_sheet("sheet13")
        sheet13.title = "IAM User"
        cell_widths = [6, 6, 30, 30, 38, 33, 13, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
        sheet_cell_width(sheet13, cell_widths)

        sheet13_cell_start = 2
        make_header(sheet13, sheet13_cell_start, "IAM User")
        sheet13_cell_start += 1

        cell_headers = ["No.", "User Name", "Group Names", "Policies (Attached to Groups)", "Policies (Attached to User)",
                        "Creation Date"]
        make_cell_header(sheet13, sheet13_cell_start, cell_headers)

        sheet13_cell_start += 1

        for idx, user_detail in enumerate(iam_cli.get_account_authorization_details(Filter=["User"])["UserDetailList"]):
            add_cell(sheet13, sheet13_cell_start, 2, idx + 1)
            add_cell(sheet13, sheet13_cell_start, 3, str(user_detail.get("UserName")))

            if len(user_detail.get("GroupList")) != 0:
                # Group 리스트 생성
                group_list = []
                for idx, group in enumerate(user_detail.get("GroupList")):
                    group_list.append(group)

                str_group_list = ", \n".join(group_list)
                add_cell(sheet13, sheet13_cell_start, 4, str_group_list)

                # Group Policies
                group_policies = []
                for each_group in group_list:
                    iam_res_group = iam_res.Group(each_group)
                    policy_generator = iam_res_group.attached_policies.all()
                    for policy in policy_generator:
                        group_policies.append(policy.policy_name)

                if len(group_policies) != 0:
                    str_group_policies = ", \n".join(group_policies)
                    add_cell(sheet13, sheet13_cell_start, 5, str_group_policies)
                else:
                    add_cell(sheet13, sheet13_cell_start, 5, "-")

                # User Policies
                user_policies = []
                for policy in user_detail.get("AttachedManagedPolicies"):
                    user_policies.append(str(policy["PolicyName"]))

                only_user_policies = [x for x in user_policies if x not in group_policies]
                if len(only_user_policies) != 0:
                    str_only_user_policies = ", \n".join(only_user_policies)
                    add_cell(sheet13, sheet13_cell_start, 6, str_only_user_policies)
                else:
                    add_cell(sheet13, sheet13_cell_start, 6, "-")

            else:
                add_cell(sheet13, sheet13_cell_start, 4, "-")
                add_cell(sheet13, sheet13_cell_start, 5, "-")

                gpolicy = ""
                for idx, attachpolicy in enumerate(user_detail.get("AttachedManagedPolicies")):
                    if idx != 0:
                        gpolicy += ", \n"
                    gpolicy += attachpolicy["PolicyName"]
                add_cell(sheet13, sheet13_cell_start, 6, gpolicy)

            # Creation Date
            created_date = str(user_detail.get("CreateDate"))
            iamUser_date = created_date.split(" ")[0]
            add_cell(sheet13, sheet13_cell_start, 7, iamUser_date)

            sheet13_cell_start += 1

        # sheet14
        # IAM Roles
        iam_roles = iam_cli.list_roles()["Roles"]
        iam_roles_cnt = 0
        for i in iam_roles:
            if i["Path"] == "/":
                iam_roles_cnt += 1

        # Custom하게 생성한 IAM Roles가 1개 이상일 경우
        if iam_roles_cnt != 0:
            sheet14 = wb.create_sheet("sheet14")
            sheet14.title = "IAM Role"
            cell_widths = [6, 6, 60, 20, 70, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet14, cell_widths)

            sheet14_cell_start = 2
            make_header(sheet14, sheet14_cell_start, "IAM Roles")
            sheet14_cell_start += 1

            cell_headers = ["No.", "Role Name", "Creation Date", "Description"]
            make_cell_header(sheet14, sheet14_cell_start, cell_headers)

            sheet14_cell_start += 1

            count = 1
            for idx, iam_role in enumerate(iam_roles):
                # 직접 생성한 IAM Roles만 엑셀에 추가함
                if iam_role["Path"] == "/":
                    # No.
                    add_cell(sheet14, sheet14_cell_start, 2, count)

                    # Role Name
                    roleName = iam_role["RoleName"]
                    add_cell(sheet14, sheet14_cell_start, 3, roleName)

                    # Creation Date
                    role_creation_date = str(iam_role.get('CreateDate')).split(" ")[0]
                    add_cell(sheet14, sheet14_cell_start, 4, role_creation_date)

                    # Description
                    try:
                        iamRoleDesc = iam_role["Description"]
                        add_cell(sheet14, sheet14_cell_start, 5, iamRoleDesc)
                    except Exception:
                        add_cell(sheet14, sheet14_cell_start, 5, "-")

                    sheet14_cell_start = sheet14_cell_start + 1
                    count += 1
        else:
            print("There is no self made IAM Roles")
                
        # sheet15
        # Lambda

        if lambda_cli.list_functions().get("Functions"):
            sheet15 = wb.create_sheet("sheet15")
            sheet15.title = "Lambda"
            cell_widths = [6, 6, 30, 12, 55, 11, 12, 9, 12, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet15, cell_widths)

            sheet15_cell_start = 2
            make_header(sheet15, sheet15_cell_start, "Lambda")
            sheet15_cell_start += 1

            cell_headers = ["No.", "Function Name", "Runtime", "Role", "Code Size", "Memory Size", "Time Out", "Package Type"]
            make_cell_header(sheet15, sheet15_cell_start, cell_headers)
            sheet15_cell_start += 1

            for idx, each_lambda in enumerate(lambda_cli.list_functions().get("Functions")):
                add_cell(sheet15, sheet15_cell_start, 2, idx + 1)
                # Function Name
                add_cell(sheet15, sheet15_cell_start, 3, each_lambda.get("FunctionName"))
                # Runtime
                add_cell(sheet15, sheet15_cell_start, 4, each_lambda.get("Runtime"))
                # Role Name
                roleName = each_lambda.get("Role").split(":")[-1]
                add_cell(sheet15, sheet15_cell_start, 5, roleName)
                # Code Size
                readable_code_size = humanbytes(each_lambda.get("CodeSize"))
                add_cell(sheet15, sheet15_cell_start, 6, readable_code_size)
                # Memory Size
                add_cell(sheet15, sheet15_cell_start, 7, each_lambda.get("MemorySize"))
                # Timeout
                add_cell(sheet15, sheet15_cell_start, 8, each_lambda.get("Timeout"))
                # Package Type
                add_cell(sheet15, sheet15_cell_start, 9, each_lambda.get("PackageType"))
                sheet15_cell_start = sheet15_cell_start + 1

        else:
            print("There is no Lambda.")

        # Sheet16

        if route53_cli.list_hosted_zones_by_name()["HostedZones"]:
            sheet16 = wb.create_sheet("sheet16")
            sheet16.title = "Route 53"
            cell_widths = [6, 6, 33, 9, 65, 11, 75, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet16, cell_widths)

            sheet16_cell_start = 2
            make_header(sheet16, sheet16_cell_start, "Route 53")
            sheet16_cell_start += 1

            cell_headers = ["No.", "Domain Name", "Type", "Record Name", "Record Type", "Record Value"]
            make_cell_header(sheet16, sheet16_cell_start, cell_headers)
            sheet16_cell_start += 1

            hosted_zones = route53_cli.list_hosted_zones_by_name()["HostedZones"]
            # record_sets = route53_cli.list_resource_record_sets
            for idx, zone in enumerate(hosted_zones):
                zoneId = zone["Id"]

                # No.
                add_cell(sheet16, sheet16_cell_start, 2, idx + 1)

                # Domain Name
                zoneName = zone["Name"]
                real_zoneName = zoneName[:-1]
                add_cell(sheet16, sheet16_cell_start, 3, real_zoneName)

                # Type
                if zone["Config"]["PrivateZone"] == False:
                    add_cell(sheet16, sheet16_cell_start, 4, "Public")
                else:
                    add_cell(sheet16, sheet16_cell_start, 4, "Private")

                # Record Sets
                init_row_cnt = sheet16_cell_start
                row_cnt = 0
                for record_set in route53_cli.list_resource_record_sets(HostedZoneId=zoneId)["ResourceRecordSets"]:
                    row_cnt += 1

                    # Record Name
                    record_name = record_set.get("Name")

                    # Record Type
                    record_type = record_set.get("Type")

                    # Record Values
                    # Alias : AliasTarget
                    # Value : ResourceRecords
                    record_value = ""
                    if "ResourceRecords" in record_set:
                        for idx, value in enumerate(record_set.get("ResourceRecords")):
                            if idx != 0:
                                record_value = record_value + ", \n"
                            record_value += value.get("Value")
                    elif "AliasTarget" in record_set:
                        record_value = record_set["AliasTarget"]["DNSName"]

                    else:
                        record_value = "-"

                    add_cell(sheet16, sheet16_cell_start, 5, record_name)
                    add_cell(sheet16, sheet16_cell_start, 6, record_type)
                    add_cell(sheet16, sheet16_cell_start, 7, record_value)
                    sheet16_cell_start += 1

                sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=2, end_column=2)
                sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=3, end_column=3)
                sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=4, end_column=4)

                # sheet16_cell_start = sheet16_cell_start+1

        else:
            print("There is no Route53")

        # Sheet17
        # ECS
        response = ecs_cli.list_clusters().get('clusterArns')

        if len(response) != 0:
            sheet17 = wb.create_sheet("sheet18")
            sheet17.title = "ECS"
            cell_widths = [6, 6, 30, 7, 30, 13, 13, 16, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet17, cell_widths)

            sheet17_cell_start = 2
            make_header(sheet17, sheet17_cell_start, "ECS")
            sheet17_cell_start += 1

            cell_headers = ["No.", "Cluster Name", "Status", "Registered Container Instances", "Running Tasks",
                            "Active Services", "Container Insight"]
            make_cell_header(sheet17, sheet17_cell_start, cell_headers)
            sheet17_cell_start += 1

            # Get Cluster Names from ARN
            cluster_list = []
            for clus in response:
                name = str(clus).split('/')[-1]
                cluster_list.append(name)

            response2 = ecs_cli.describe_clusters(clusters=cluster_list).get('clusters')
            for idx, cluster in enumerate(response2):
                # No.
                add_cell(sheet17, sheet17_cell_start, 2, idx + 1)

                # Cluster Name
                ecs_name = cluster.get('clusterName')
                add_cell(sheet17, sheet17_cell_start, 3, ecs_name)

                # Status
                ecs_status = str(cluster.get('status')).lower().capitalize()
                add_cell(sheet17, sheet17_cell_start, 4, ecs_status)

                # Registered Container Instances
                ecs_registered_containers_cnt = cluster.get('registeredContainerInstancesCount')
                add_cell(sheet17, sheet17_cell_start, 5, ecs_registered_containers_cnt)

                # Running Tasks
                ecs_running_tasks_cnt = cluster.get('runningTasksCount')
                add_cell(sheet17, sheet17_cell_start, 6, ecs_running_tasks_cnt)

                # Active Services
                ecs_active_services_cnt = cluster.get('activeServicesCount')
                add_cell(sheet17, sheet17_cell_start, 7, ecs_active_services_cnt)

                # Container Insight
                for setting in cluster.get('settings'):
                    if setting.get('name') == 'containerInsights':
                        cont_insight = str(setting.get('value')).capitalize()
                        break
                add_cell(sheet17, sheet17_cell_start, 8, cont_insight)

                sheet17_cell_start += 1

        else:
            print("There is no ECS Clusters")


        # Sheet18

        # EKS가 있을 경우
        if eks_cli.list_clusters()["clusters"]:
            response = eks_cli.list_clusters()["clusters"]
            sheet18 = wb.create_sheet("sheet18")
            sheet18.title = "EKS"
            cell_widths = [6, 6, 18, 13, 14, 8, 46, 23, 20, 26, 22, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet18, cell_widths)

            sheet18_cell_start = 2
            make_header(sheet18, sheet18_cell_start, "EKS")
            sheet18_cell_start += 1

            cell_headers = ["No.", "Cluster Name", "Cluster Version", "Platform Version", "Status", "Subnet IDs",
                        "Cluster Security Group IDs", "Security Group IDs", "Node Group Names", "Fargate Profile Names"]
            make_cell_header(sheet18, sheet18_cell_start, cell_headers)
            sheet18_cell_start += 1

            for idx, cluster_name in enumerate(response):
                cluster = eks_cli.describe_cluster(name=cluster_name)["cluster"]

                # No.
                add_cell(sheet18, sheet18_cell_start, 2, idx + 1)

                # Cluster name
                name = cluster.get("name")
                add_cell(sheet18, sheet18_cell_start, 3, name)

                # Version
                version = cluster.get("version")
                add_cell(sheet18, sheet18_cell_start, 4, version)

                # Platform Version
                platform_ver = cluster.get("platformVersion")
                add_cell(sheet18, sheet18_cell_start, 5, platform_ver)

                # Status
                status = str(cluster.get("status")).lower().capitalize()
                add_cell(sheet18, sheet18_cell_start, 6, status)

                # Subnet IDs
                subnetIds = ""
                for i, subnetId in enumerate(cluster["resourcesVpcConfig"]["subnetIds"]):
                    if i != 0:
                        subnetIds += ", "
                    elif i != 0 and i % 2 == 0:
                        subnetIds += "\n"
                    subnetIds += subnetId
                add_cell(sheet18, sheet18_cell_start, 7, subnetIds)

                # Cluster Security Group IDs
                cluster_sg_ids = cluster["resourcesVpcConfig"].get("clusterSecurityGroupId")
                if cluster_sg_ids:
                    add_cell(sheet18, sheet18_cell_start, 8, cluster_sg_ids)
                else:
                    add_cell(sheet18, sheet18_cell_start, 8, "-")

                # Security Group IDs
                sg_ids_list = cluster["resourcesVpcConfig"].get("securityGroupIds")
                sg_ids = ""
                for i, sg_id in enumerate(sg_ids_list):
                    if i != 0:
                        sg_ids += ", \n"
                    sg_ids += sg_id
                add_cell(sheet18, sheet18_cell_start, 9, sg_ids)

                # Node Group Names
                node_groups = eks_cli.list_nodegroups(clusterName=name)["nodegroups"]
                if len(node_groups) != 0:
                    node_group_names = ""
                    for i, node_group in enumerate(node_groups):
                        if i != 0:
                            node_group_names += ", \n"
                        node_group_names += node_group
                    add_cell(sheet18, sheet18_cell_start, 10, node_group_names)
                else:
                    add_cell(sheet18, sheet18_cell_start, 10, "-")

                # Fargate Profile Names
                try:
                    fargates = eks_cli.list_fargate_profiles(clusterName="eks-test")["fargateProfileNames"]
                except:
                    fargates = []
                if len(fargates) != 0:
                    fg_names = ""
                    for i, fargate in enumerate(fargates):
                        if i != 0:
                            fg_names += ", \n"
                        fg_names += fargate
                    add_cell(sheet18, sheet18_cell_start, 11, fg_names)
                else:
                    add_cell(sheet18, sheet18_cell_start, 11, "-")
                
                sheet18_cell_start += 1

        else:
            print("There is no EKS")

        # ECR
        response = ecr_cli.describe_repositories().get('repositories')

        if len(response) != 0:

            sheet19 = wb.create_sheet("sheet19")
            sheet19.title = "ECR"
            cell_widths = [5, 5, 28, 70, 13, 13, 20, 8, 8, 5, 5, 8, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7]
            sheet_cell_width(sheet19, cell_widths)

            sheet19_cell_start = 2
            make_header(sheet19, sheet19_cell_start, "ECR")
            sheet19_cell_start += 1

            cell_headers = ["No.", "Repo Name", "Repo URI", "Creation Date", "Scan on Push",
                            "Encryption configuration"]
            make_cell_header(sheet19, sheet19_cell_start, cell_headers)
            sheet19_cell_start += 1

            for idx, ecr in enumerate(response):
                # No.
                add_cell(sheet19, sheet19_cell_start, 2, idx + 1)

                # Repository Name
                repo_name = ecr.get('repositoryName')
                add_cell(sheet19, sheet19_cell_start, 3, repo_name)

                # Repository URI
                repo_uri = ecr.get('repositoryUri')
                add_cell(sheet19, sheet19_cell_start, 4, repo_uri)

                # Creation Date
                repo_created = str(ecr.get('createdAt'))
                repo_created_date = repo_created.split(' ')[0]
                add_cell(sheet19, sheet19_cell_start, 5, repo_created_date)

                # Scan on push
                repo_scan_on_push = str(ecr.get('imageScanningConfiguration').get('scanOnPush')).lower().capitalize()
                add_cell(sheet19, sheet19_cell_start, 6, repo_scan_on_push)

                # Encryption Configuration
                repo_encryption_conf = ecr.get('encryptionConfiguration').get('encryptionType')
                add_cell(sheet19, sheet19_cell_start, 7, repo_encryption_conf)

                sheet19_cell_start += 1

        else:
            print("There is no ECR Repositories")


        # sheet19
        # ElastiCache
        response = els_cli.describe_cache_clusters()["CacheClusters"]
        if len(response) != 0:

            sheet20 = wb.create_sheet("sheet19")
            sheet20.title = "ElastiCache"
            cell_widths = [6, 6, 25, 15, 12, 13, 12, 12, 21, 28, 13, 8, 8.10, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet20, cell_widths)

            sheet20_cell_start = 2
            make_header(sheet20, sheet20_cell_start, "ElastiCache")
            sheet20_cell_start += 1

            cell_headers = ["No.", "Cluster Id", "Node Type", "Engine", "Engine Version", "Status", "Nodes (EA)", "Security Groups", "Subnet Group", "Creation Date"]
            make_cell_header(sheet20, sheet20_cell_start, cell_headers)

            sheet20_cell_start += 1

            for idx, node in enumerate(response):

                # No.
                add_cell(sheet20, sheet20_cell_start, 2, idx + 1)

                # Cluster Id
                cluster_id = node.get("CacheClusterId")
                add_cell(sheet20, sheet20_cell_start, 3, cluster_id)

                # Node Type
                node_type = node.get("CacheNodeType")
                add_cell(sheet20, sheet20_cell_start, 4, node_type)

                # Engine
                node_engine = str(node.get("Engine")).capitalize()
                add_cell(sheet20, sheet20_cell_start, 5, node_engine)

                # Engine Version
                engine_version = node.get("EngineVersion")
                add_cell(sheet20, sheet20_cell_start, 6, engine_version)

                # Status
                node_status = node.get("CacheClusterStatus")
                add_cell(sheet20, sheet20_cell_start, 7, node_status)

                # Number of Nodes
                node_count = node.get("NumCacheNodes")
                add_cell(sheet20, sheet20_cell_start, 8, node_count)

                # Security Groups
                node_sgs = ""
                for idx, sg in enumerate(node.get("SecurityGroups")):
                    if idx != 0:
                        node_sgs += ", \n"
                    node_sgs += sg.get("SecurityGroupId")
                add_cell(sheet20, sheet20_cell_start, 9, node_sgs)

                # Subnet Group
                node_subnet_group = node.get('CacheSubnetGroupName')
                add_cell(sheet20, sheet20_cell_start, 10, node_subnet_group)

                # Creation Date
                node_creation_date = str(node.get('CacheClusterCreateTime')).split(" ")[0]
                add_cell(sheet20, sheet20_cell_start, 11, node_creation_date)

                sheet20_cell_start = sheet20_cell_start + 1

        else:
            print("There is no ElastiCache")


        # sheet20
        response = efs_cli.describe_file_systems().get("FileSystems")

        if len(response) != 0:

            sheet21 = wb.create_sheet("sheet20")
            sheet21.title = "EFS"
            cell_widths = [6, 6, 13, 25, 13, 18, 14, 12, 21 , 17, 9, 17, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7]
            sheet_cell_width(sheet21, cell_widths)
            sheet21_cell_start = 2
            make_header(sheet21, sheet21_cell_start, "EFS")
            sheet21_cell_start += 1

            cell_headers = ["No.", "File System Id", "Name", "Creation Date", "Mount Targets (EA)", "Size (Standard)", "Size (IA)", "LifeCycle Configuration", "Performance Mode",
                        "Encrypted", "Throughput Mode"]
            make_cell_header(sheet21, sheet21_cell_start, cell_headers)
            sheet21_cell_start += 1

            for idx, efs in enumerate(response):
                # No.
                add_cell(sheet21, sheet21_cell_start, 2, idx + 1)

                # File System Id
                file_sys_id = efs.get("FileSystemId")
                add_cell(sheet21, sheet21_cell_start, 3, file_sys_id)

                # Name
                efs_name = efs.get("Name")
                add_cell(sheet21, sheet21_cell_start, 4, efs_name)

                # Creation Date
                creation_date = str(efs.get('CreationTime')).split(" ")[0]
                add_cell(sheet21, sheet21_cell_start, 5, creation_date)

                # Number of Mount Targets
                mnt_count = efs.get("NumberOfMountTargets")
                add_cell(sheet21, sheet21_cell_start, 6, mnt_count)

                # Size in Standard
                efs_standard_size = efs["SizeInBytes"].get("ValueInStandard")
                readable_standard_size = humanbytes(efs_standard_size)
                add_cell(sheet21, sheet21_cell_start, 7, readable_standard_size)
                
                # Size in IA
                efs_ia_size = efs["SizeInBytes"].get("ValueInIA")
                readable_ia_size = humanbytes(efs_ia_size)
                add_cell(sheet21, sheet21_cell_start, 8, readable_ia_size)

                # LifeCycle Configuration
                efs_lifecycle = efs_cli.describe_lifecycle_configuration(
                    FileSystemId=file_sys_id
                    ).get('LifecyclePolicies')
                if len(efs_lifecycle) != 0:
                    efs_transition_to_IA = efs_lifecycle[0].get('TransitionToIA')
                else:
                    efs_transition_to_IA = "None"
                add_cell(sheet21, sheet21_cell_start, 9, efs_transition_to_IA)
                

                # Performance Mode
                perf_mode = str(efs.get("PerformanceMode")).capitalize()
                add_cell(sheet21, sheet21_cell_start, 10, perf_mode)

                # Enctypted
                is_encrypted = str(efs.get("Encrypted")).lower().capitalize()
                add_cell(sheet21, sheet21_cell_start, 11, is_encrypted)

                # Throughput Mode
                through_mode = str(efs.get("ThroughputMode")).capitalize()
                add_cell(sheet21, sheet21_cell_start, 12, through_mode)

                sheet21_cell_start = sheet21_cell_start + 1
        else:
            print("There is no EFS")

        # Sheet 22
        # DynamoDB

        response = dynamo_cli.list_tables().get('TableNames')
        if len(response) != 0:

            sheet22 = wb.create_sheet("sheet20")
            sheet22.title = "DynamoDB"
            cell_widths = [5, 5, 35, 7, 13, 20, 20, 18, 18, 10, 10, 8, 8, 8, 8, 8, 8, 8, 7, 7, 7, 7]
            sheet_cell_width(sheet22, cell_widths)

            sheet22_cell_start = 2
            make_header(sheet22, sheet22_cell_start, "DynamoDB")
            sheet22_cell_start += 1

            cell_headers = ["No.", "Table Name", "Status", "Creation Date", "Partition Key", "Sort Key", "Total Read Capacity",
                            "Total Write Capacity", "Table Size", "Item Count"]
            make_cell_header(sheet22, sheet22_cell_start, cell_headers)
            sheet22_cell_start += 1

            for idx, table_name in enumerate(response):
                table = dynamo_cli.describe_table(TableName=table_name).get('Table')
                # No.
                add_cell(sheet22, sheet22_cell_start, 2, idx + 1)

                # Table name
                add_cell(sheet22, sheet22_cell_start, 3, table_name)

                # Status
                dynamo_status = str(table.get('TableStatus')).lower().capitalize()
                add_cell(sheet22, sheet22_cell_start, 4, dynamo_status)

                # Creation Date
                dynamo_created = str(table.get('CreationDateTime'))
                dynamo_created_date = dynamo_created.split(' ')[0]
                add_cell(sheet22, sheet22_cell_start, 5, dynamo_created_date)

                # Get Partition Key, Sort Key from KeySchema
                dynamo_part_key = "-"
                dynamo_sort_key = "-"
                for key in table.get('KeySchema'):
                    # HASH : Partition Key
                    if key.get('KeyType') == "HASH":
                        dynamo_part_key = key.get('AttributeName')
                    # RANGE : Sort Key
                    if key.get('KeyType') == 'RANGE':
                        dynamo_sort_key = key.get('AttributeName')

                # Get Key`s Attribute Type (Number, String, Binary)
                for attribute in table.get('AttributeDefinitions'):
                    if attribute.get('AttributeName') == dynamo_part_key:
                        dynamo_part_key_type = str(attribute.get('AttributeType'))
                        if dynamo_part_key_type == "S":
                            dynamo_part_key_type = "String"
                        elif dynamo_part_key_type == "N":
                            dynamo_part_key_type = "Number"
                        else:
                            dynamo_part_key_type = "Binary"
                    try:
                        if attribute.get('AttributeName') == dynamo_sort_key:
                            dynamo_sort_key_type = attribute.get('AttributeType')
                            if dynamo_sort_key_type == "S":
                                dynamo_sort_key_type = "String"
                            elif dynamo_sort_key_type == "N":
                                dynamo_sort_key_type = "Number"
                            else:
                                dynamo_sort_key_type = "Binary"
                    except:
                        pass


                # Primary Key
                dynamo_partition_key = "-"
                if dynamo_part_key != "-":
                    dynamo_partition_key = str(dynamo_part_key) + " (" + dynamo_part_key_type + ")"
                add_cell(sheet22, sheet22_cell_start, 6, dynamo_partition_key)

                # Sort Key
                dynamo_sorted_key = "-"
                if dynamo_sort_key != "-":
                    dynamo_sorted_key = str(dynamo_sort_key) + " (" + dynamo_sort_key_type + ")"
                add_cell(sheet22, sheet22_cell_start, 7, dynamo_sorted_key)

                # Total Read Capacity
                dynamo_total_read_cap = table.get('ProvisionedThroughput').get('ReadCapacityUnits')
                add_cell(sheet22, sheet22_cell_start, 8, dynamo_total_read_cap)

                # Total Write Capacity
                dynamo_total_write_cap = table.get('ProvisionedThroughput').get('WriteCapacityUnits')
                add_cell(sheet22, sheet22_cell_start, 9, dynamo_total_write_cap)

                # Table Size
                dynamo_table_size = table.get('TableSizeBytes')
                readable_dynamo_table_size = humanbytes(dynamo_table_size)
                add_cell(sheet22, sheet22_cell_start, 10, readable_dynamo_table_size)

                # Item Count
                dynamo_item_count = table.get('ItemCount')
                add_cell(sheet22, sheet22_cell_start, 11, dynamo_item_count)

                sheet22_cell_start += 1


        else:
            print("There is no DynamoDB")


        # sheet23
        # Elastic Search
        response = es_cli.list_domain_names().get('DomainNames')

        if len(response) != 0:
            try:
                sheet23 = wb.create_sheet("sheet23")
                sheet23.title = "ElasticSearch"
                cell_widths = [6, 5, 25, 10, 13, 16, 20, 24, 20, 12, 20, 12, 10, 12, 16, 8, 8, 8, 7, 7, 7, 7]
                sheet_cell_width(sheet23, cell_widths)

                sheet23_cell_start = 2
                make_header(sheet23, sheet23_cell_start, "ElasticSearch")
                sheet23_cell_start += 1

                cell_headers = ["No.", "Domain Name", "ES Version", "Creation Date", "Availability Zones", "VPC ID", "Subnet IDs", "Security Group IDs", 
                                "Cluster State", "Instance Type", "Instances (EA)", "EBS Type", "EBS Size (GB)", "Auto Tune Option"]
                make_cell_header(sheet23, sheet23_cell_start, cell_headers)
                sheet23_cell_start += 1

                for idx, domain_name in enumerate(response):
                    es_domain_name = domain_name.get('DomainName')

                    es_config = es_cli.describe_elasticsearch_domain_config(
                        DomainName=es_domain_name
                        ).get('DomainConfig')
                    
                    # No.
                    add_cell(sheet23, sheet23_cell_start, 2, idx + 1)

                    # Domain Name
                    add_cell(sheet23, sheet23_cell_start, 3, es_domain_name)

                    # ES Version
                    es_version = es_config.get('ElasticsearchVersion').get('Options')
                    add_cell(sheet23, sheet23_cell_start, 4, es_version)

                    # Creation Date
                    es_cluster_creation_date = str(es_config.get('ElasticsearchClusterConfig').get('Status').get('CreationDate')).split(" ")[0]
                    add_cell(sheet23, sheet23_cell_start, 5, es_cluster_creation_date)
                    
                    # Availability Zones
                    es_azs = ""
                    azs = es_config.get('VPCOptions').get('Options').get('AvailabilityZones')
                    for i, az in enumerate(azs):
                        if i != 0:
                            es_azs += ", \n"
                        es_azs += az
                    add_cell(sheet23, sheet23_cell_start, 6, es_azs)

                    # VPC 
                    es_vpc_id = es_config.get('VPCOptions').get('Options').get('VPCId')
                    add_cell(sheet23, sheet23_cell_start, 7, es_vpc_id)
                    
                    # Subnet IDs
                    es_subnet_ids = ""
                    subnets = es_config.get('VPCOptions').get('Options').get('SubnetIds')
                    for i, subnet in enumerate(subnets):
                        if i != 0:
                            es_subnet_ids += ", \n"
                        es_subnet_ids += subnet
                    add_cell(sheet23, sheet23_cell_start, 8, es_subnet_ids)
                    
                    # Security Group IDs
                    es_sg_ids = ""
                    sgs = es_config.get('VPCOptions').get('Options').get('SecurityGroupIds')
                    for i, sg in enumerate(sgs):
                        if i != 0:
                            es_sg_ids += ", \n"
                        es_sg_ids += sg
                    add_cell(sheet23, sheet23_cell_start, 9, es_sg_ids)
                    
                    # ES Cluster State
                    es_cluster_state = es_config.get('ElasticsearchClusterConfig').get('Status').get('State')
                    add_cell(sheet23, sheet23_cell_start, 10, es_cluster_state)

                    # Instance Type
                    es_instance_type = es_config.get('ElasticsearchClusterConfig').get('Options').get('InstanceType')
                    add_cell(sheet23, sheet23_cell_start, 11, es_instance_type)

                    # Instance Count
                    es_instance_count = es_config.get('ElasticsearchClusterConfig').get('Options').get('InstanceCount')
                    add_cell(sheet23, sheet23_cell_start, 12, es_instance_count)

                    # EBS (Volume Type, Size)
                    is_es_ebs_enabled = es_config.get('EBSOptions').get('Options').get('EBSEnabled')
                    if is_es_ebs_enabled:
                        # EBS Volume Type
                        es_ebs_type = es_config.get('EBSOptions').get('Options').get('VolumeType')

                        # EBS Volume Size
                        es_ebs_size = es_config.get('EBSOptions').get('Options').get('VolumeSize')
                    else:
                        es_ebs_type = "-"
                        es_ebs_size = "-"
                    
                    add_cell(sheet23, sheet23_cell_start, 13, es_ebs_type)
                    add_cell(sheet23, sheet23_cell_start, 14, es_ebs_size)
                    
                    # Auto Tune Options
                    es_auto_tune = str(es_config.get("AutoTuneOptions").get("Options").get("DesiredState")).lower().capitalize()
                    add_cell(sheet23, sheet23_cell_start, 15, es_auto_tune)
                    
                    sheet23_cell_start += 1
            except Exception as e:
                print(f"Elastic Search 오류 발생, {e}")

        else:
            print("There is no Elasticsearch.")
        
        # Sheet 24
        # MSK

        wb.save(storing_path + p_name + "_자산 내역_" + today_date + ".xlsx")
        
        print("")
        sleep(5)
    
    except Exception as e:
        print(f"Error 발생, 고객사 : {p_name}, 에러 내용: {e}")
        print("")
    
    finally:
        wb.close()