import boto3
import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment, Color, PatternFill
import datetime
from dateutil.relativedelta import relativedelta
from time import sleep

# cell box
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),)


# custom 함수 생성
def add_cell(sheet, ro, col, values):
    sheet.cell(row=ro, column=col).value = values
    sheet.cell(row=ro, column=col).font = Font(name="맑은 고딕", size=10)
    sheet.cell(row=ro, column=col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    sheet.cell(row=ro, column=col).border = thin_border


def humanbytes(B):
    """
    Return the given bytes as a human friendly KiB, MiB, GiB, or TiB string
    """
    B = float(B)
    KiB = float(1024)
    MiB = float(KiB ** 2)  # 1,048,576
    GiB = float(KiB ** 3)  # 1,073,741,824
    TiB = float(KiB ** 4)  # 1,099,511,627,776
    if B < KiB:
        return "{0} {1}".format(B, "Bytes" if 0 == B > 1 else "Byte")
    elif KiB <= B < MiB:
        return "{0:.1f} KiB".format(B / KiB)
    elif MiB <= B < GiB:
        return "{0:.1f} MiB".format(B / MiB)
    elif GiB <= B < TiB:
        return "{0:.1f} GiB".format(B / GiB)
    elif TiB <= B:
        return "{0:.1f} TiB".format(B / TiB)


# cell width
def sheet_cell_width(sheet, sheet_cell_widths):
    dimensions = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", 
                "N", "O", "P", 'Q', 'R', 'S', 'T', 'U', 'V']
    for k in range(0, 22):
        sheet.column_dimensions[dimensions[k]].width = sheet_cell_widths[k]


# Transport Port Number into Type ( Dictionary )
def check_type(ip_type, number):
    """
    Transport Port Number into Type
    ipType(str), number(int)
    """
    transported_type = ip_type  # tcp, upd, icmp, icmpv6 or numbers
    target_dict = {22: "SSH", 25: "SMTP", 53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP", 389: "LDAP", 443: "HTTPS",
                445: "SMB", 465: "SMTPS", 993: "IMAPS", 995: "POP3S", 1433: "MSSQL", 2049: "NFS", 3306: "MySQL/Aurora",
                3389: "RDP", 5439: "Redshift", 5432: "PostgreSQL", 1521: "Oracle-RDS", 5985: "WirnRM-HTTP",
                5986: "WinRM-HTTPS", 2007: "Elastic-Graphics"}

    if type(number) == int:
        if number in target_dict:
            transported_type = target_dict.get(number)
        elif type(number) == "tcp" or type(number) == "udp":
            transported_type = "Custom {0} Rule".format(ip_type.upper())
        return transported_type
    else:
        if "-" in str(number):
            transported_type = "Custom {0} Rule".format(ip_type.upper())

    return transported_type


def make_header(sheet, start_row, title):
    sheet.cell(row=start_row, column=2).value = title
    sheet.cell(row=start_row, column=2).font = Font(name="맑은 고딕", size=12, bold=True)


def make_cell_header(sheet, cell_start, head):
    for colu_index, hea in enumerate(head):
        sheet.cell(row=cell_start, column=colu_index + 2).value = hea
        sheet.cell(row=cell_start, column=colu_index + 2).font = Font(name="맑은 고딕", size=10, bold=True)
        sheet.cell(row=cell_start, column=colu_index + 2).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=cell_start, column=colu_index + 2).fill = PatternFill(patternType="solid", fgColor=Color("E3E3E3"))
        sheet.cell(row=cell_start, column=colu_index + 2).border = thin_border


p_name = "default"
r_name = "ap-northeast-2"

# Profile Name의 Credentials 정보를 이용하여 Session 맺음
session = boto3.session.Session(profile_name=p_name)

# Session을 이용하여 Resource or Client 객체 생성
ec2_res = session.resource(service_name="ec2", region_name=r_name)
ec2_cli = session.client(service_name="ec2", region_name=r_name)
elb_cli = session.client(service_name="elb", region_name=r_name)  # CLB
elbv2_cli = session.client(service_name="elbv2", region_name=r_name)  # ALB, NLB, GLB
rds_cli = session.client("rds", region_name=r_name)
s3_cli = session.client(service_name="s3", region_name=r_name)
s3_res = session.resource(service_name="s3", region_name=r_name)
cloudfront_cli = session.client(service_name="cloudfront", region_name=r_name)
cloudtrail_cli = session.client(service_name="cloudtrail", region_name=r_name)
cloudwatch_cli = session.client(service_name="cloudwatch", region_name=r_name)
lambda_cli = session.client(service_name="lambda", region_name=r_name)
iam_cli = session.client(service_name="iam", region_name=r_name)
iam_res = session.resource(service_name="iam", region_name=r_name)
route53_cli = session.client(service_name="route53", region_name=r_name)
eks_cli = session.client(service_name="eks", region_name=r_name)
ecs_cli = session.client(service_name="ecs", region_name=r_name)
els_cli = session.client(service_name="elasticache", region_name=r_name)
efs_cli = session.client(service_name="efs", region_name=r_name)
ecr_cli = session.client(service_name="ecr", region_name=r_name)
dynamo_cli = session.client(service_name="dynamodb", region_name=r_name)
es_cli = session.client(service_name="es", region_name=r_name)



# Workbook 생성
wb = openpyxl.Workbook()


vpc = wb.active
vpc.title = "VPC"

print(vpc)

print(type(vpc))

print(dir(vpc))