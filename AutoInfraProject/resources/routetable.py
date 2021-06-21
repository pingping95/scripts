from .settings import Common
import openpyxl
import boto3
import main

class RouteTable(Common):
    def __init__(self, name, workbook, ses, info, log, is_run = False):
        Common.__init__(self, name)
        if is_run:
            self.log = log
            self.wb = workbook
            self.profile = info.get('porfile')
            self.region = info.get('region')
            self.resource = ses.resource(service_name="ec2", region_name=self.region)
            self.run()

    def run(self):
        try:
            print(f"name: {self.name}, profile: {self.profile}, res: {self.resource}")
            # Initialize
            self.sheet = self.wb.create_sheet(self.name)
            self.sheet.title = self.name
            
            # Cell width
            cell_widths = [5, 5, 25, 23, 22, 25, 16, 20, 22, 15, 20, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7]
            self.fit_cell_width(cell_widths)
            
            # Header
            self.make_header(self.cell_start, self.name)
            
            # Cell header
            cell_headers = ["No.", "Name", "ID", "Destination", "Target"]
            self.make_cell_header(self.cell_start, cell_headers)
            
            # For loop
            for idx, route in enumerate(self.resource.route_tables.all()):
                long_start = self.cell_start
                short_start = self.cell_start
                route_table = self.resource.RouteTable(route.id)
                route_asso = route_table.associations_attribute
                # No.
                self.add_cell(self.cell_start, 2, idx + 1)
                # Name
                try:
                    self.add_cell(self.cell_start, 3, route_table.tags[0].get("Value"))
                except:
                    self.add_cell(self.cell_start, 3, "-")
                # Id
                self.add_cell(self.cell_start, 4, route_table.id)
                for i, info in enumerate(route_table.routes_attribute):
                    self.add_cell(long_start, 5, info.get('DestinationCidrBlock'))
                    if info.get('GatewayId'):
                        self.add_cell(long_start, 6, info.get("GatewayId"))
                    elif info.get('NatGatewayId'):
                        self.add_cell(long_start, 6, info.get("NatGatewayId"))
                    elif info.get('TransitGatewayId'):
                        self.add_cell(long_start, 6, info.get("TransitGatewayId"))
                    elif info.get('LocalGatewayId'):
                        self.add_cell(long_start, 6, info.get("LocalGatewayId"))
                    else:
                        self.add_cell(long_start, 6, info.get("NetworkInterfaceId"))
                    long_start += 1
                
                # Cell Merge
                self.sheet.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=2, end_column=2)
                self.sheet.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=3, end_column=3)
                self.sheet.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=4, end_column=4)
                
                self.cell_start = long_start
                
        except Exception as e:
            self.log.write(f"Error 발생, 리소스: {self.name}, 내용: {e}")