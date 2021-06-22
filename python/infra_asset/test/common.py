import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment, Color, PatternFill
from dateutil.relativedelta import relativedelta
import datetime
import boto3

# Profile Name와 Region 기입 필수
p_name = 'gcube_ssm_user'
r_name = ''

# Profile Name의 Credentials 정보를 이용하여 Session 맺음
session = boto3.session.Session(profile_name=p_name)


class Common:
    # cell box
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # custom 함수 생성
    def add_cell(sheet, ro, col, values):
        sheet.cell(row=ro, column=col).value = values
        sheet.cell(row=ro, column=col).font = Font(name='맑은 고딕', size=10)
        sheet.cell(row=ro, column=col).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        sheet.cell(row=ro, column=col).border = thin_border

    def humanbytes(B):
        '''
        Return the given bytes as a human friendly KiB, MiB, GiB, or TiB string
        '''
        B = float(B)
        KiB = float(1024)
        MiB = float(KiB ** 2)  # 1,048,576
        GiB = float(KiB ** 3)  # 1,073,741,824
        TiB = float(KiB ** 4)  # 1,099,511,627,776
        if B < KiB:
            return '{0} {1}'.format(B, 'Bytes' if 0 == B > 1 else 'Byte')
        elif KiB <= B < MiB:
            return '{0:.1f} KiB'.format(B / KiB)
        elif MiB <= B < GiB:
            return '{0:.1f} MiB'.format(B / MiB)
        elif GiB <= B < TiB:
            return '{0:.1f} GiB'.format(B / GiB)
        elif TiB <= B:
            return '{0:.1f} TiB'.format(B / TiB)

    # custom 함수 생성
    def add_cell(sheet, ro, col, values):
        sheet.cell(row=ro, column=col).value = values
        sheet.cell(row=ro, column=col).font = Font(name='맑은 고딕', size=10)
        sheet.cell(row=ro, column=col).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        sheet.cell(row=ro, column=col).border = thin_border

    # cell width
    def sheet_cell_width(sheet, cell_widths):
        sheet.column_dimensions['A'].width = cell_widths[0]
        sheet.column_dimensions['B'].width = cell_widths[1]
        sheet.column_dimensions['C'].width = cell_widths[2]
        sheet.column_dimensions['D'].width = cell_widths[3]
        sheet.column_dimensions['E'].width = cell_widths[4]
        sheet.column_dimensions['F'].width = cell_widths[5]
        sheet.column_dimensions['G'].width = cell_widths[6]
        sheet.column_dimensions['H'].width = cell_widths[7]
        sheet.column_dimensions['I'].width = cell_widths[8]
        sheet.column_dimensions['J'].width = cell_widths[9]
        sheet.column_dimensions['K'].width = cell_widths[10]
        sheet.column_dimensions['L'].width = cell_widths[11]
        sheet.column_dimensions['M'].width = cell_widths[12]
        sheet.column_dimensions['N'].width = cell_widths[13]
        sheet.column_dimensions['O'].width = cell_widths[14]
        sheet.column_dimensions['P'].width = cell_widths[15]
        sheet.column_dimensions['Q'].width = cell_widths[16]
        sheet.column_dimensions['R'].width = cell_widths[17]
        sheet.column_dimensions['S'].width = cell_widths[18]
        sheet.column_dimensions['T'].width = cell_widths[19]
        sheet.column_dimensions['U'].width = cell_widths[20]
        sheet.column_dimensions['V'].width = cell_widths[21]

    # Transport Port Number into Type ( Dictionary )
    def TypeCheck(ipType, number):
        """
        Transport Port Number into Type
        ipType(str), number(int)
        """
        transported_type = ipType  # tcp, upd, icmp, icmpv6 or numbers
        target_dict = {22: "SSH", 25: "SMTP", 53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP", 389: "LDAP",
                       443: "HTTPS", 445: "SMB", 465: "SMTPS", 993: "IMAPS", 995: "POP3S", 1433: "MSSQL", 2049: "NFS",
                       3306: "MySQL/Aurora", 3389: "RDP", 5439: "Redshift", 5432: "PostgreSQL", 1521: "Oracle-RDS",
                       5985: "WirnRM-HTTP", 5986: "WinRM-HTTPS", 2007: "Elastic-Graphics"}

        if type(number) == int:
            if number in target_dict:
                transported_type = target_dict.get(number)
            elif type(number) == "tcp" or type(number) == "udp":
                transported_type = "Custom {0} Rule".format(ipType.upper())
            return transported_type
        else:
            if '-' in str(number):
                transported_type = "Custom {0} Rule".format(ipType.upper())

        return transported_type
