import boto3
import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment, Color, PatternFill
import datetime
import pprint
from dateutil.relativedelta import relativedelta

# Profile Name와 Region 기입 필수
p_name = 'hyundai_brazil_prd'
r_name = 'sa-east-1'

# Profile Name의 Credentials 정보를 이용하여 Session 맺음
session = boto3.session.Session(profile_name=p_name)

# Session을 이용하여 Resource or Client 객체 생성
ec2_res = session.resource(service_name="ec2", region_name=r_name)
ec2_cli = session.client(service_name="ec2", region_name=r_name)
elb_cli = session.client(service_name='elb', region_name=r_name)        # CLB
elbv2_cli = session.client(service_name='elbv2', region_name=r_name)    # ALB, NLB, GLB
rds_cli = session.client('rds', region_name=r_name)
s3_cli = session.client(service_name="s3", region_name=r_name)
s3_res = session.resource(service_name="s3", region_name=r_name)
cloudfront_cli = session.client(service_name="cloudfront", region_name=r_name)
cloudtrail_cli = session.client(service_name="cloudtrail", region_name=r_name)
cloudwatch_cli = session.client(service_name="cloudwatch", region_name=r_name)
lambda_cli = session.client(service_name="lambda", region_name=r_name)
iam_cli = session.client(service_name="iam", region_name=r_name)
iam_res = session.resource(service_name="iam", region_name=r_name)
route53_cli = session.client(service_name="route53", region_name=r_name)
eks_cli = session.client(service_name='eks', region_name=r_name)
ecs_cli = session.client(service_name='ecs', region_name=r_name)
els_cli = session.client(service_name='elasticache', region_name=r_name)


right_now = datetime.datetime.now()
now = right_now.date()
target_date = str(now + relativedelta(days=-2))

# cell box
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# custom 함수 생성
def add_cell(sheet, ro, col, values):
    sheet.cell(row=ro, column=col).value = values
    sheet.cell(row=ro, column=col).font = Font(name='맑은 고딕', size=10)
    sheet.cell(row=ro, column=col).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet.cell(row=ro, column=col).border = thin_border

def humanbytes(B):
    '''
    Return the given bytes as a human friendly KiB, MiB, GiB, or TiB string
    '''
    B = float(B)
    KiB = float(1024)
    MiB = float(KiB ** 2)  # 1,048,576
    GiB = float(KiB ** 3)  # 1,073,741,824
    TiB = float(KiB ** 4)  # 1,099,511,627,776
    if B < KiB:
        return '{0} {1}'.format(B, 'Bytes' if 0 == B > 1 else 'Byte')
    elif KiB <= B < MiB:
        return '{0:.1f} KiB'.format(B / KiB)
    elif MiB <= B < GiB:
        return '{0:.1f} MiB'.format(B / MiB)
    elif GiB <= B < TiB:
        return '{0:.1f} GiB'.format(B / GiB)
    elif TiB <= B:
        return '{0:.1f} TiB'.format(B / TiB)

# cell width
def sheet_cell_width(sheet, cell_widths):
    sheet.column_dimensions['A'].width = cell_widths[0]
    sheet.column_dimensions['B'].width = cell_widths[1]
    sheet.column_dimensions['C'].width = cell_widths[2]
    sheet.column_dimensions['D'].width = cell_widths[3]
    sheet.column_dimensions['E'].width = cell_widths[4]
    sheet.column_dimensions['F'].width = cell_widths[5]
    sheet.column_dimensions['G'].width = cell_widths[6]
    sheet.column_dimensions['H'].width = cell_widths[7]
    sheet.column_dimensions['I'].width = cell_widths[8]
    sheet.column_dimensions['J'].width = cell_widths[9]
    sheet.column_dimensions['K'].width = cell_widths[10]
    sheet.column_dimensions['L'].width = cell_widths[11]
    sheet.column_dimensions['M'].width = cell_widths[12]
    sheet.column_dimensions['N'].width = cell_widths[13]
    sheet.column_dimensions['O'].width = cell_widths[14]
    sheet.column_dimensions['P'].width = cell_widths[15]
    sheet.column_dimensions['Q'].width = cell_widths[16]
    sheet.column_dimensions['R'].width = cell_widths[17]
    sheet.column_dimensions['S'].width = cell_widths[18]
    sheet.column_dimensions['T'].width = cell_widths[19]
    sheet.column_dimensions['U'].width = cell_widths[20]
    sheet.column_dimensions['V'].width = cell_widths[21]

# Transport Port Number into Type ( Dictionary )
def TypeCheck(ipType, number):
    """
    Transport Port Number into Type
    ipType(str), number(int)
    """
    transported_type = ipType  # tcp, upd, icmp, icmpv6 or numbers
    target_dict = {22: "SSH", 25: "SMTP", 53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP", 389: "LDAP",
                   443: "HTTPS", 445: "SMB", 465: "SMTPS", 993: "IMAPS", 995: "POP3S", 1433: "MSSQL", 2049: "NFS",
                   3306: "MySQL/Aurora", 3389: "RDP", 5439: "Redshift", 5432: "PostgreSQL", 1521: "Oracle-RDS",
                   5985: "WirnRM-HTTP", 5986: "WinRM-HTTPS", 2007: "Elastic-Graphics"}

    if type(number) == int:
        if number in target_dict:
            transported_type = target_dict.get(number)
        elif type(number) == "tcp" or type(number) == "udp":
            transported_type = "Custom {0} Rule".format(ipType.upper())
        return transported_type
    else:
        if '-' in str(number):
            transported_type = "Custom {0} Rule".format(ipType.upper())

    return transported_type

# Workbook 생성
wb = openpyxl.Workbook()
print(f"profile name : {p_name}\n region name : {r_name}\n(고객 환경마다 엑셀 생성 시간 상이)\n")

'''
sheet 1
'''

sheet1 = wb.active
sheet1.title = 'VPC'
# 10
cell_widths = [5, 5, 18, 25, 25, 25, 17, 15, 20, 20, 5, 5, 18, 24, 24, 13, 13, 13, 13, 13, 7, 7]
sheet_cell_width(sheet1, cell_widths)

title = sheet1['B2']
title.value = "<VPC>"
title.font = Font(name='맑은 고딕', size=12, bold=True)

'''
VPC
'''

sheet1_cell_start = 3
cell_header = ["No.", "VPC Name", "VPC ID", "VPC CIDR Block"]

for col_index, header in enumerate(cell_header):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).border = thin_border

sheet1_cell_start = sheet1_cell_start + 1

for idx, vpc in enumerate(ec2_res.vpcs.all()):
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=2, values=idx + 1)
    try:
        for tag in vpc.tags:
            if tag["Key"] == "Name":
                add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values=tag["Value"])
    except Exception as e:
        add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values='-')
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=4, values=vpc.id)
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=5, values=vpc.cidr_block)

    sheet1_cell_start = sheet1_cell_start + 1

'''
Network ACLS
'''

sheet1_cell_start = sheet1_cell_start + 2
sheet1.cell(row=sheet1_cell_start, column=2).value = "<Network ACL>"
sheet1.cell(row=sheet1_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)
sheet1_cell_start = sheet1_cell_start + 1

# 1. Inbound
acl_start_row = sheet1_cell_start

cell_header_upper = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Inbound Rule", "", "", "", ""]
for col_index, header in enumerate(cell_header_upper):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).border = thin_border
sheet1_cell_start = sheet1_cell_start + 1

acl_end_row = sheet1_cell_start

cell_header_lower = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Rule", "Protocol", "Port Range", "Source",
                     "Allow / Deny"]
for col_index, header in enumerate(cell_header_lower):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).border = thin_border
sheet1_cell_start = sheet1_cell_start + 1

sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=2, end_column=2)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=3, end_column=3)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=4, end_column=4)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=5, end_column=5)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_start_row, start_column=6, end_column=10)

# 2. OutBound
sheet1_cell_start = acl_start_row

acl_start_row = sheet1_cell_start

cell_header_upper = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Outbound Rule", "", "", "", ""]
for col_index, header in enumerate(cell_header_upper):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).border = thin_border
sheet1_cell_start = sheet1_cell_start + 1

acl_end_row = sheet1_cell_start

cell_header_lower = ["No.", "Network ACL Name", "Network ACL Id", "VPC ID", "Rule", "Protocol", "Port Range", "Source",
                     "Allow / Deny"]
for col_index, header in enumerate(cell_header_lower):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 12).border = thin_border
sheet1_cell_start = sheet1_cell_start + 1

sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=12, end_column=12)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=13, end_column=13)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=14, end_column=14)
sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=15, end_column=15)

sheet1.merge_cells(start_row=acl_start_row, end_row=acl_start_row, start_column=16, end_column=20)

# 값 입력
# Inbound Rule

init_sheet1_cell_start = sheet1_cell_start

for idx, acl in enumerate(ec2_cli.describe_network_acls()['NetworkAcls']):

    acl_start_row = sheet1_cell_start

    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=2, values=idx + 1)  # Number
    name = '-'
    try:
        for tag in acl['Tags']:
            if tag['Key'] == 'Name':
                name = tag['Value']
    except Exception:
        pass

    acl_id = acl['NetworkAclId']
    vpc_id = acl['VpcId']

    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values=name)  # Network ACL Name
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=4, values=acl_id)
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=5, values=vpc_id)

    in_count = 0
    for entry in acl['Entries']:
        if entry['Egress'] == False:  # Ingress
            in_count += 1

            if 'CidrBlock' in entry:
                source = entry['CidrBlock']
            else:
                source = "-"
            protocol = "All"
            portRange = "All"

            # ruleNumber
            if entry['RuleNumber'] >= 32737:
                ruleNumber = "*"
            else:
                ruleNumber = entry['RuleNumber']

            # RuleAction
            ruleAction = entry['RuleAction']

            # protocol, PortRange
            if entry['Protocol'] == '-1':
                protocol = "All"
            elif entry['Protocol'] == '6':
                protocol = "TCP"
                if entry['PortRange']['To'] == entry['PortRange']['From']:
                    portRange = entry['PortRange']['From']
                else:
                    portRange = str(entry['PortRange']['From']) + " - " + str(entry['PortRange']['To'])
            elif entry['Protocol'] == '17':
                protocol = "UDP"
                if entry['PortRange']['To'] == entry['PortRange']['From']:
                    portRange = entry['PortRange']['From']
                else:
                    portRange = str(entry['PortRange']['From']) + " - " + str(entry['PortRange']['To'])

            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=6, values=ruleNumber)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=7, values=protocol)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=8, values=portRange)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=9, values=source)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=10, values=ruleAction.capitalize())

            sheet1_cell_start += 1

    acl_end_row = sheet1_cell_start - 1

    # merge cells
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=2, end_column=2)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=3, end_column=3)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=4, end_column=4)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=5, end_column=5)

subnet_sheet1_cell_start = sheet1_cell_start

# Outbound

sheet1_cell_start = init_sheet1_cell_start

for idx, acl in enumerate(ec2_cli.describe_network_acls()['NetworkAcls']):
    acl_start_row = sheet1_cell_start

    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=12, values=idx + 1)  # Number
    name = '-'
    try:
        for tag in acl['Tags']:
            if tag['Key'] == 'Name':
                name = tag['Value']
    except Exception:
        pass

    acl_id = acl['NetworkAclId']
    vpc_id = acl['VpcId']

    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=13, values=name)  # Network ACL Name
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=14, values=acl_id)
    add_cell(sheet=sheet1, ro=sheet1_cell_start, col=15, values=vpc_id)

    in_count = 0
    for entry in acl['Entries']:
        if entry['Egress'] == True:  # Egress
            in_count += 1

            if 'CidrBlock' in entry:
                source = entry['CidrBlock']
            else:
                source = "-"
            protocol = "All"
            portRange = "All"

            # ruleNumber
            if entry['RuleNumber'] >= 32737:
                ruleNumber = "*"
            else:
                ruleNumber = entry['RuleNumber']

            # RuleAction
            ruleAction = entry['RuleAction']

            # protocol, PortRange
            if entry['Protocol'] == '-1':
                protocol = "All"
            elif entry['Protocol'] == '6':
                protocol = "TCP"
                if entry['PortRange']['To'] == entry['PortRange']['From']:
                    portRange = entry['PortRange']['From']
                else:
                    portRange = str(entry['PortRange']['From']) + " - " + str(entry['PortRange']['To'])
            elif entry['Protocol'] == '17':
                protocol = "UDP"
                if entry['PortRange']['To'] == entry['PortRange']['From']:
                    portRange = entry['PortRange']['From']
                else:
                    portRange = str(entry['PortRange']['From']) + " - " + str(entry['PortRange']['To'])

            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=16, values=ruleNumber)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=17, values=protocol)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=18, values=portRange)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=19, values=source)
            add_cell(sheet=sheet1, ro=sheet1_cell_start, col=20, values=ruleAction.capitalize())

            sheet1_cell_start += 1

    acl_end_row = sheet1_cell_start - 1

    # merge cells
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=12, end_column=12)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=13, end_column=13)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=14, end_column=14)
    sheet1.merge_cells(start_row=acl_start_row, end_row=acl_end_row, start_column=15, end_column=15)

'''
Subnets
'''

sheet1_cell_start = subnet_sheet1_cell_start + 2
sheet1.cell(row=sheet1_cell_start, column=2).value = "<Subnet>"
sheet1.cell(row=sheet1_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)

sheet1_cell_start = sheet1_cell_start + 1

cell_header = ["No.", "Subnet Type", "VPC ID", "Subnet Name", "Subnet ID", "Subnet CIDR Block", "Availability Zone",
               "Network ACLs", "Route Tables"]
for col_index, header in enumerate(cell_header):
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).value = header
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet1.cell(row=sheet1_cell_start, column=col_index + 2).border = thin_border

# Public Subnet을 pubSubnetDict에 넣음
pubSubnetDict = {}

for routeTable in ec2_cli.describe_route_tables()['RouteTables']:
    associations = routeTable['Associations']
    routes = routeTable['Routes']
    isPublic = False

    for route in routes:
        gid = route.get('GatewayId', '')
        if gid.startswith('igw-'):
            isPublic = True

    if (not isPublic):
        continue

    for assoc in associations:
        subnetId = assoc.get('SubnetId', None)  # This checks for explicit associations, only
        if subnetId:
            pubSubnetDict[subnetId] = isPublic

sheet1_cell_start = sheet1_cell_start + 1

for idx, subnet in enumerate(ec2_res.subnets.all()):
    add_cell(sheet1, sheet1_cell_start, 2, idx + 1)  # Number

    # Subnet Type (Check if this subnet is Public or Private)
    if subnet.id in pubSubnetDict:
        add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values="Public Network")
    else:
        add_cell(sheet=sheet1, ro=sheet1_cell_start, col=3, values="Private Network")

    add_cell(sheet1, sheet1_cell_start, 4, subnet.vpc_id)  # VPC Id

    # Subnet Name
    try:
        for tags in subnet.tags:
            if tags["Key"] == 'Name':
                add_cell(sheet1, sheet1_cell_start, 5, tags["Value"])
    except Exception as e:
        add_cell(sheet1, sheet1_cell_start, 5, '-')

    add_cell(sheet1, sheet1_cell_start, 6, subnet.subnet_id)  # Subnet Id
    add_cell(sheet1, sheet1_cell_start, 7, subnet.cidr_block)  # Subnet CIDR Block
    add_cell(sheet1, sheet1_cell_start, 8, subnet.availability_zone)  # Subnet Availability Zone

    # NetworkAcls
    for acls in ec2_cli.describe_network_acls()['NetworkAcls']:
        for i in acls['Associations']:
            if i.get('SubnetId') == subnet.subnet_id:
                add_cell(sheet1, sheet1_cell_start, 9, i.get('NetworkAclId'))

    # Route Tables
    try:
        route_table = ec2_cli.describe_route_tables(
            Filters=[
                {
                    'Name': 'association.subnet-id',
                    'Values': [
                        subnet.subnet_id,
                    ]
                },
            ],
        )['RouteTables']
        add_cell(sheet1, sheet1_cell_start, 10, route_table[0].get('Associations')[0].get('RouteTableId'))
    except:
        add_cell(sheet1, sheet1_cell_start, 10, '-')

    sheet1_cell_start = sheet1_cell_start + 1

'''
Route Table
'''

sheet3 = wb.create_sheet('sheet3')
sheet3.title = 'Route Table'

cell_widths = [5, 5, 25, 23, 22, 25, 16, 20, 22, 15, 20, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7]
sheet_cell_width(sheet3, cell_widths)

title = sheet3['B2']
title.value = "<Route Table>"
title.font = Font(name='맑은 고딕', size=12, bold=True)

sheet3_cell_start = 3

cell_header = ["No.", "Name", "ID", "Destination", "Target"]

for col_index, header in enumerate(cell_header):
    sheet3.cell(row=sheet3_cell_start, column=col_index + 2).value = header
    sheet3.cell(row=sheet3_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet3.cell(row=sheet3_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet3.cell(row=sheet3_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet3.cell(row=sheet3_cell_start, column=col_index + 2).border = thin_border
sheet3_cell_start = sheet3_cell_start + 1

for idx, route in enumerate(ec2_res.route_tables.all()):
    long_start = sheet3_cell_start
    short_start = sheet3_cell_start
    route_table = ec2_res.RouteTable(route.id)
    route_asso = route_table.associations_attribute

    add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

    try:
        add_cell(sheet3, sheet3_cell_start, 3, route_table.tags[0].get('Value'))
    except:
        add_cell(sheet3, sheet3_cell_start, 3, '-')

    add_cell(sheet3, sheet3_cell_start, 4, route_table.id)

    for idx, info in enumerate(route_table.routes_attribute):
        add_cell(sheet3, long_start, 5, info.get('DestinationCidrBlock'))
        if info.get('GatewayId'):
            add_cell(sheet3, long_start, 6, info.get('GatewayId'))
        elif info.get('NatGatewayId'):
            add_cell(sheet3, long_start, 6, info.get('NatGatewayId'))
        elif info.get('TransitGatewayId'):
            add_cell(sheet3, long_start, 6, info.get('TransitGatewayId'))
        elif info.get('LocalGatewayId'):
            add_cell(sheet3, long_start, 6, info.get('LocalGatewayId'))
        else:
            add_cell(sheet3, long_start, 6, info.get('NetworkInterfaceId'))
        long_start = long_start + 1

    # cell merge
    sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=2, end_column=2)
    sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=3, end_column=3)
    sheet3.merge_cells(start_row=short_start, end_row=long_start - 1, start_column=4, end_column=4)

    sheet3_cell_start = long_start



'''
Peering Connection
'''

response = ec2_cli.describe_vpc_peering_connections()['VpcPeeringConnections']
if len(response) != 0:

    sheet3_cell_start = sheet3_cell_start + 2
    sheet3.cell(row=sheet3_cell_start, column=2).value = "<Peering Connections>"
    sheet3.cell(row=sheet3_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)
    sheet3_cell_start = sheet3_cell_start + 1

    cell_header = ["No.", "Name", "Status", "Peering Connection ID", "Requester VPC ID", "Requester Region", "Requester CIDR Block", "Accepter VPC ID", "Accepter Region", "Accepter CIDR Block"]
    for col_index, header in enumerate(cell_header):
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).value = header
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).border = thin_border
    sheet3_cell_start = sheet3_cell_start + 1

    for idx, peering in enumerate(response):
        # No.
        add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

        # Name
        peer_name = '-'
        try:
            for i in peering.get('Tags'):
                if i.get('Key') == 'Name':
                    peer_name = i.get('Value')
        except Exception:
            pass
        add_cell(sheet3, sheet3_cell_start, 3, peer_name)

        # Status, Accepter CIDR Block
        # Statue가 Active가 아닐 경우 Accepter의 CIDR Block을 알아낼 수 없음
        peer_status = peering['Status'].get('Code')
        if peer_status == 'active':
            add_cell(sheet3, sheet3_cell_start, 4, peer_status)

            # Accepter CIDR Block
            accepter_cidr = peering['AccepterVpcInfo'].get('CidrBlock')
            add_cell(sheet3, sheet3_cell_start, 11 , accepter_cidr)

        else:
            add_cell(sheet3, sheet3_cell_start, 4, peer_status)
            add_cell(sheet3, sheet3_cell_start, 11 , '-')

        # Peering Connection ID
        peer_id = peering.get('VpcPeeringConnectionId')
        add_cell(sheet3, sheet3_cell_start, 5, peer_id)

        # Requester
        requester = peering.get('RequesterVpcInfo')

        # Requester VPC ID
        req_id = requester.get('VpcId')
        add_cell(sheet3, sheet3_cell_start, 6, req_id)

        # Requester Region
        req_region = requester.get('Region')
        add_cell(sheet3, sheet3_cell_start, 7, req_region)

        # Requester CIDR Block
        req_cidr = requester.get('CidrBlock')
        add_cell(sheet3, sheet3_cell_start, 8, req_cidr)

        # Accepter
        accepter = peering.get('AccepterVpcInfo')

        # Accepter VPC ID
        accep_id = accepter.get('VpcId')
        add_cell(sheet3, sheet3_cell_start, 9, accep_id)

        # Accepter Region
        accep_region = accepter.get('Region')
        add_cell(sheet3, sheet3_cell_start, 10, accep_region)

        sheet3_cell_start = sheet3_cell_start + 1


else:
    print('There is no Peering Connections')

'''
Transit Gateway
'''

response = ec2_cli.describe_transit_gateways()['TransitGateways']
if len(response) != 0:

    sheet3_cell_start = sheet3_cell_start + 2
    sheet3.cell(row=sheet3_cell_start, column=2).value = "<Transit Gateway>"
    sheet3.cell(row=sheet3_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)
    sheet3_cell_start = sheet3_cell_start + 1

    cell_header = ["No.", "Name", "State", "Transit Gateway ID", "Creation Date", "CIDR Blocks"]
    for col_index, header in enumerate(cell_header):
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).value = header
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet3.cell(row=sheet3_cell_start, column=col_index + 2).border = thin_border
    sheet3_cell_start = sheet3_cell_start + 1

    for idx, tgw in enumerate(response):
        # No.
        add_cell(sheet3, sheet3_cell_start, 2, idx + 1)

        # Name
        tgw_name = '-'
        try:
            for i in tgw.get('Tags'):
                if i.get('Key') == 'Name':
                    tgw_name = i.get('Value')
        except Exception:
            pass
        add_cell(sheet3, sheet3_cell_start, 3, tgw_name)

        # State
        tgw_state = tgw.get('State')
        add_cell(sheet3, sheet3_cell_start, 4, tgw_state)

        # Transit Gateway ID
        tgw_id = tgw.get('TransitGatewayId')
        add_cell(sheet3, sheet3_cell_start, 5, tgw_id)

        # Creation Date
        tgw_created_date = str(tgw.get('CreationTime'))
        tgw_date = tgw_created_date.split(' ')[0]
        tgw_time = tgw_created_date.split(' ')[1]
        add_cell(sheet3, sheet3_cell_start, 6, tgw_date + ',  ' + tgw_time)

        
        # CIDR Blocks
        try:
            tgw_cidr = ''
            for i, cidr in enumerate(tgw['Options']['TransitGatewayCidrBlocks']):
                if i != 0:
                    tgw_cidr += ', \n'
                tgw_cidr += cidr
            add_cell(sheet3, sheet3_cell_start, 7, tgw_cidr)
        except Exception as e:
            add_cell(sheet3, sheet3_cell_start, 7, '-')

        sheet3_cell_start = sheet3_cell_start + 1

else:
    print('There is no Transit Gateway')



# sheet 4
# Nat GW, IGW

'''
Nat Gateway
'''

sheet4 = wb.create_sheet('sheet4')
sheet4.title = 'NAT GW & IGW'

cell_widths = [5, 5, 27, 22, 22, 22, 24, 20, 20, 20, 20, 20, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7]
sheet_cell_width(sheet4, cell_widths)

title = sheet4['B2']
title.value = "<Nat Gateways>"
title.font = Font(name='맑은 고딕', size=12, bold=True)
# title.alignment = Alignment(horizontal='center', vertical='center')


sheet4_cell_start = 3
cell_header = ["No.", "Name", "ID", "Elastic IP", "VPC", "Subnet", "Status"]

for col_index, header in enumerate(cell_header):
    sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
    sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
    sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border

sheet4_cell_start = sheet4_cell_start + 1

'''
Nat Gateway
'''

if ec2_cli.describe_nat_gateways()['NatGateways']:

    for idx, nat in enumerate(ec2_cli.describe_nat_gateways()['NatGateways']):
        add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
        natName = '-'
        try:
            for i in nat.get('Tags'):
                if i.get('Key') == 'Name':
                    natName = i.get('Value')
        except Exception:
            pass

        id = nat.get('NatGatewayId')
        eip = nat.get('NatGatewayAddresses')[0].get('PublicIp')
        vpc = nat.get('VpcId')
        subnet = nat.get('SubnetId')
        state = nat.get('State')

        add_cell(sheet4, sheet4_cell_start, 3, natName)
        add_cell(sheet4, sheet4_cell_start, 4, id)
        add_cell(sheet4, sheet4_cell_start, 5, eip)
        add_cell(sheet4, sheet4_cell_start, 6, vpc)
        add_cell(sheet4, sheet4_cell_start, 7, subnet)
        add_cell(sheet4, sheet4_cell_start, 8, state.capitalize())

        sheet4_cell_start = sheet4_cell_start + 1
else:
    print('There is no NAT Gateway.')


'''
IGW
'''

if ec2_cli.describe_internet_gateways():

    sheet4_cell_start = sheet4_cell_start + 2

    sheet4.cell(row=sheet4_cell_start, column=2).value = "<Internet Gateways>"
    sheet4.cell(row=sheet4_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)

    sheet4_cell_start = sheet4_cell_start + 1

    cell_header = ["No.", "Name", "ID", "VPC", "Status"]

    for col_index, header in enumerate(cell_header):
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border

    sheet4_cell_start = sheet4_cell_start + 1

    for idx, igw in enumerate(ec2_cli.describe_internet_gateways()['InternetGateways']):
        # No.
        add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
        # Name 가져옴, 없을 시 '-'
        igwName = '-'
        try:
            for i in igw.get('Tags'):
                if i.get('Key') == 'Name':
                    igwName = i.get('Value')
        except Exception:
            pass
        # igw ID, vpc ID, Status
        igwId = igw.get('InternetGatewayId')
        if len(igw['Attachments']) != 0:
            vpcId = igw['Attachments'][0].get('VpcId')
            status = igw['Attachments'][0].get('State')
        else:
            vpcId = '-'
            status = 'Detached'

        # Name
        add_cell(sheet4, sheet4_cell_start, 3, igwName)
        # IGW ID
        add_cell(sheet4, sheet4_cell_start, 4, igwId)
        # VPC ID
        add_cell(sheet4, sheet4_cell_start, 5, vpcId)
        # Status
        add_cell(sheet4, sheet4_cell_start, 6, status.capitalize())

        sheet4_cell_start = sheet4_cell_start + 1

else:
    print('There is no Internet Gateway.')


'''
VPN Connections
'''
response = ec2_cli.describe_vpn_connections()['VpnConnections']

if len(response) != 0:
    sheet4_cell_start = sheet4_cell_start + 2

    sheet4.cell(row=sheet4_cell_start, column=2).value = "<VPN Connection>"
    sheet4.cell(row=sheet4_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)

    sheet4_cell_start = sheet4_cell_start + 1
    cell_header = ["No.", "Name", "VPN ID", "State", "Virtual Private Gateway", "Transit Gateway", "Customer Gateway", "Routing", "Type", "Local IPv4 CIDR", "Remote IPv4 CIDR"]

    for col_index, header in enumerate(cell_header):
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border
    sheet4_cell_start = sheet4_cell_start + 1

    for idx, vpn in enumerate(response):
        # No.
        add_cell(sheet4, sheet4_cell_start, 2, idx + 1)
        # Name
        vpn_name = '-'
        try:
            for i in vpn.get('Tags'):
                if i.get('Key') == 'Name':
                    vpn_name = i.get('Value')
        except Exception:
            pass
        add_cell(sheet4, sheet4_cell_start, 3, vpn_name)

        # VPN ID
        vpn_id = vpn.get('VpnGatewayId')
        add_cell(sheet4, sheet4_cell_start, 4, vpn_id)

        # State
        vpn_state = vpn.get('State')
        add_cell(sheet4, sheet4_cell_start, 5, vpn_state.capitalize())

        # Virtual Private Gateway ID
        try:
            vgw_id = vpn.get('VpnGatewayId')
            add_cell(sheet4, sheet4_cell_start, 6, vgw_id)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 6, '-')

        # Transit Gateway ID
        try:
            tgw_id = vpn.get('TransitGatewayId')
            add_cell(sheet4, sheet4_cell_start, 7, tgw_id)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 7, '-')

        # Customer Gateway ID
        try:
            cgw_id = vpn.get('CustomerGatewayId')
            add_cell(sheet4, sheet4_cell_start, 8, cgw_id)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 8, '-')

        # Routing
        routing = vpn['Options']['StaticRoutesOnly']
        if routing == False:
            add_cell(sheet4, sheet4_cell_start, 9, 'Dynamic')
        else:
            add_cell(sheet4, sheet4_cell_start, 9, 'Static')

        # Type
        vpn_type = vpn.get('Type')
        add_cell(sheet4, sheet4_cell_start, 10, vpn_type)

        # Local IPv4 CIDR
        local_cidr = vpn['Options']['LocalIpv4NetworkCidr']
        add_cell(sheet4, sheet4_cell_start, 11, local_cidr)

        # Remote IPV4 CIDR
        remote_cidr = vpn['Options']['RemoteIpv4NetworkCidr']
        add_cell(sheet4, sheet4_cell_start, 12, remote_cidr)

        sheet4_cell_start = sheet4_cell_start + 1
else:
    print('There is no VPN Connections.')


'''
Virtual Private Gateway
'''

response = ec2_cli.describe_vpn_gateways()['VpnGateways']
if len(response) != 0 :

    sheet4_cell_start = sheet4_cell_start + 2

    sheet4.cell(row=sheet4_cell_start, column=2).value = "<Virtual Private Gateway>"
    sheet4.cell(row=sheet4_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)

    sheet4_cell_start = sheet4_cell_start + 1

    cell_header = ["No.", "Name", "ID", "State", "Type", "VPC", "Amazon Side SAN"]

    for col_index, header in enumerate(cell_header):
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border

    sheet4_cell_start = sheet4_cell_start + 1

    for idx, vgw in enumerate(response):
        # No.
        add_cell(sheet4, sheet4_cell_start, 2, idx + 1)

        # Name
        vgw_name = '-'
        try:
            for i in vgw.get('Tags'):
                if i.get('Key') == 'Name':
                    vgw_name = i.get('Value')
            add_cell(sheet4, sheet4_cell_start, 3, vgw_name)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 3, vgw_name)

        # ID
        vgw_id = vgw.get('VpnGatewayId')
        add_cell(sheet4, sheet4_cell_start, 4, vgw_id)

        # State
        vgw_state = vgw.get('State')
        add_cell(sheet4, sheet4_cell_start, 5, str(vgw_state).capitalize())

        # Type
        vgw_type = vgw.get('Type')
        add_cell(sheet4, sheet4_cell_start, 6, vgw_type)

        # VPC
        # TODO 여기 작성해야함

        # Amazon Side SAN
        try:
            vgw_san = vgw.get('AmazonSideAsn')
            add_cell(sheet4, sheet4_cell_start, 8, vgw_san)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 8, '-')

        sheet4_cell_start = sheet4_cell_start + 1

else:
    print('There is no Virtual Private Gateways.')


'''
Customer Gateway
'''

response = ec2_cli.describe_customer_gateways()['CustomerGateways']
if len(response) != 0:
    sheet4_cell_start = sheet4_cell_start + 2

    sheet4.cell(row=sheet4_cell_start, column=2).value = "<Customer Gateway>"
    sheet4.cell(row=sheet4_cell_start, column=2).font = Font(name='맑은 고딕', size=12, bold=True)
    sheet4_cell_start = sheet4_cell_start + 1

    cell_header = ["No.", "Name", "ID", "State", "Type", "IP Address", "BGP ASN", "Device Name"]

    for col_index, header in enumerate(cell_header):
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).value = header
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet4.cell(row=sheet4_cell_start, column=col_index + 2).border = thin_border
    sheet4_cell_start = sheet4_cell_start + 1

    for idx, cgw in enumerate(response):

        # No.
        add_cell(sheet4, sheet4_cell_start, 2, idx + 1)

        # Name
        cgw_name = '-'
        try:
            for i in vgw.get('Tags'):
                if i.get('Key') == 'Name':
                    cgw_name = i.get('Value')
            add_cell(sheet4, sheet4_cell_start, 3, cgw_name)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 3, cgw_name)

        # CGW ID
        cgw_id = cgw.get('CustomerGatewayId')
        add_cell(sheet4, sheet4_cell_start, 4, cgw_id)

        # State
        cgw_state = cgw.get('State')
        add_cell(sheet4, sheet4_cell_start, 5, str(cgw_state).capitalize())

        # Type
        cgw_type = cgw.get('Type')
        add_cell(sheet4, sheet4_cell_start, 6, cgw_type)

        # IP Address
        cgw_ip = cgw.get('IpAddress')
        add_cell(sheet4, sheet4_cell_start, 7, cgw_id)

        # BGP SAN
        bgw_asn = cgw.get('BgpAsn')
        add_cell(sheet4, sheet4_cell_start, 8, bgw_asn)

        # Device Name
        try:
            device_name = cgw['DeviceName']
            add_cell(sheet4, sheet4_cell_start, 9, device_name)
        except Exception:
            add_cell(sheet4, sheet4_cell_start, 9, '-')

else:
    print('There is no Customet Gateways.')



# sheet5

sheet5 = wb.create_sheet('sheet5')
sheet5.title = 'Security Group'

title = sheet5['B2']
title.value = "<Inbound>"
title.font = Font(name='맑은 고딕', size=12, bold=True)
# title.alignment = Alignment(horizontal='center', vertical='center')

cell_widths = [5, 5, 55, 20, 22, 13, 24, 45, 7, 7.8, 55, 23, 22, 13, 20, 30, 10, 11, 7, 7, 7, 7]
sheet_cell_width(sheet5, cell_widths)

'''
Security Group
'''

sheet5_cell_start = 3
cell_header1 = ["No.", "Security Groups Name", "Group ID", "Inbound Rule", "Inbound Rule", "Inbound Rule",
                "비고(Description)"]
cell_header2 = ["No.", "Security Groups Name", "Group ID", "Type", "Port Range", "source", "비고(Description)"]

for col_index, header in enumerate(cell_header1):
    add_cell(sheet5, sheet5_cell_start, col_index + 2, header)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                fgColor=Color('E3E3E3'))
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).border = thin_border
sheet5_cell_start = sheet5_cell_start + 1

for col_index, header in enumerate(cell_header2):
    add_cell(sheet5, sheet5_cell_start, col_index + 2, header)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                fgColor=Color('E3E3E3'))
    sheet5.cell(row=sheet5_cell_start, column=col_index + 2).border = thin_border
sheet5_cell_start = sheet5_cell_start + 1

sheet5.merge_cells(start_row=3, end_row=4, start_column=2, end_column=2)
sheet5.merge_cells(start_row=3, end_row=4, start_column=3, end_column=3)
sheet5.merge_cells(start_row=3, end_row=4, start_column=4, end_column=4)
sheet5.merge_cells(start_row=3, end_row=3, start_column=5, end_column=7)
sheet5.merge_cells(start_row=3, end_row=4, start_column=8, end_column=8)

previous = 4
flag = 0

# Inbound Rule
for idx, security_group in enumerate(ec2_res.security_groups.all()):
    short_start = sheet5_cell_start
    long_start1 = sheet5_cell_start
    short_start1 = sheet5_cell_start
    long_start2 = sheet5_cell_start
    short_start2 = sheet5_cell_start
    sec_group = ec2_res.SecurityGroup(security_group.id)

    # No.
    add_cell(sheet5, sheet5_cell_start, 2, idx + 1)
    # Security Group Name
    add_cell(sheet5, sheet5_cell_start, 3, sec_group.group_name)
    # Security Group Id
    add_cell(sheet5, sheet5_cell_start, 4, sec_group.group_id)

    # inbound
    # Type, Port Range, Source, Description
    for inbound in sec_group.ip_permissions:
        ip_range = []
        desc = []

        # Type
        if inbound.get('IpProtocol') == '-1':
            ipType = 'ALL Traffic'
        else:
            ipType = inbound.get('IpProtocol')

        # portRange
        if inbound.get('FromPort') == inbound.get('ToPort'):
            portrange = inbound.get('FromPort')
        else:
            portrange = str(inbound.get('FromPort')) + " - " + str(inbound.get('ToPort'))

        if portrange == '0--1' or portrange == -1:
            portrange = 'N/A'

        if ipType == "ALL Traffic":
            portrange = "All"

        ipType = TypeCheck(ipType, portrange)

        # source
        for ips in inbound.get('IpRanges'):
            ip_range.append(ips.get('CidrIp'))
            description = ips.get('Description')
            if description == None:
                desc.append("-")
            else:
                desc.append(description)

        # 비고
        for ips in inbound.get('Ipv6Ranges'):
            ip_range.append(ips.get('CidrIpv6'))
            description = ips.get('Description')
            if description == None:
                desc.append("-")
            else:
                desc.append(description)

        for group in inbound.get('UserIdGroupPairs'):
            ip_range.append(group.get('GroupId'))
            description = group.get('Description')
            if description == None:
                desc.append("-")
            else:
                desc.append(description)

        add_cell(sheet5, short_start1, 5, ipType)
        add_cell(sheet5, short_start1, 6, portrange)

        tmp1 = long_start1

        for ip, desc in zip(ip_range, desc):
            add_cell(sheet5, long_start1, 7, ip)
            add_cell(sheet5, long_start1, 8, desc)
            long_start1 += 1

        tmp1_1 = long_start1

        sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=5,
                           end_column=5)
        sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=6,
                           end_column=6)

        short_start1 = long_start1

    try:
        if long_start1 >= long_start2:
            sheet5_cell_start = long_start1
            # cell merge
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=2,
                               end_column=2)
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=3,
                               end_column=3)
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=4,
                               end_column=4)
        else:
            sheet5_cell_start = long_start2
            # cell merge

            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=2,
                               end_column=2)
            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=3,
                               end_column=3)
            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=4,
                               end_column=4)
    except:
        pass

##

title = sheet5['J2']
title.value = "<Outbound>"
title.font = Font(name='맑은 고딕', size=12, bold=True)
# title.alignment = Alignment(horizontal='center', vertical='center')

sheet5_cell_start = 3
cell_header1 = ["No.", "Security Groups Name", "Group ID", "Outbound Rule", "Outbound Rule", "Outbound Rule",
                "비고(Description)"]
cell_header2 = ["No.", "Security Groups Name", "Group ID", "Type", "Port Range", "source", "비고(Description)"]

for col_index, header in enumerate(cell_header1):
    add_cell(sheet5, sheet5_cell_start, col_index + 10, header)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).fill = PatternFill(patternType='solid',
                                                                                 fgColor=Color('E3E3E3'))
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).border = thin_border
sheet5_cell_start = sheet5_cell_start + 1

for col_index, header in enumerate(cell_header2):
    add_cell(sheet5, sheet5_cell_start, col_index + 10, header)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).font = Font(name='맑은 고딕', size=10, bold=True)
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).fill = PatternFill(patternType='solid',
                                                                                 fgColor=Color('E3E3E3'))
    sheet5.cell(row=sheet5_cell_start, column=col_index + 10).border = thin_border
sheet5_cell_start = sheet5_cell_start + 1

sheet5.merge_cells(start_row=3, end_row=4, start_column=10, end_column=10)
sheet5.merge_cells(start_row=3, end_row=4, start_column=11, end_column=11)
sheet5.merge_cells(start_row=3, end_row=4, start_column=12, end_column=12)
sheet5.merge_cells(start_row=3, end_row=3, start_column=13, end_column=15)
sheet5.merge_cells(start_row=3, end_row=4, start_column=16, end_column=16)

# outbound

for idx, security_group in enumerate(ec2_res.security_groups.all()):
    add_cell(sheet5, sheet5_cell_start, 10, idx + 1)
    short_start = sheet5_cell_start
    long_start1 = sheet5_cell_start
    short_start1 = sheet5_cell_start
    long_start2 = sheet5_cell_start
    short_start2 = sheet5_cell_start

    sec_group = ec2_res.SecurityGroup(security_group.id)
    add_cell(sheet5, sheet5_cell_start, 11, sec_group.group_name)
    add_cell(sheet5, sheet5_cell_start, 12, sec_group.group_id)

    # inbound
    for outbound in sec_group.ip_permissions_egress:
        if outbound:
            ip_range = []
            desc = []

            # Type
            if outbound.get('IpProtocol') == '-1':
                ipType = 'ALL Traffic'
            else:
                ipType = outbound.get('IpProtocol')

            # portRange
            if outbound.get('FromPort') == outbound.get('ToPort'):
                portrange = outbound.get('FromPort')
            else:
                portrange = str(outbound.get('FromPort')) + " - " + str(outbound.get('ToPort'))

            if portrange == '0--1' or portrange == -1:
                portrange = 'N/A'

            if ipType == "ALL Traffic":
                portrange = "All"

            ipType = TypeCheck(ipType, portrange)

            # source
            for ips in outbound.get('IpRanges'):
                ip_range.append(ips.get('CidrIp'))
                description = ips.get('Description')
                if description == None:
                    desc.append("-")
                else:
                    desc.append(description)

            # 비고
            for ips in outbound.get('Ipv6Ranges'):
                ip_range.append(ips.get('CidrIpv6'))
                description = ips.get('Description')
                if description == None:
                    desc.append("-")
                else:
                    desc.append(description)

            for group in outbound.get('UserIdGroupPairs'):
                ip_range.append(group.get('GroupId'))
                description = group.get('Description')
                if description == None:
                    desc.append("-")
                else:
                    desc.append(description)

            add_cell(sheet5, short_start1, 10, idx + 1)
            add_cell(sheet5, short_start1, 11, sec_group.group_name)
            add_cell(sheet5, short_start1, 12, sec_group.group_id)

            add_cell(sheet5, short_start1, 13, ipType)
            add_cell(sheet5, short_start1, 14, portrange)

            tmp1 = long_start1

            for ip, desc in zip(ip_range, desc):
                add_cell(sheet5, long_start1, 15, ip)
                add_cell(sheet5, long_start1, 16, desc)
                add_cell(sheet5, long_start1, 10, idx + 1)
                add_cell(sheet5, long_start1, 11, sec_group.group_name)
                add_cell(sheet5, long_start1, 12, sec_group.group_id)
                long_start1 += 1

            tmp1_1 = long_start1
            try:
                sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=13,
                                   end_column=13)
                sheet5.merge_cells(start_row=tmp1, end_row=tmp1_1 - 1, start_column=14,
                                   end_column=14)
            except:
                pass
            short_start1 = long_start1

        sheet5_cell_start = long_start1

    try:
        if long_start1 >= long_start2:
            sheet5_cell_start = long_start1
            # cell merge
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=10,
                               end_column=10)
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=11,
                               end_column=11)
            sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=12,
                               end_column=12)
        else:
            sheet5_cell_start = long_start2
            # cell merge

            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=10,
                               end_column=10)
            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=11,
                               end_column=11)
            sheet5.merge_cells(start_row=short_start, end_row=long_start2 - 1, start_column=12,
                               end_column=12)
    except:
        pass

        # # cell merge
        # sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=10, end_column=10)
        # sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=11, end_column=11)
        # sheet5.merge_cells(start_row=short_start, end_row=long_start1 - 1, start_column=12, end_column=12)
    # else:
    #     pass

# sheet6
response = elb_cli.describe_load_balancers()
response2 = elbv2_cli.describe_load_balancers()
if response['LoadBalancerDescriptions'] or response2['LoadBalancers']:

    sheet6 = wb.create_sheet('sheet6')
    sheet6.title = 'ELB'
    cell_widths = [5, 5, 13, 28, 65, 11, 20, 30, 18, 20.5, 25, 22, 14, 25, 30, 12, 12, 12, 7, 7, 7, 7]
    sheet_cell_width(sheet6, cell_widths)

    title = sheet6['B2']
    title.value = "<ELB>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)
    # title.alignment = Alignment(horizontal='center', vertical='center')

    '''
    ELB 9
    '''

    sheet6_cell_start = 3

    cell_header1 = ["필수사항", "", "", "", "", "", "", "", ""]
    for col_index, header in enumerate(cell_header1):
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).value = header
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).border = thin_border

    cell_header2 = ["선택사항", "", "", "", "", "", ""]
    for col_index, header in enumerate(cell_header2):
        sheet6.cell(row=sheet6_cell_start, column=col_index + 11).value = header
        sheet6.cell(row=sheet6_cell_start, column=col_index + 11).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet6.cell(row=sheet6_cell_start, column=col_index + 11).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')
        sheet6.cell(row=sheet6_cell_start, column=col_index + 11).fill = PatternFill(patternType='solid',
                                                                                     fgColor=Color('F4FA58'))
        sheet6.cell(row=sheet6_cell_start, column=col_index + 11).border = thin_border
    sheet6_cell_start = sheet6_cell_start + 1

    cell_header3 = ["No.", "Scheme", "ELB Name", "DNS Name", "Type", "Port Configuration",
                    "Instance IDs or Target Groups", "Availability Zones", "ELB Security Group",
                    "Cross-Zone Load Balancing", "Connection Draining (s)", "Idle Timeout (s)", "Health Check",
                    "Certification", "Stickiness", "Access Logs"]
    for col_index, header in enumerate(cell_header3):
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).value = header
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).font = Font(name='맑은 고딕', size=10, bold=True)
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet6.cell(row=sheet6_cell_start, column=col_index + 2).border = thin_border
    sheet6_cell_start = sheet6_cell_start + 1

    sheet6.merge_cells(start_row=3, end_row=3, start_column=2, end_column=10)
    sheet6.merge_cells(start_row=3, end_row=3, start_column=11, end_column=17)

    num = 0

    # CLB
    for idx, response in enumerate(response['LoadBalancerDescriptions']):
        # No.
        num = idx + 1
        add_cell(sheet6, sheet6_cell_start, 2, idx + 1)
        # Scheme
        schemeType = response['Scheme']
        add_cell(sheet6, sheet6_cell_start, 3, schemeType)
        # ELB Name
        ElbName = response['LoadBalancerName']
        add_cell(sheet6, sheet6_cell_start, 4, ElbName)
        # DNS Name
        DnsName = response['DNSName']
        add_cell(sheet6, sheet6_cell_start, 5, DnsName)
        # Type
        add_cell(sheet6, sheet6_cell_start, 6, "Classic")
        # Port Configuration
        ListenerDesc = response['ListenerDescriptions']
        Listenerstr = ""
        for idx, Listener in enumerate(ListenerDesc):
            if idx != 0:
                Listenerstr += ", \n"
            Listenerstr += str(Listener['Listener'].get('InstancePort')) + ' forwarding to ' + str(
                Listener['Listener'].get('LoadBalancerPort'))
        add_cell(sheet6, sheet6_cell_start, 7, Listenerstr)
        # Instance IDs
        idstr = ""
        if len(response['Instances']) != 0:
            for idx, id in enumerate(response['Instances']):
                if idx != 0:
                    idstr += ", \n"
                idstr += id.get('InstanceId')
            add_cell(sheet6, sheet6_cell_start, 8, idstr)
        else:
            add_cell(sheet6, sheet6_cell_start, 8, '-')
        # Availability Zones
        zonestr = ""
        for idx, zone in enumerate(response['AvailabilityZones']):
            if idx != 0:
                zonestr += ", \n"
            zonestr += zone
        add_cell(sheet6, sheet6_cell_start, 9, zonestr)
        # Security Groups
        secstr = ""
        for idx, sec in enumerate(response['SecurityGroups']):
            if idx != 0:
                secstr += ", \n"
            secstr += sec
        add_cell(sheet6, sheet6_cell_start, 10, secstr)

        # describe_lb_attributes()를 이용하여 CrossZone, AccessLog, ConnectionDraining, ConnectionSettings 얻음
        # 각 clb마다 ElbName을 얻어 response_attr 객체 생성
        response_attr = elb_cli.describe_load_balancer_attributes(LoadBalancerName=ElbName)['LoadBalancerAttributes']

        # Cross-Zone Load Balancing
        crossZone = response_attr['CrossZoneLoadBalancing']['Enabled']
        if crossZone == True:
            add_cell(sheet6, sheet6_cell_start, 11, "Enabled")
        else:
            add_cell(sheet6, sheet6_cell_start, 11, "Disabled")

        # Connection Draining
        conDraining = response_attr['ConnectionDraining']
        if conDraining['Enabled'] == True:
            conDrainingTime = conDraining['Timeout']
            add_cell(sheet6, sheet6_cell_start, 12, 'Enabled, ' + str(conDrainingTime))
        else:
            add_cell(sheet6, sheet6_cell_start, 12, '-')

        # Idle Timeout
        idleTimeout = response_attr['ConnectionSettings']['IdleTimeout']
        add_cell(sheet6, sheet6_cell_start, 13, int(idleTimeout))

        # Health Checks
        healthCheck = response['HealthCheck']
        target = healthCheck['Target']
        interval = healthCheck['Interval']
        timeout = healthCheck['Timeout']
        unhealthyThres = healthCheck['UnhealthyThreshold']
        healthyThres = healthCheck['HealthyThreshold']

        add_cell(sheet6, sheet6_cell_start, 14, 'Target : ' + str(target) + '\nInterval : ' + str(interval) +
                 'seconds \nTimeout : ' + str(timeout) + 'seconds \nUnhealthy Threshold : ' + str(unhealthyThres) +
                 '\nHealthy Threshold : ' + str(healthyThres))

        # Certificate
        ListenerDesc = response['ListenerDescriptions']
        certIds = ""

        for idx, Listener in enumerate(ListenerDesc):
            if "SSLCertificateId" in Listener['Listener']:
                certIds = certIds + Listener['Listener'].get('SSLCertificateId') + ", \n"

        if len(certIds) != 0:
            add_cell(sheet6, sheet6_cell_start, 15, certIds)
        else:
            add_cell(sheet6, sheet6_cell_start, 15, "-")

        # Stickiness
        appCookies = response['Policies']['AppCookieStickinessPolicies']
        lbCookies = response['Policies']['LBCookieStickinessPolicies']

        if len(appCookies) == 0 and len(lbCookies):  # ApplicationCookieStickiness, LBCookieStickiness 둘 다 없을 경우
            add_cell(sheet6, sheet6_cell_start, 16, "Disabled")
        else:
            add_cell(sheet6, sheet6_cell_start, 16, "Enabled")

        # Access Logs
        accessLog = response_attr['AccessLog']['Enabled']
        if accessLog == True:
            add_cell(sheet6, sheet6_cell_start, 17, "Enabled")
        else:
            add_cell(sheet6, sheet6_cell_start, 17, "Disabled")

        sheet6_cell_start = sheet6_cell_start + 1

    # ALB, NLB, GAW
    for idx, response in enumerate(response2['LoadBalancers']):

        # No.
        num = num + 1
        add_cell(sheet6, sheet6_cell_start, 2, num)
        # Scheme
        schemeType = ""
        if 'Scheme' in response:
            schemeType = response['Scheme']
        if len(schemeType) != 0:
            add_cell(sheet6, sheet6_cell_start, 3, schemeType)
        else:
            add_cell(sheet6, sheet6_cell_start, 3, "-")
        # ELB Name
        ElbName = response['LoadBalancerName']
        add_cell(sheet6, sheet6_cell_start, 4, ElbName)
        # DNS Name
        DnsName = ""
        if 'DNSName' in response:
            DnsName = response['DNSName']
        if len(DnsName) != 0:
            add_cell(sheet6, sheet6_cell_start, 5, DnsName)
        else:
            add_cell(sheet6, sheet6_cell_start, 5, "-")
        # Type
        lbType = str(response['Type'])
        add_cell(sheet6, sheet6_cell_start, 6, lbType.capitalize())
        # if lbType == "network":
        #     add_cell(sheet6, sheet6_cell_start, 6, "Network")
        # elif lbType == "application":
        #     add_cell(sheet6, sheet6_cell_start, 6, "Application")
        # else:
        #     add_cell(sheet6, sheet6_cell_start, 6, "Gateway")

        # Port Configuration
        # Instance IDs
        targetgroup = elbv2_cli.describe_target_groups(LoadBalancerArn=response['LoadBalancerArn'])
        portstr = ""
        targetstr = ""
        try:
            for idx, port in enumerate(targetgroup['TargetGroups']):
                if idx != 0:
                    portstr += ", \n"
                    targetstr += ", \n"
                portstr += str(port.get('Port')) + " " + port.get('Protocol')
                targetstr += port.get('TargetGroupName')
            add_cell(sheet6, sheet6_cell_start, 7, portstr)
            add_cell(sheet6, sheet6_cell_start, 8, targetstr)

        except:
            add_cell(sheet6, sheet6_cell_start, 7, '-')
            add_cell(sheet6, sheet6_cell_start, 8, '-')

        # Availability Zones
        zonestr = ""
        for idx, zone in enumerate(response['AvailabilityZones']):
            if idx != 0:
                zonestr += ", \n"
            zonestr += zone.get('ZoneName')
        add_cell(sheet6, sheet6_cell_start, 9, zonestr)
        # Security Groups
        secstr = ""
        try:
            for idx, sec in enumerate(response['SecurityGroups']):
                if idx != 0:
                    secstr += ", \n"
                secstr += sec
                add_cell(sheet6, sheet6_cell_start, 10, secstr)
        except:
            add_cell(sheet6, sheet6_cell_start, 10, '-')

        # Attributes
        Attributes = elbv2_cli.describe_load_balancer_attributes(LoadBalancerArn=response['LoadBalancerArn'])[
            'Attributes']

        # Cross-Zone Load Balancing
        # Access Logs
        # Idle Timeout

        # Gateway LB : CorssZone
        if response['Type'] == 'gateway':
            accessLogsEnabled = '-'
            crossZoneLBEnabled = Attributes[1]['Value']
            idleTimeout = '-'

        # Network LB : AccessLogs, CrossZone
        elif response['Type'] == 'network':
            accessLogsEnabled = Attributes[0]['Value']
            crossZoneLBEnabled = Attributes[4]['Value']
            idleTimeout = '-'

        # Application LB : AccessLogs, CrossZone, Idle Timeout
        else:
            accessLogsEnabled = Attributes[0]['Value']
            crossZoneLBEnabled = "Enabled"  # 기본으로 활성화되어 있음
            idleTimeout = Attributes[3]['Value']

        # CrossZone
        if crossZoneLBEnabled == "true" or crossZoneLBEnabled == "Enabled":
            add_cell(sheet6, sheet6_cell_start, 11, "Enabled")
        else:
            add_cell(sheet6, sheet6_cell_start, 11, "Disabled")

        # IdleTimeout
        add_cell(sheet6, sheet6_cell_start, 13, idleTimeout)

        # Access Logs
        if accessLogsEnabled == "true":
            add_cell(sheet6, sheet6_cell_start, 17, "Enabled")
        elif accessLogsEnabled == "false":
            add_cell(sheet6, sheet6_cell_start, 17, "Disabled")
        else:
            add_cell(sheet6, sheet6_cell_start, 17, '-')

        # Connection Draining
        add_cell(sheet6, sheet6_cell_start, 12, '')

        # Health Checks
        add_cell(sheet6, sheet6_cell_start, 14, '')

        # Certificate
        add_cell(sheet6, sheet6_cell_start, 15, '')

        # Stickiness
        add_cell(sheet6, sheet6_cell_start, 16, '')

        sheet6_cell_start = sheet6_cell_start + 1

    '''
    ELB
    '''
else:
    print('There is no ELB.')


# Elastic IP 리스트 뽑아냄
addr_list = ec2_cli.describe_addresses().get('Addresses')
eip_list = []
try:
    for eip in addr_list:
        eip_list.append(eip.get('PublicIp'))
except Exception:
    pass

response = ec2_cli.describe_instances()
if response['Reservations']:

    # sheet7
    sheet7 = wb.create_sheet('sheet7')
    sheet7.title = 'EC2'
    cell_widths = [5, 5, 14, 35, 8, 8, 12, 11, 20, 23, 19, 12, 12, 13, 21, 15, 20, 38, 15, 21, 15, 10]
    sheet_cell_width(sheet7, cell_widths)

    title = sheet7['B2']
    title.value = "<EC2>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' EC2
    '''
    sheet7_cell_start = 3
    cell_header = ["No.", "Availability Zone", "Instance Name", "Status", "AMI (OS)", "Instance Type", "Subnet Type",
                   "VPC ID", "Subnet ID", "Instance ID",
                   "Private IP", "Public IP", "Elstic IP", "Root Volume ID", "Root Volume (GB)", "Key Pair",
                   "Security Groups", "IAM role", "Data Volume ID", "Data Volume (GB)"]
    for col_index, header in enumerate(cell_header):
        add_cell(sheet7, sheet7_cell_start, col_index + 2, header)
        sheet7.cell(row=sheet7_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet7.cell(row=sheet7_cell_start, column=col_index + 2).border = thin_border

    sheet7_cell_start = sheet7_cell_start + 1

    for idx, reservation in enumerate(response['Reservations']):
        for Instances in reservation['Instances']:
            # No.
            add_cell(sheet7, sheet7_cell_start, 2, idx + 1)
            # Availability Zone
            Zone = Instances['Placement'].get('AvailabilityZone')
            add_cell(sheet7, sheet7_cell_start, 3, Zone)
            # Instance Name
            InstanceName = '-'
            try:
                for i in Instances['Tags']:
                    if i.get('Key') == 'Name':
                        InstanceName = i.get('Value')
                        break

                add_cell(sheet7, sheet7_cell_start, 4, InstanceName)
            except:
                pass
            # Status
            Status = Instances['State'].get('Name')
            add_cell(sheet7, sheet7_cell_start, 5, Status)

            # AMI (OS)
            try:
                OS = Instances['Platform']
                add_cell(sheet7, sheet7_cell_start, 6, OS)
            except:
                add_cell(sheet7, sheet7_cell_start, 6, 'Linux')
            # Instance Type
            Type = Instances['InstanceType']
            add_cell(sheet7, sheet7_cell_start, 7, Type)

            # Subnet Type
            if Instances.get('PublicIpAddress'):
                add_cell(sheet7, sheet7_cell_start, 8, 'Public')
            else:
                add_cell(sheet7, sheet7_cell_start, 8, 'Private')
            # VPC ID
            vpc_id = Instances['VpcId']
            add_cell(sheet7, sheet7_cell_start, 9, vpc_id)
            # Subnet ID
            Subnet_ID = Instances['SubnetId']
            add_cell(sheet7, sheet7_cell_start, 10, Subnet_ID)
            # Instance ID
            Instance_ID = Instances['InstanceId']
            add_cell(sheet7, sheet7_cell_start, 11, Instance_ID)
            # Private IP
            Private_IP = Instances['PrivateIpAddress']
            add_cell(sheet7, sheet7_cell_start, 12, Private_IP)
            # Public IP
            if Instances.get('PublicIpAddress'):
                Public_IP = Instances['PublicIpAddress']
                add_cell(sheet7, sheet7_cell_start, 13, Public_IP)
            else:
                Public_IP = ''
                add_cell(sheet7, sheet7_cell_start, 13, Public_IP)
            # EIP
            if Public_IP in eip_list:
                add_cell(sheet7, sheet7_cell_start, 14, Public_IP)
                Public_IP = ''
            else:
                add_cell(sheet7, sheet7_cell_start, 14, '-')

            # data volume 리스트
            data_volume_list = []

            # Root Volume ID
            root_volume_id = ''
            for ebs in Instances['BlockDeviceMappings']:
                data_volume_list.append(ebs['Ebs'].get('VolumeId'))  # EC2 인스턴스에 있는 모든 volume을 추가

                if ebs.get('DeviceName') == Instances['RootDeviceName']:
                    data_volume_list.remove(ebs['Ebs'].get('VolumeId'))  # Root Volume은 리스트에서 제외
                    root_volume_id = str(ebs['Ebs'].get('VolumeId'))
                    add_cell(sheet7, sheet7_cell_start, 15, root_volume_id)

            # Root Volume (GB)
            response = ec2_cli.describe_volumes(
                VolumeIds=[
                    root_volume_id,
                ],
            )
            for volume in response['Volumes']:
                Size = volume.get('Size')
                add_cell(sheet7, sheet7_cell_start, 16, Size)

            # Key Pair
            try:
                KeyPair = Instances['KeyName']
                add_cell(sheet7, sheet7_cell_start, 17, KeyPair)
            except:
                pass
            # Security Group
            sec_groups = ""
            for idx, sec in enumerate(Instances['SecurityGroups']):
                if idx != 0:
                    sec_groups += ', \n'
                sec_groups += sec.get('GroupName')
                # SG = sec.get('GroupName')
            add_cell(sheet7, sheet7_cell_start, 18, sec_groups)
            # IAM role
            try:
                IAM = Instances['IamInstanceProfile'].get('Arn').split('/')[-1]
                add_cell(sheet7, sheet7_cell_start, 19, IAM)
            except:
                add_cell(sheet7, sheet7_cell_start, 19, '-')

            # Data Volume이 존재할 경우
            if len(data_volume_list) != 0:
                # Data Volume ID
                data_volume_id = ''
                data_volume_size = ''
                for idx, data_volume in enumerate(data_volume_list):
                    if idx != 0:
                        data_volume_id += ', \n'
                        data_volume_size += ', \n'
                    data_volume_id += data_volume
                    response2 = ec2_cli.describe_volumes(
                        VolumeIds=[
                            data_volume,
                        ],
                    )
                    for volume in response2['Volumes']:
                        data_volume_size += str(volume.get('Size'))

                add_cell(sheet7, sheet7_cell_start, 20, data_volume_id)
                add_cell(sheet7, sheet7_cell_start, 21, data_volume_size)
                # start_sheet7_row = sheet7_cell_start
                # # Data Volume ID
                # # Data Volume (GB)
                # for data_volume in data_volume_list:
                #     add_cell(sheet7, sheet7_cell_start, 20, data_volume)
                #     response2 = ec2_cli.describe_volumes(
                #         VolumeIds=[
                #             data_volume,
                #         ],
                #     )
                #     for volume in response2['Volumes']:
                #         data_size = volume.get('Size')
                #         add_cell(sheet7, sheet7_cell_start, 21, data_size)
                #     sheet7_cell_start += 1
                # sheet7_cell_start -= 1      # 마지막 Data Volume은 +1 한거 취소
            else:
                add_cell(sheet7, sheet7_cell_start, 20, '-')
                add_cell(sheet7, sheet7_cell_start, 21, '-')

            end_sheet7_row = sheet7_cell_start
            # if len(data_volume_list) != 0:
            #     for i in range(2, 20):
            #         sheet7.merge_cells(start_row = start_sheet7_row , end_row = end_sheet7_row, start_column = i, end_column = i)

        sheet7_cell_start = sheet7_cell_start + 1

    ''' EC2
    '''
else:
    print("There is no EC2.")

# sheet8
rdsres = rds_cli.describe_db_instances()

if rdsres['DBInstances']:

    sheet8 = wb.create_sheet('sheet8')
    sheet8.title = 'RDS'
    cell_widths = [5, 5, 14.25, 35, 11.67, 13.08, 22.42, 15.00, 15.50, 21, 45, 49.00, 23.83, 23.83, 12.42, 27.17, 22.42,
                   20.00, 7, 7, 7, 7]
    sheet_cell_width(sheet8, cell_widths)
    title = sheet8['B2']
    title.value = "<RDS>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' 
    RDS
    '''
    sheet8_cell_start = 3
    cell_header = ["No.", "Availability Zone", "RDS Name", "RDS Engine", "Engine Version", "DB Instance Class",
                   "Storage Type", "Master Username", "Master Password", "VPC ID", "Subnet Group", "Parameter Group",
                   "Option Group", "Database Port", "Preferred Maintenance Window", "Preferred Backup Window",
                   "Backup Retention Time"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet8, sheet8_cell_start, col_index + 2, header)
        sheet8.cell(row=sheet8_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet8.cell(row=sheet8_cell_start, column=col_index + 2).border = thin_border

    sheet8_cell_start = sheet8_cell_start + 1

    for idx, rdsdata in enumerate(rdsres['DBInstances']):
        # pprint.pprint(rdsdata)
        # No.
        add_cell(sheet8, sheet8_cell_start, 2, idx + 1)
        # Availability Zone
        add_cell(sheet8, sheet8_cell_start, 3, rdsdata['AvailabilityZone'])
        # RDS Name
        add_cell(sheet8, sheet8_cell_start, 4, rdsdata['DBInstanceIdentifier'])
        # DB Engine
        add_cell(sheet8, sheet8_cell_start, 5, rdsdata['Engine'])
        # Engine Version
        add_cell(sheet8, sheet8_cell_start, 6, rdsdata['EngineVersion'])
        # DB Instance Class
        add_cell(sheet8, sheet8_cell_start, 7, rdsdata['DBInstanceClass'])
        # Storage Type
        add_cell(sheet8, sheet8_cell_start, 8, rdsdata['StorageType'])
        # Master Username
        add_cell(sheet8, sheet8_cell_start, 9, rdsdata['MasterUsername'])
        # Master Password
        add_cell(sheet8, sheet8_cell_start, 10, '-')
        # VPC ID
        add_cell(sheet8, sheet8_cell_start, 11, rdsdata['DBSubnetGroup'].get('VpcId'))
        # Subnet Group

        subnetstr = rdsdata['DBSubnetGroup'].get('DBSubnetGroupName')
        subnetstr += '\n( '
        for idx, subnet in enumerate(rdsdata['DBSubnetGroup'].get('Subnets')):
            if idx != 0:
                subnetstr += ', \n'
            subnetstr += subnet.get('SubnetIdentifier')
        subnetstr += ' )'
        add_cell(sheet8, sheet8_cell_start, 12, subnetstr)

        # # DB Security Group
        # print(rdsdata['DBSecurityGroups'])
        # for dbsec in rdsdata['DBSecurityGroups']:
        #     if dbsec :
        #         # 수정 필요
        #         pass
        #     else:
        #         add_cell(sheet8, sheet8_cell_start, 13, '-')
        # Parameter Group
        for dbparam in rdsdata['DBParameterGroups']:
            add_cell(sheet8, sheet8_cell_start, 13, dbparam.get('DBParameterGroupName'))
        # Option Group
        for dboption in rdsdata['OptionGroupMemberships']:
            add_cell(sheet8, sheet8_cell_start, 14, dboption.get('OptionGroupName'))
        # Database Port
        add_cell(sheet8, sheet8_cell_start, 15, rdsdata['Endpoint'].get('Port'))

        # Maintenance Time
        add_cell(sheet8, sheet8_cell_start, 16, rdsdata['PreferredMaintenanceWindow'])
        # PreferredMaintenanceWindowPreferredBackupWindow Backup Retention Time
        add_cell(sheet8, sheet8_cell_start, 17, rdsdata['PreferredBackupWindow'])
        add_cell(sheet8, sheet8_cell_start, 18, rdsdata['BackupRetentionPeriod'])
        # 용도
        sheet8_cell_start = sheet8_cell_start + 1
else:
    print("There is no RDS.")

# sheet9
response = s3_cli.list_buckets()

if response['Buckets']:
    sheet9 = wb.create_sheet('sheet9')
    sheet9.title = 'S3'
    cell_widths = [5, 5, 55, 22.83, 23, 17, 12, 12, 12, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 7, 7, 7,
                   7]
    sheet_cell_width(sheet9, cell_widths)

    title = sheet9['B2']
    title.value = "<S3>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)
    s3_size_list = []

    '''
    S3
    '''
    sheet9_cell_start = 3

    # Bucket name	Access	Date created	Size
    cell_header = ["No.", "Bucket Name", "Access", "Creation Date", "Size\n(" + str(target_date) + " 측정값)", "Logging",
                   "Versioning"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet9, sheet9_cell_start, col_index + 2, header)
        sheet9.cell(row=sheet9_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                    fgColor=Color('E3E3E3'))
        sheet9.cell(row=sheet9_cell_start, column=col_index + 2).border = thin_border

    sheet9_cell_start = sheet9_cell_start + 1

    for idx, each_bucket in enumerate(s3_res.buckets.all()):
        # No.
        add_cell(sheet9, sheet9_cell_start, 2, idx + 1)

        # Bucket name
        bucket_name = each_bucket.name
        add_cell(sheet9, sheet9_cell_start, 3, bucket_name)

        # Access
        ''' 만들어야 함
            response = s3_cli.get_public_access_block(Bucket='test-pingping95-bucket')
            response['PublicAccessBlockConfiguration']['BlockPublicAcls']
            response['PublicAccessBlockConfiguration']['BlockPublicPolicy']
        '''
        add_cell(sheet9, sheet9_cell_start, 4, '')

        # Date created
        bucket_created_date = str(each_bucket.creation_date)
        bucket_date = bucket_created_date.split(' ')[0]
        bucket_time = bucket_created_date.split(' ')[1]
        add_cell(sheet9, sheet9_cell_start, 5, bucket_date + ',  ' + bucket_time)

        # Bucket Size
        readable_bucket_size = 0
        response = cloudwatch_cli.get_metric_statistics(Namespace='AWS/S3',
                                                        MetricName='BucketSizeBytes',
                                                        Dimensions=[
                                                            {'Name': 'BucketName', 'Value': bucket_name},
                                                            {'Name': 'StorageType', 'Value': 'StandardStorage'}
                                                        ],
                                                        Statistics=['Average'],
                                                        Period=3600,
                                                        StartTime=(now - datetime.timedelta(days=2)).isoformat(),
                                                        EndTime=now.isoformat()
                                                        )
        if len(response['Datapoints']) != 0:
            bucket_size = int(response['Datapoints'][0].get('Average'))
        else:
            bucket_size = '-'

        if type(bucket_size) == int:
            readable_bucket_size = humanbytes(bucket_size)

        add_cell(sheet9, sheet9_cell_start, 6, readable_bucket_size)

        try:
            # Logging
            bucketLogging = s3_cli.get_bucket_logging(Bucket=bucket_name)
            if 'LoggingEnabled' in bucketLogging:
                add_cell(sheet9, sheet9_cell_start, 7, "Enabled")
            else:
                add_cell(sheet9, sheet9_cell_start, 7, "Disabled")

            # Versioning
            bucketVersioning = s3_cli.get_bucket_versioning(Bucket=bucket_name)
            if 'Status' in bucketVersioning:
                add_cell(sheet9, sheet9_cell_start, 8, "Enabled")
            else:
                add_cell(sheet9, sheet9_cell_start, 8, "Disabled")
        except Exception as e:
            add_cell(sheet9, sheet9_cell_start, 7, "-")
            add_cell(sheet9, sheet9_cell_start, 8, "-")
            print(f"{bucket_name} : 비 정상적인 Bucket으로 추정")

        sheet9_cell_start = sheet9_cell_start + 1
else:
    print("There is no S3.")

# sheet10
cloudfront = cloudfront_cli.list_distributions()['DistributionList']
try:
    test = cloudfront['Items']
    sheet10 = wb.create_sheet('sheet10')
    sheet10.title = 'CloudFront'
    cell_widths = [6, 6, 18.92, 33, 35, 27, 12, 12, 8.08, 8.08, 8.08, 8.08, 8.08, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7,
                   7]
    sheet_cell_width(sheet10, cell_widths)

    title = sheet10['B2']
    title.value = "<CloudFront>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)
    # title.alignment = Alignment(horizontal='center', vertical='center')

    ''' 
    CloudFront
    '''
    sheet10_cell_start = 3
    cell_header = ["No.", "ID", "Domain Name", "Origin Domain Name", "CNAMEs", "State"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet10, sheet10_cell_start, col_index + 2, header)
        sheet10.cell(row=sheet10_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                      fgColor=Color('E3E3E3'))
        sheet10.cell(row=sheet10_cell_start, column=col_index + 2).border = thin_border

    sheet10_cell_start = sheet10_cell_start + 1

    for idx, items in enumerate(cloudfront['Items']):
        # No.
        add_cell(sheet10, sheet10_cell_start, 2, idx + 1)
        # Id
        add_cell(sheet10, sheet10_cell_start, 3, items.get('Id'))
        # Domain Name
        add_cell(sheet10, sheet10_cell_start, 4, items.get('DomainName'))

        # Origin Domain Name
        originstr = ''
        origin = items.get('Origins')
        for idx, item in enumerate(origin['Items']):
            if idx != 0:
                originstr += ', \n'
            originstr += item.get('DomainName')
        add_cell(sheet10, sheet10_cell_start, 5, originstr)

        # CNAMEs
        cnameItems = ''
        cname = items.get('Aliases')
        for idx, item in enumerate(cname['Items']):
            if idx != 0:
                cnameItems += ', \n'
            cnameItems += item
        add_cell(sheet10, sheet10_cell_start, 6, cnameItems)

        # State ( Enabled, Disabled )
        state = items['Enabled']
        if state == True:
            state = 'Enabled'
        elif state == False:
            state = 'Disabled'
        add_cell(sheet10, sheet10_cell_start, 7, state)

        sheet10_cell_start += 1

    ''' CloudFront
    '''
except:
    print("There is no CloudFront.")

# sheet11

if cloudtrail_cli.describe_trails()['trailList']:
    # print(cloudtrail_cli.describe_trails()['trailList'])

    sheet11 = wb.create_sheet('sheet11')
    sheet11.title = 'CloudTrail'
    cell_widths = [6, 6, 18, 14, 16, 13.00, 15, 45, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7,
                   7]
    sheet_cell_width(sheet11, cell_widths)

    title = sheet11['B2']
    title.value = "<CloudTrail>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' CloudTrail
    '''
    sheet11_cell_start = 3

    cell_header = ["No.", "Trail name", "Home Region", "Multi-region Trail", "Trail Insights", "Organization Trail",
                   "S3 bucket"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet11, sheet11_cell_start, col_index + 2, header)
        sheet11.cell(row=sheet11_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                      fgColor=Color('E3E3E3'))
        sheet11.cell(row=sheet11_cell_start, column=col_index + 2).border = thin_border

    sheet11_cell_start = sheet11_cell_start + 1

    # NO.
    # Home Region
    # Trail name
    # Multi-region trails
    # trail	Insights
    # Organization trail	
    # S3 bucket
    for idx, trail in enumerate(cloudtrail_cli.describe_trails()['trailList']):
        add_cell(sheet11, sheet11_cell_start, 2, idx + 1)
        add_cell(sheet11, sheet11_cell_start, 3, trail.get('Name'))
        add_cell(sheet11, sheet11_cell_start, 4, trail.get('HomeRegion'))
        add_cell(sheet11, sheet11_cell_start, 5, trail.get('IsMultiRegionTrail'))
        add_cell(sheet11, sheet11_cell_start, 6, trail.get('HasInsightSelectors'))
        add_cell(sheet11, sheet11_cell_start, 7, trail.get('IsOrganizationTrail'))
        add_cell(sheet11, sheet11_cell_start, 8, trail.get('S3BucketName'))
        sheet11_cell_start = sheet11_cell_start + 1

    ''' CloudTrail
    '''
else:
    print("There is no CloudTrail.")

# sheet12

sheet12 = wb.create_sheet('sheet12')
sheet12.title = 'CloudWatch'
cell_widths = [6, 6, 56, 56, 15, 23, 28, 10.75, 19, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 7, 7, 7, 7]
sheet_cell_width(sheet12, cell_widths)

title = sheet12['B2']
title.value = "<CloudWatch>"
title.font = Font(name='맑은 고딕', size=12, bold=True)

''' CloudWatch
'''
sheet12_cell_start = 3

cell_header = ["No.", "Dashboard Name", "Last Updated(UTC)", "Size"]

for col_index, header in enumerate(cell_header):
    add_cell(sheet12, sheet12_cell_start, col_index + 2, header)
    sheet12.cell(row=sheet12_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                  fgColor=Color('E3E3E3'))
    sheet12.cell(row=sheet12_cell_start, column=col_index + 2).border = thin_border

sheet12_cell_start = sheet12_cell_start + 1

dashboards = cloudwatch_cli.list_dashboards()
for idx, dashboard in enumerate(dashboards['DashboardEntries']):
    add_cell(sheet12, sheet12_cell_start, 2, idx + 1)
    add_cell(sheet12, sheet12_cell_start, 3, dashboard.get('DashboardName'))
    date = str(dashboard.get('LastModified'))
    add_cell(sheet12, sheet12_cell_start, 4, date.split('+')[0])
    add_cell(sheet12, sheet12_cell_start, 5, dashboard.get('Size'))
    sheet12_cell_start = sheet12_cell_start + 1

sheet12_cell_start = sheet12_cell_start + 1
cell_header = ["No.", "Alarm Name", "AlarmDescription", "Namespace", "MetricName", "ComparisonOperator", "Threshold",
               "StateValue"]

for col_index, header in enumerate(cell_header):
    add_cell(sheet12, sheet12_cell_start, col_index + 2, header)
    sheet12.cell(row=sheet12_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                  fgColor=Color('E3E3E3'))
    sheet12.cell(row=sheet12_cell_start, column=col_index + 2).border = thin_border

sheet12_cell_start = sheet12_cell_start + 1

alarms = cloudwatch_cli.describe_alarms()
for idx, alarm in enumerate(alarms['MetricAlarms']):
    add_cell(sheet12, sheet12_cell_start, 2, idx + 1)
    add_cell(sheet12, sheet12_cell_start, 3, alarm.get('AlarmName'))
    add_cell(sheet12, sheet12_cell_start, 4, alarm.get('AlarmDescription'))
    add_cell(sheet12, sheet12_cell_start, 5, alarm.get('Namespace'))
    add_cell(sheet12, sheet12_cell_start, 6, alarm.get('MetricName'))
    add_cell(sheet12, sheet12_cell_start, 7, alarm.get('ComparisonOperator'))
    add_cell(sheet12, sheet12_cell_start, 8, alarm.get('Threshold'))
    add_cell(sheet12, sheet12_cell_start, 9, alarm.get('StateValue'))
    sheet12_cell_start = sheet12_cell_start + 1

''' CloudWatch
'''

# sheet13

sheet13 = wb.create_sheet('sheet13')
sheet13.title = 'IAM User'
cell_widths = [6, 6, 30, 30, 38, 33, 25, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7]
sheet_cell_width(sheet13, cell_widths)

title = sheet13['B2']
title.value = "<IAM Users>"
title.font = Font(name='맑은 고딕', size=12, bold=True)

''' IAM Users
'''
sheet13_cell_start = 3

cell_header = ["No.", "User Name", "Group Names", "Policies\n(Attached to Groups)", "Policies\n(Attached to User)",
               "Creation Date"]

for col_index, header in enumerate(cell_header):
    add_cell(sheet13, sheet13_cell_start, col_index + 2, header)
    sheet13.cell(row=sheet13_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                  fgColor=Color('E3E3E3'))
    sheet13.cell(row=sheet13_cell_start, column=col_index + 2).border = thin_border

sheet13_cell_start = sheet13_cell_start + 1

for idx, user_detail in enumerate(iam_cli.get_account_authorization_details(Filter=['User'])['UserDetailList']):
    add_cell(sheet13, sheet13_cell_start, 2, idx + 1)
    add_cell(sheet13, sheet13_cell_start, 3, str(user_detail.get('UserName')))

    if len(user_detail.get('GroupList')) != 0:
        # Group 리스트 생성
        group_list = []
        for idx, group in enumerate(user_detail.get('GroupList')):
            group_list.append(group)

        str_group_list = ', \n'.join(group_list)
        add_cell(sheet13, sheet13_cell_start, 4, str_group_list)

        # Group Policies
        group_policies = []
        for each_group in group_list:
            iam_res_group = iam_res.Group(each_group)
            policy_generator = iam_res_group.attached_policies.all()
            for policy in policy_generator:
                group_policies.append(policy.policy_name)

        if len(group_policies) != 0:
            str_group_policies = ', \n'.join(group_policies)
            add_cell(sheet13, sheet13_cell_start, 5, str_group_policies)
        else:
            add_cell(sheet13, sheet13_cell_start, 5, '-')

        # User Policies
        user_policies = []
        for policy in user_detail.get('AttachedManagedPolicies'):
            user_policies.append(str(policy['PolicyName']))

        only_user_policies = [x for x in user_policies if x not in group_policies]
        if len(only_user_policies) != 0:
            str_only_user_policies = ', \n'.join(only_user_policies)
            add_cell(sheet13, sheet13_cell_start, 6, str_only_user_policies)
        else:
            add_cell(sheet13, sheet13_cell_start, 6, '-')

    else:
        add_cell(sheet13, sheet13_cell_start, 4, '-')
        add_cell(sheet13, sheet13_cell_start, 5, '-')

        gpolicy = ""
        for idx, attachpolicy in enumerate(user_detail.get('AttachedManagedPolicies')):
            if idx != 0:
                gpolicy += ', \n'
            gpolicy += attachpolicy['PolicyName']
        add_cell(sheet13, sheet13_cell_start, 5, gpolicy)
        add_cell(sheet13, sheet13_cell_start, 6, gpolicy)

    # Creation Date
    created_date = str(user_detail.get('CreateDate'))
    iamUser_date = created_date.split(' ')[0]
    iamUser_time = created_date.split(' ')[1]
    add_cell(sheet13, sheet13_cell_start, 7, iamUser_date + ',  ' + iamUser_time)

    sheet13_cell_start = sheet13_cell_start + 1
''' IAM Users
'''

# sheet14
iam_roles = iam_cli.list_roles()['Roles']
iam_roles_cnt = 0
for i in iam_roles:
    if i['Path'] == '/':
        iam_roles_cnt += 1

# Custom하게 생성한 IAM Roles가 1개 이상일 경우
if iam_roles_cnt != 0:
    sheet14 = wb.create_sheet('sheet14')
    sheet14.title = 'IAM Role'
    cell_widths = [6, 6, 45, 70, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7, 7]
    sheet_cell_width(sheet14, cell_widths)

    title = sheet14['B2']
    title.value = "<IAM Roles>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' IAM Roles
    '''

    sheet14_cell_start = 3
    cell_header = ["No.", "Role Name", "Description"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet14, sheet14_cell_start, col_index + 2, header)
        sheet14.cell(row=sheet14_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet14.cell(row=sheet14_cell_start, column=col_index + 2).border = thin_border

    sheet14_cell_start = sheet14_cell_start + 1


    count = 1
    for idx, iam_role in enumerate(iam_roles):
        # 직접 생성한 IAM Roles만 엑셀에 추가함
        if iam_role['Path'] == '/':
            # No.
            add_cell(sheet14, sheet14_cell_start, 2, count)

            # Role Name
            roleName = iam_role['RoleName']
            add_cell(sheet14, sheet14_cell_start, 3, roleName)

            # Description
            try:
                iamRoleDesc = iam_role['Description']
                add_cell(sheet14, sheet14_cell_start, 4, iamRoleDesc)
            except Exception:
                add_cell(sheet14, sheet14_cell_start, 4, '-')

            sheet14_cell_start = sheet14_cell_start + 1
            count += 1
else:
    print("There is no self made IAM Roles")

# sheet15

if lambda_cli.list_functions().get('Functions'):
    sheet15 = wb.create_sheet('sheet15')
    sheet15.title = 'Lambda'
    cell_widths = [6, 6, 30, 12, 55, 11, 12, 9, 12, 8.08, 8.08, 8.08, 8.08, 8.08, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7]
    sheet_cell_width(sheet15, cell_widths)

    title = sheet15['B2']
    title.value = "<Lambda>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)
    title.alignment = Alignment(horizontal='center', vertical='center')

    ''' Lambda
    '''
    sheet15_cell_start = 3

    cell_header = ["No.", "Function Name", "Runtime", "Role", "Code Size", "Memory Size", "Time Out", "Package Type"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet15, sheet15_cell_start, col_index + 2, header)
        sheet15.cell(row=sheet15_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                      fgColor=Color('E3E3E3'))
        sheet15.cell(row=sheet15_cell_start, column=col_index + 2).border = thin_border

    sheet15_cell_start = sheet15_cell_start + 1

    for idx, each_lambda in enumerate(lambda_cli.list_functions().get('Functions')):
        add_cell(sheet15, sheet15_cell_start, 2, idx + 1)
        # Function Name
        add_cell(sheet15, sheet15_cell_start, 3, each_lambda.get('FunctionName'))
        # Runtime
        add_cell(sheet15, sheet15_cell_start, 4, each_lambda.get('Runtime'))
        # Role Name
        roleName = each_lambda.get('Role').split(':')[-1]
        add_cell(sheet15, sheet15_cell_start, 5, roleName)
        # Code Size
        readable_code_size = humanbytes(each_lambda.get('CodeSize'))
        add_cell(sheet15, sheet15_cell_start, 6, readable_code_size)
        # Memory Size
        add_cell(sheet15, sheet15_cell_start, 7, each_lambda.get('MemorySize'))
        # Timeout
        add_cell(sheet15, sheet15_cell_start, 8, each_lambda.get('Timeout'))
        # Package Type
        add_cell(sheet15, sheet15_cell_start, 9, each_lambda.get('PackageType'))
        sheet15_cell_start = sheet15_cell_start + 1
    ''' Lambda
    '''

else:
    print("There is no Lambda.")

# Sheet16

if route53_cli.list_hosted_zones_by_name()['HostedZones']:
    sheet16 = wb.create_sheet('sheet16')
    sheet16.title = 'Route 53'
    cell_widths = [6, 6, 33, 9, 55, 10, 75, 8, 8, 8.08, 8.08, 8.08, 8.08, 8.08, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7]
    sheet_cell_width(sheet16, cell_widths)

    title = sheet16['B2']
    title.value = "<Route 53>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' Route 53
    '''
    # No, Domain Name, Type, Record Name, Type, Value
    sheet16_cell_start = 3

    cell_header = ["No.", "Domain Name", "Type", "Record Name", "Record Type", "Record Value"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet16, sheet16_cell_start, col_index + 2, header)
        sheet16.cell(row=sheet16_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                      fgColor=Color('E3E3E3'))
        sheet16.cell(row=sheet16_cell_start, column=col_index + 2).border = thin_border

    sheet16_cell_start = sheet16_cell_start + 1

    hosted_zones = route53_cli.list_hosted_zones_by_name()['HostedZones']
    # record_sets = route53_cli.list_resource_record_sets
    for idx, zone in enumerate(hosted_zones):
        zoneId = zone['Id']
        # No.
        add_cell(sheet16, sheet16_cell_start, 2, idx + 1)
        # Domain Name
        zoneName = zone['Name']
        real_zoneName = zoneName[:-1]
        add_cell(sheet16, sheet16_cell_start, 3, real_zoneName)
        # Type
        if zone['Config']['PrivateZone'] == False:
            add_cell(sheet16, sheet16_cell_start, 4, "Public")
        else:
            add_cell(sheet16, sheet16_cell_start, 4, "Private")

        # Record Sets
        init_row_cnt = sheet16_cell_start
        row_cnt = 0
        for record_set in route53_cli.list_resource_record_sets(HostedZoneId=zoneId)['ResourceRecordSets']:
            row_cnt += 1
            # Record Name
            record_name = record_set.get('Name')
            # Record Type
            record_type = record_set.get('Type')
            # Record Values
            # Alias : AliasTarget
            # Value : ResourceRecords
            record_value = ''
            if 'ResourceRecords' in record_set:
                for idx, value in enumerate(record_set.get('ResourceRecords')):
                    if idx != 0:
                        record_value = record_value + ', \n'
                    record_value += value.get('Value')
            elif 'AliasTarget' in record_set:
                record_value = record_set['AliasTarget']['DNSName']

            else:
                record_value = '-'

            add_cell(sheet16, sheet16_cell_start, 5, record_name)
            add_cell(sheet16, sheet16_cell_start, 6, record_type)
            add_cell(sheet16, sheet16_cell_start, 7, record_value)

            sheet16_cell_start = sheet16_cell_start + 1
        # print(f"row_cnt : {row_cnt}")
        # print(f"init_row_cnt : {init_row_cnt}")
        sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=2, end_column=2)
        sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=3, end_column=3)
        sheet16.merge_cells(start_row=init_row_cnt, end_row=init_row_cnt + row_cnt - 1, start_column=4, end_column=4)

        # sheet16_cell_start = sheet16_cell_start+1

else:
    print("There is no Route53")

# Sheet17

''' ElastiCache
'''

# Sheet18

''' EKS & ECR
'''

# EKS가 있을 경우
if eks_cli.list_clusters()['clusters']:
    response = eks_cli.list_clusters()['clusters']
    sheet18 = wb.create_sheet('sheet18')
    sheet18.title = 'EKS'
    cell_widths = [6, 6, 18, 13, 14, 8, 46, 23, 21, 22, 22, 8, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7]
    sheet_cell_width(sheet18, cell_widths)

    title = sheet18['B2']
    title.value = "<EKS>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    ''' EKS
    '''
    sheet18_cell_start = 3

    cell_header = ["No.", "Cluster Name", "Cluster Version", "Platform Version", "Status", "Subnet IDs",
                   "Cluster Security Group IDs", "Security Group IDs", "Node Group Names", "Fargate Profile Names"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet18, sheet18_cell_start, col_index + 2, header)
        sheet18.cell(row=sheet18_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid',
                                                                                      fgColor=Color('E3E3E3'))
        sheet18.cell(row=sheet18_cell_start, column=col_index + 2).border = thin_border

    sheet18_cell_start = sheet18_cell_start + 1

    for idx, cluster_name in enumerate(response):
        cluster = eks_cli.describe_cluster(name=cluster_name)['cluster']

        # No.
        add_cell(sheet18, sheet18_cell_start, 2, idx + 1)

        # Cluster name
        name = cluster.get('name')
        add_cell(sheet18, sheet18_cell_start, 3, name)

        # Version
        version = cluster.get('version')
        add_cell(sheet18, sheet18_cell_start, 4, version)

        # Platform Version
        platform_ver = cluster.get('platformVersion')
        add_cell(sheet18, sheet18_cell_start, 5, platform_ver)


        # Status
        status = cluster.get('status')
        add_cell(sheet18, sheet18_cell_start, 6, status)

        # Subnet IDs
        subnetIds = ''
        for i, subnetId in enumerate(cluster['resourcesVpcConfig']['subnetIds']):
            if i != 0:
                subnetIds += ', '
            elif i != 0 and i % 2 == 0:
                subnetIds += '\n'
            subnetIds += subnetId
        add_cell(sheet18, sheet18_cell_start, 7, subnetIds)

        # Cluster Security Group IDs
        cluster_sg_ids = cluster['resourcesVpcConfig'].get('clusterSecurityGroupId')
        if cluster_sg_ids:
            add_cell(sheet18, sheet18_cell_start, 8, cluster_sg_ids)
        else:
            add_cell(sheet18, sheet18_cell_start, 8, '-')

        # Security Group IDs
        sg_ids_list = cluster['resourcesVpcConfig'].get('securityGroupIds')
        sg_ids = ''
        for i, sg_id in enumerate(sg_ids_list):
            if i != 0:
                sg_ids += ', \n'
            sg_ids += sg_id
        add_cell(sheet18, sheet18_cell_start, 9, sg_ids)

        # Node Group Names
        node_groups = eks_cli.list_nodegroups(clusterName=name)['nodegroups']
        if len(node_groups) != 0:
            node_group_names = ''
            for i, node_group in enumerate(node_groups):
                if i != 0:
                    node_group_names += ', \n'
                node_group_names += node_group
            add_cell(sheet18, sheet18_cell_start, 10, node_group_names)
        else:
            add_cell(sheet18, sheet18_cell_start, 10, '-')

        # Fargate Profile Names
        fargates = eks_cli.list_fargate_profiles(clusterName='eks-test')['fargateProfileNames']
        if len(fargates) != 0:
            fg_names = ''
            for i, fargate in enumerate(fargates):
                if i != 0:
                    fg_names += ', \n'
                fg_names += fargate
            add_cell(sheet18, sheet18_cell_start, 11, fg_names)
        else:
            add_cell(sheet18, sheet18_cell_start, 11, '-')

else:
    print("There is no EKS")

# sheet19
response = els_cli.describe_cache_clusters()['CacheClusters']
if len(response) != 0:

    sheet19 = wb.create_sheet('sheet19')
    sheet19.title = 'ElastiCache'
    cell_widths = [6, 6, 18, 13, 14, 8, 46, 23, 21, 22, 22, 8, 8.10, 8.10, 8.10, 8.10, 8.10, 8.10, 7, 7, 7, 7]
    sheet_cell_width(sheet19, cell_widths)

    title = sheet19['B2']
    title.value = "<ElastiCache>"
    title.font = Font(name='맑은 고딕', size=12, bold=True)

    '''
    ElastiCache
    '''
    sheet19_cell_start = 3
    cell_header = ["No.", "Cluster Name"]

    for col_index, header in enumerate(cell_header):
        add_cell(sheet19, sheet19_cell_start, col_index + 2, header)
        sheet19.cell(row=sheet19_cell_start, column=col_index + 2).fill = PatternFill(patternType='solid', fgColor=Color('E3E3E3'))
        sheet19.cell(row=sheet19_cell_start, column=col_index + 2).border = thin_border

    sheet19_cell_start = sheet19_cell_start + 1
    '''
    ElastiCache
    '''

else:
    print('There is no ElastiCache')


wb.save(p_name + '_asset.xlsx')
wb.close()